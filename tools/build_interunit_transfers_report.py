# -*- coding: utf-8 -*-
"""Consolidated inter-unit transfer report: Transfer-Out / Cold Transfer-Out /
Transfer-In / Cold Transfer-In, all four sheets in the SAME item-level layout.

    Challan | Date | From | To | Item | Category | Material | Lot |
    Qty | Net Weight | Total Weight | Boxes | Status | Received

READ-ONLY against prod. Usage:
    python tools/build_interunit_transfers_report.py [YYYY-MM-DD cutoff] [out.xlsx]
"""
import os
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

CUTOFF = date.fromisoformat(sys.argv[1]) if len(sys.argv) > 1 else date(2026, 7, 26)
OUT_XLSX = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "data" / f"interunit_transfers_report_till_{CUTOFF:%d%b%Y}.xlsx"
CUT_TS = CUTOFF + timedelta(days=1)   # exclusive upper bound on timestamps

_raw = os.environ["DATABASE_URL"]
engine = create_engine(_raw.replace("postgresql://", "postgresql+psycopg://", 1)
                       if _raw.startswith("postgresql://") else _raw)

# Cold sites, lower-cased. A dispatch FROM one of these (or from 'Cold Storage')
# is a Cold Transfer-Out; a receipt INTO one is a Cold Transfer-In.
COLD_SITES = {"savla d-39", "savla d-514", "rishi", "supreme", "eskimo"}
COLD_SQL = "('savla d-39','savla d-514','rishi','supreme','eskimo')"

COLS = ["Challan", "Date", "From", "To", "Item", "Category", "Material", "Lot",
        "Qty", "Net Weight", "Total Weight", "Boxes", "Status", "Received"]
IN_EXTRA = ["GRN No", "Out Date", "Received By", "Condition", "Box Source", "Scan Rows"]


def fetch(sql, **params):
    with engine.connect() as c:
        c.execute(text("SET TRANSACTION READ ONLY"))
        return [dict(r._mapping) for r in c.execute(text(sql), params)]


def n3(x):
    return None if x is None else round(float(x), 3)


def is_cold(site):
    s = (site or "").strip().lower()
    return s.startswith("cold") or s in COLD_SITES


def received_state(status, has_grn):
    if status == "Received":
        return "Received"
    return "Pending" if has_grn else "Not Received"


# ── 1. Transfer-Out / Cold Transfer-Out ─────────────────────────────────
# Grain = (challan, item, category, material, lot). Cold dispatches write one
# LINE per carton, so Boxes = COUNT(lines) — same measure the previous report used.
OUT_SQL = f"""
SELECT h.id, h.challan_no, h.stock_trf_date::text AS doc_date, h.from_site, h.to_site,
       h.status, h.from_cold_unit,
       COALESCE(l.item_desc_raw,'') AS item, COALESCE(l.item_category,'') AS category,
       COALESCE(l.rm_pm_fg_type,'') AS material, COALESCE(l.lot_number,'') AS lot,
       SUM(l.qty) AS qty, SUM(l.net_weight) AS net_wt, SUM(l.total_weight) AS tot_wt,
       COUNT(*) AS boxes,
       (EXISTS (SELECT 1 FROM interunit_transfer_in_header i WHERE i.transfer_out_id = h.id)
        OR EXISTS (SELECT 1 FROM cold_transfer_in_headers c WHERE c.transfer_out_id = h.id)) AS has_grn
FROM interunit_transfers_lines l
JOIN interunit_transfers_header h ON h.id = l.header_id
WHERE h.stock_trf_date <= :cut OR h.created_ts < :cut_ts
GROUP BY h.id, h.challan_no, h.stock_trf_date, h.from_site, h.to_site, h.status,
         h.from_cold_unit, l.item_desc_raw, l.item_category, l.rm_pm_fg_type, l.lot_number
ORDER BY h.stock_trf_date, h.challan_no, l.item_desc_raw, l.lot_number
"""


def build_out():
    wh, cold = [], []
    for r in fetch(OUT_SQL, cut=CUTOFF, cut_ts=CUT_TS):
        row = [r["challan_no"], r["doc_date"], r["from_site"], r["to_site"], r["item"],
               r["category"], r["material"], r["lot"], n3(r["qty"]), n3(r["net_wt"]),
               n3(r["tot_wt"]), r["boxes"], r["status"],
               received_state(r["status"], r["has_grn"])]
        (cold if (is_cold(r["from_site"]) or r["from_cold_unit"]) else wh).append(row)
    return wh, cold


# ── 2. Category / Material lookup from the matching transfer-OUT lines ───
LINE_MAP_SQL = """
SELECT header_id, LOWER(TRIM(COALESCE(item_desc_raw,''))) AS item,
       COALESCE(lot_number,'') AS lot,
       MIN(COALESCE(item_category,'')) AS category, MIN(COALESCE(rm_pm_fg_type,'')) AS material
FROM interunit_transfers_lines GROUP BY 1,2,3
"""


def line_lookup():
    by_item_lot, by_item, by_lot = {}, {}, {}
    for r in fetch(LINE_MAP_SQL):
        v = (r["category"], r["material"])
        by_item_lot[(r["header_id"], r["item"], r["lot"])] = v
        by_item.setdefault((r["header_id"], r["item"]), v)
        if r["lot"]:
            by_lot.setdefault((r["header_id"], r["lot"]), v)

    def look(out_id, item, lot):
        it = (item or "").strip().lower()
        return (by_item_lot.get((out_id, it, lot or ""))
                or by_item.get((out_id, it))
                or by_lot.get((out_id, lot))
                or ("", ""))
    return look


# ── 3. Transfer-In (warehouse destinations) ─────────────────────────────
IN_SQL = f"""
SELECT i.id, i.grn_number, i.grn_date::date::text AS grn_date, i.status, i.received_by,
       i.box_condition, i.transfer_out_id,
       COALESCE(o.challan_no, i.transfer_out_no, '') AS challan,
       o.stock_trf_date::text AS out_date, o.status AS out_status, o.from_site,
       COALESCE(NULLIF(TRIM(i.receiving_warehouse),''), o.to_site) AS to_site,
       COALESCE(b.article,'') AS item, COALESCE(b.lot_number,'') AS lot,
       COUNT(b.id) AS boxes, SUM(b.net_weight) AS net_wt, SUM(b.gross_weight) AS tot_wt
FROM interunit_transfer_in_header i
LEFT JOIN interunit_transfer_in_boxes b ON b.header_id = i.id
LEFT JOIN interunit_transfers_header o ON o.id = i.transfer_out_id
WHERE i.grn_date < :cut_ts
  AND LOWER(TRIM(COALESCE(i.receiving_warehouse,''))) NOT IN {COLD_SQL}
  AND LOWER(TRIM(COALESCE(o.to_site,''))) NOT IN {COLD_SQL}
GROUP BY i.id, i.grn_number, i.grn_date, i.status, i.received_by, i.box_condition,
         i.transfer_out_id, o.challan_no, i.transfer_out_no, o.stock_trf_date, o.status,
         o.from_site, i.receiving_warehouse, o.to_site, b.article, b.lot_number
ORDER BY i.grn_date, i.grn_number, b.article, b.lot_number
"""


def build_in(look):
    rows = []
    for r in fetch(IN_SQL, cut_ts=CUT_TS):
        cat, mat = look(r["transfer_out_id"], r["item"], r["lot"])
        rows.append([r["challan"], r["grn_date"], r["from_site"], r["to_site"], r["item"],
                     cat, mat, r["lot"], r["boxes"], n3(r["net_wt"]), n3(r["tot_wt"]),
                     r["boxes"], r["status"],
                     received_state(r["out_status"], True),
                     r["grn_number"], r["out_date"], r["received_by"], r["box_condition"],
                     "interunit_transfer_in_boxes", r["boxes"]])
    return rows


# ── 4. Cold Transfer-In (cold destinations) ─────────────────────────────
# Two receive paths write two tables. Take the per-box cold table when it holds
# real per-box detail (>1 row); else fall back to the interunit scan ledger.
COLD_IN_HDR_SQL = f"""
SELECT COALESCE(i.id, c.id) AS id,
       COALESCE(c.grn_number, i.grn_number) AS grn_number,
       COALESCE(c.grn_date, i.grn_date)::date::text AS grn_date,
       COALESCE(c.status, i.status) AS status,
       COALESCE(c.received_by, i.received_by) AS received_by,
       COALESCE(c.box_condition, i.box_condition) AS box_condition,
       COALESCE(i.transfer_out_id, c.transfer_out_id) AS transfer_out_id,
       COALESCE(o.challan_no, c.transfer_out_no, i.transfer_out_no, '') AS challan,
       o.stock_trf_date::text AS out_date, o.status AS out_status,
       COALESCE(c.from_site, o.from_site) AS from_site,
       COALESCE(NULLIF(TRIM(c.to_site),''), NULLIF(TRIM(i.receiving_warehouse),''), o.to_site) AS to_site
FROM interunit_transfer_in_header i
FULL OUTER JOIN cold_transfer_in_headers c ON c.id = i.id
LEFT JOIN interunit_transfers_header o ON o.id = COALESCE(i.transfer_out_id, c.transfer_out_id)
WHERE COALESCE(c.grn_date, i.grn_date) < :cut_ts
  AND (c.id IS NOT NULL
       OR LOWER(TRIM(COALESCE(i.receiving_warehouse,''))) IN {COLD_SQL}
       OR LOWER(TRIM(COALESCE(o.to_site,''))) IN {COLD_SQL})
ORDER BY 3, 2
"""

COLD_IN_BOX_SQL = """
SELECT header_id, COALESCE(item_description,'') AS item, COALESCE(lot_no,'') AS lot,
       COUNT(*) AS scan_rows, SUM(COALESCE(no_of_cartons,1)) AS cartons, SUM(weight_kg) AS net_wt
FROM cold_transfer_inboxes WHERE header_id = ANY(:ids)
GROUP BY 1,2,3 ORDER BY 1,2,3
"""

IU_IN_BOX_SQL = """
SELECT header_id, COALESCE(article,'') AS item, COALESCE(lot_number,'') AS lot,
       COUNT(*) AS scan_rows, SUM(net_weight) AS net_wt, SUM(gross_weight) AS tot_wt
FROM interunit_transfer_in_boxes WHERE header_id = ANY(:ids)
GROUP BY 1,2,3 ORDER BY 1,2,3
"""


def build_cold_in(look):
    hdrs = fetch(COLD_IN_HDR_SQL, cut_ts=CUT_TS)
    ids = [h["id"] for h in hdrs]
    cold_boxes, iu_boxes = defaultdict(list), defaultdict(list)
    for r in fetch(COLD_IN_BOX_SQL, ids=ids):
        cold_boxes[r["header_id"]].append(r)
    for r in fetch(IU_IN_BOX_SQL, ids=ids):
        iu_boxes[r["header_id"]].append(r)

    rows = []
    for h in hdrs:
        cb, ib = cold_boxes.get(h["id"], []), iu_boxes.get(h["id"], [])
        use_cold = len(cb) > 1 or (cb and not ib)
        src, groups = ("cold_transfer_inboxes", cb) if use_cold else ("interunit_transfer_in_boxes", ib)
        if not groups:
            src, groups = ("(no box rows)", [{"item": "", "lot": "", "scan_rows": 0,
                                              "cartons": 0, "net_wt": None, "tot_wt": None}])
        for g in groups:
            cat, mat = look(h["transfer_out_id"], g["item"], g["lot"])
            # cold table counts cartons (1 carton = 1 box); the scan ledger counts rows
            boxes = int(g["cartons"]) if use_cold else g["scan_rows"]
            rows.append([h["challan"], h["grn_date"], h["from_site"], h["to_site"], g["item"],
                         cat, mat, g["lot"], boxes, n3(g["net_wt"]),
                         n3(g.get("tot_wt", g["net_wt"])), boxes, h["status"],
                         received_state(h["out_status"], True),
                         h["grn_number"], h["out_date"], h["received_by"], h["box_condition"],
                         src, g["scan_rows"]])
    return rows


# ── 5. Write ────────────────────────────────────────────────────────────
def add_sheet(wb, title, header, rows):
    ws = wb.create_sheet(title)
    ws.append(header)
    for r in rows:
        ws.append(r)
    for c in range(1, len(header) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
        ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
        width = max([len(str(header[c - 1]))] +
                    [len(str(r[c - 1])) for r in rows[:400] if r[c - 1] is not None]) + 2
        ws.column_dimensions[get_column_letter(c)].width = min(max(width, 9), 48)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws


def main():
    look = line_lookup()
    wh_out, cold_out = build_out()
    wh_in = build_in(look)
    cold_in = build_cold_in(look)

    wb = Workbook()
    wb.remove(wb.active)
    sheets = [("Transfer Out", COLS, wh_out), ("Cold Transfer Out", COLS, cold_out),
              ("Transfer In", COLS + IN_EXTRA, wh_in),
              ("Cold Transfer In", COLS + IN_EXTRA, cold_in)]
    for title, header, rows in sheets:
        add_sheet(wb, title, header, rows)

    notes = wb.create_sheet("Notes")
    notes.append(["Inter-unit transfer report", f"IMS inception -> {CUTOFF:%d %b %Y}"])
    notes.append([])
    notes.append(["Sheet", "Rows", "Challans / GRNs", "First date", "Last date", "Total boxes"])
    for title, _h, rows in sheets:
        keys = {r[0] for r in rows} if title.endswith("Out") else {r[14] for r in rows}
        dates = sorted(d for d in (r[1] for r in rows) if d)
        notes.append([title, len(rows), len(keys), dates[0] if dates else "",
                      dates[-1] if dates else "", sum(r[11] or 0 for r in rows)])
    for line in [
        [],
        ["Definitions"],
        ["Grain", "one row per (document, item, category, material, lot)"],
        ["Cold Transfer Out", "dispatch FROM a cold site (Cold Storage / Savla D-39 / D-514 / Rishi / Supreme / Eskimo)"],
        ["Cold Transfer In", "receipt INTO a cold site; union of cold_transfer_in_headers + interunit_transfer_in_header"],
        ["Challan", "the transfer-OUT challan no on every sheet, so IN rows join straight back to OUT rows"],
        ["Date", "OUT sheets = stock transfer date; IN sheets = GRN date (transfer-out date is in 'Out Date')"],
        ["Boxes (OUT)", "COUNT of transfer lines - cold dispatch writes one line per carton"],
        ["Qty / Boxes (IN)", "boxes actually scanned in. The OUT-side Qty can be a pack count for PM/RM, so Qty may differ by design"],
        ["Total Weight (IN)", "gross weight where scanned; cold receipts record net only, so net = total there"],
        ["Status", "OUT = challan status; IN = GRN status"],
        ["Received", "Received if the challan is Received, else Pending when a GRN exists, else Not Received"],
        ["Date cut", f"document date <= {CUTOFF:%Y-%m-%d} OR created before {CUT_TS:%Y-%m-%d} (keeps future-dated challans entered in the period)"],
        [],
        ["Known data caveats"],
        ["GRN-20260608124120", "interunit scan ledger holds 369 box rows for 185 cartons (duplicate-box incident); the cold table's 185 is used"],
        ["GRN-20260330132729 / 133149", "cold table holds a single summary row, so the per-box interunit ledger is used instead"],
        ["Box Source", "which table the IN rows were counted from, per GRN"],
        ["Category / Material (IN)", "carried over from the matching transfer-OUT line (by lot + item); blank when no match"],
    ]:
        notes.append(line)
    notes.column_dimensions["A"].width = 30
    notes.column_dimensions["B"].width = 110
    notes["A1"].font = Font(bold=True, size=13)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    for title, _h, rows in sheets:
        print(f"  {title:20s} {len(rows):6d} rows")


if __name__ == "__main__":
    main()
