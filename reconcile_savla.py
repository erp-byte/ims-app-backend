"""
SAVLA / RISHI PHYSICAL RECONCILIATION  (READ-ONLY — no DB writes)

Physical sheet = source of truth (Savla_Rishi_Inventory_4th June 2026.xlsx).
Compares lot-by-lot against cfpl_cold_stocks / cdpl_cold_stocks (1 row = 1 carton),
explains every mismatch using outbound movements:
  - Inter-unit Transfers (interunit_transfer_boxes + header)
  - Job Work OUT      (jb_materialout_lines + header)
  - Direct Out        ({co}_cold_storage_direct_out.removed_stock_snapshot)
  - Pending / In-Transit (pending_transfer_stock)
  - Disposition ledger cross-check (cold_stock_disposition, live since 2026-05-29)

Output -> ../Corrections/Savla_Rishi_Reconciliation_<date>.xlsx
"""
import io, sys, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import psycopg2
from collections import defaultdict, Counter

SHEET_PATH = "../Corrections/Savla_Rishi_Inventory_4th June 2026.xlsx"
OUT_PATH = "../Corrections/Savla_Rishi_Reconciliation_4thJun2026.xlsx"
CARTON_TOL = 0          # cartons must match exactly
KG_TOL = 1.0            # kg diff <= this and cartons equal -> still MATCH

# ───────────────────────── DB connect ─────────────────────────
env = {}
with open(".env", encoding="utf-8") as f:
    for ln in f:
        ln = ln.strip()
        if ln and not ln.startswith("#") and "=" in ln:
            k, v = ln.split("=", 1); env[k.strip()] = v.strip()
conn = psycopg2.connect(host=env["DB_HOST"], port=int(env.get("DB_PORT", "5432")),
                        dbname=env["DB_NAME"], user=env["DB_USER"], password=env["DB_PASSWORD"])
cur = conn.cursor()


def fnum(x):
    try: return float(x)
    except: return 0.0


# ───────────────────────── parse physical sheet ─────────────────────────
wb = openpyxl.load_workbook(SHEET_PATH, read_only=True, data_only=True)
ws = wb["4th June 26"]
phys = {}   # lot -> aggregate
for r in ws.iter_rows(min_row=7, values_only=True):
    lot = r[5]
    if lot is None or str(lot).strip() == "":
        continue
    lot = str(lot).strip()
    d = phys.setdefault(lot, {
        "lot": lot, "cartons": 0.0, "total_kg": 0.0, "rows": 0,
        "units": set(), "companies": set(), "locs": set(),
        "items": set(), "inwards": set(), "inward_dts": set(),
    })
    d["cartons"] += fnum(r[6])
    d["total_kg"] += fnum(r[8])
    d["rows"] += 1
    if r[1]: d["units"].add(str(r[1]).strip())
    if r[14]: d["companies"].add(str(r[14]).strip())
    if r[15]: d["locs"].add(str(r[15]).strip())
    if r[13]: d["items"].add(str(r[13]).strip())
    elif r[9]: d["items"].add(str(r[9]).strip())
    if r[2]: d["inwards"].add(str(r[2]).strip())
    if r[0]: d["inward_dts"].add(str(r[0]).strip())
wb.close()
lots = list(phys.keys())
print(f"Physical: {len(lots)} lots, {sum(d['cartons'] for d in phys.values()):.0f} cartons")

# ───────────────────────── DB cold_stocks aggregate by lot ─────────────────────────
db_lot = {}   # lot -> {company, cartons, kg, locs, units, items}
for tbl, co in (("cdpl_cold_stocks", "CDPL"), ("cfpl_cold_stocks", "CFPL")):
    cur.execute(
        f"SELECT lot_no, count(*), COALESCE(sum(total_inventory_kgs),0), "
        f"  string_agg(DISTINCT COALESCE(storage_location,'?'), ', '), "
        f"  string_agg(DISTINCT COALESCE(unit,'?'), ', '), "
        f"  string_agg(DISTINCT COALESCE(item_description,''), ' | ') "
        f"FROM {tbl} WHERE lot_no = ANY(%s) GROUP BY lot_no", (lots,))
    for lot, n, kg, locs, units, items in cur.fetchall():
        # lot is unique to one company table (verified: 0 overlap)
        db_lot[lot] = {"company": co, "cartons": int(n), "kg": float(kg),
                       "locs": locs or "", "units": units or "", "items": (items or "")[:120]}

# ───────────────────────── bulk movement fetches ─────────────────────────
mov = defaultdict(lambda: {"transfer": [], "jobwork": [], "directout": [],
                           "pending": [], "disp": Counter()})

# Transfers OUT (boxes carry lot_number); header has status / sites / date
cur.execute("""
    SELECT b.lot_number, h.challan_no, h.status, h.from_site, h.to_site,
           h.stock_trf_date, count(*), COALESCE(sum(b.net_weight),0)
    FROM interunit_transfer_boxes b
    JOIN interunit_transfers_header h ON h.id = b.header_id
    WHERE b.lot_number = ANY(%s)
    GROUP BY b.lot_number, h.challan_no, h.status, h.from_site, h.to_site, h.stock_trf_date
""", (lots,))
for lot, ch, st, frm, to, dt, n, kg in cur.fetchall():
    mov[lot]["transfer"].append({"ref": ch, "status": st, "from": frm, "to": to,
                                 "date": str(dt) if dt else "", "cartons": int(n), "kg": float(kg)})

# Job Work OUT
cur.execute("""
    SELECT l.lot_number, h.challan_no, h.status, h.from_warehouse, h.to_party,
           h.job_work_date, COALESCE(sum(l.quantity_boxes),0), COALESCE(sum(l.quantity_kgs),0)
    FROM jb_materialout_lines l
    JOIN jb_materialout_header h ON h.id = l.header_id
    WHERE l.lot_number = ANY(%s) AND COALESCE(h.type,'OUT')='OUT'
    GROUP BY l.lot_number, h.challan_no, h.status, h.from_warehouse, h.to_party, h.job_work_date
""", (lots,))
for lot, ch, st, frm, to, dt, nb, kg in cur.fetchall():
    mov[lot]["jobwork"].append({"ref": ch, "status": st, "from": frm, "to": to,
                                "date": str(dt) if dt else "", "cartons": int(fnum(nb)), "kg": float(kg)})

# Direct Out — removed_stock_snapshot is JSONB array of removed cold_stocks rows
for tbl in ("cdpl_cold_storage_direct_out", "cfpl_cold_storage_direct_out"):
    try:
        cur.execute(f"""
            SELECT elem->>'lot_no' AS lot, d.transaction_no, d.status, d.to_customer,
                   d.entry_date, count(*),
                   COALESCE(sum( (elem->>'total_inventory_kgs')::numeric ),0)
            FROM {tbl} d, jsonb_array_elements(d.removed_stock_snapshot) elem
            WHERE elem->>'lot_no' = ANY(%s)
            GROUP BY elem->>'lot_no', d.transaction_no, d.status, d.to_customer, d.entry_date
        """, (lots,))
        for lot, tn, st, cust, dt, n, kg in cur.fetchall():
            mov[lot]["directout"].append({"ref": tn, "status": st, "from": "", "to": cust,
                                          "date": str(dt) if dt else "", "cartons": int(n), "kg": float(kg)})
    except Exception as e:
        print(f"  directout {tbl} skip: {e}"); conn.rollback()

# Pending / In-Transit
cur.execute("""
    SELECT lot_no, status, transfer_out_challan_no, to_site,
           count(*), COALESCE(sum(weight_kg),0)
    FROM pending_transfer_stock WHERE lot_no = ANY(%s)
    GROUP BY lot_no, status, transfer_out_challan_no, to_site
""", (lots,))
for lot, st, ch, to, n, kg in cur.fetchall():
    mov[lot]["pending"].append({"ref": ch, "status": st, "to": to,
                                "cartons": int(n), "kg": float(kg)})

# Disposition ledger cross-check
cur.execute("""
    SELECT lot_no, disposition_type, count(*)
    FROM cold_stock_disposition WHERE lot_no = ANY(%s) AND reverted = false
    GROUP BY lot_no, disposition_type
""", (lots,))
for lot, dt, n in cur.fetchall():
    mov[lot]["disp"][dt] += int(n)

# ───────────────────────── reverse: DB Savla/Rishi lots NOT in physical sheet ─────────────────────────
SAVLA_LOC = "(storage_location ILIKE '%%savla%%' OR storage_location ILIKE '%%rishi%%' OR storage_location ILIKE '%%supreme%%')"
reverse = []
for tbl, co in (("cdpl_cold_stocks", "CDPL"), ("cfpl_cold_stocks", "CFPL")):
    cur.execute(
        f"SELECT lot_no, count(*), COALESCE(sum(total_inventory_kgs),0), "
        f"  string_agg(DISTINCT COALESCE(storage_location,'?'),', '), "
        f"  string_agg(DISTINCT COALESCE(unit,'?'),', '), "
        f"  string_agg(DISTINCT COALESCE(item_description,''),' | ') "
        f"FROM {tbl} WHERE {SAVLA_LOC} AND COALESCE(lot_no,'') <> '' "
        f"  AND NOT (lot_no = ANY(%s)) GROUP BY lot_no ORDER BY count(*) DESC", (lots,))
    for lot, n, kg, locs, units, items in cur.fetchall():
        reverse.append({"lot": lot, "company": co, "cartons": int(n), "kg": float(kg),
                        "locs": locs or "", "units": units or "", "items": (items or "")[:120]})

# ───────────────────────── build reconciliation rows ─────────────────────────
def mv_summary(m):
    parts = []
    for kind, label in (("transfer", "TRANSFER"), ("jobwork", "JOBWORK"), ("directout", "DIRECT-OUT")):
        for e in m[kind]:
            seg = f"{label} {e['ref']} [{e.get('status','')}]"
            if e.get("from") or e.get("to"):
                seg += f" {e.get('from','')}→{e.get('to','')}"
            seg += f" : {e['cartons']} ctn / {e['kg']:.0f}kg"
            if e.get("date"): seg += f" ({e['date']})"
            parts.append(seg)
    for e in m["pending"]:
        parts.append(f"IN-TRANSIT/pending {e.get('ref','')} [{e.get('status','')}]→{e.get('to','')} : {e['cartons']} ctn / {e['kg']:.0f}kg")
    if m["disp"]:
        parts.append("ledger: " + ", ".join(f"{k}={v}" for k, v in m["disp"].items()))
    return " || ".join(parts)


def out_breakdown(m):
    """Cartons that LEFT cold per the system, de-duplicated.

    Transfer 'Received' -> shipped & arrived (no pending row) -> count.
    Transfer 'Dispatch'/'Partial' -> in transit -> a pending row should exist;
        count via pending to avoid double-count, fall back to header qty.
    Jobwork (any status) and Direct-Out (status 'pending' but stock IS removed) -> count.
    Disposition ledger = independent cross-check (recent, catches relabeled boxes).
    """
    tr_recv = sum(e["cartons"] for e in m["transfer"] if "receiv" in str(e["status"]).lower())
    tr_intransit_hdr = sum(e["cartons"] for e in m["transfer"]
                           if str(e["status"]).lower() in ("dispatch", "partial"))
    pending = sum(e["cartons"] for e in m["pending"])
    intransit = pending if pending else tr_intransit_hdr
    jobwork = sum(e["cartons"] for e in m["jobwork"])
    directout = sum(e["cartons"] for e in m["directout"])
    src_out = tr_recv + intransit + jobwork + directout
    ledger_out = sum(m["disp"].values())
    return src_out, ledger_out


recon = []
for lot in lots:
    p = phys[lot]
    pc, pk = p["cartons"], p["total_kg"]
    db = db_lot.get(lot)
    m = mov.get(lot, {"transfer": [], "jobwork": [], "directout": [], "pending": [], "disp": Counter()})
    movsum = mv_summary(m) if any(m[k] for k in ("transfer", "jobwork", "directout", "pending")) or m["disp"] else ""
    src_out, ledger_out = out_breakdown(m)
    evidence = max(src_out, ledger_out)   # best estimate of cartons the system moved out

    if db is None:
        dbc, dbk, company = 0, 0.0, (sorted(p["companies"])[0] if p["companies"] else "?")
        status = "NOT IN SYSTEM"
        if evidence > 0:
            reason = (f"Physical {pc:.0f} ctn present, but system cold-stock = 0. System recorded ~{evidence:.0f} ctn "
                      f"of this lot as moved OUT (transfer/jobwork/direct-out — see Movements/ledger). Stock was deducted "
                      f"from cold_stocks yet is physically still here -> over-dispatch / relabeled boxes / movement not "
                      f"actually shipped. Action: stock physically present; correct the movement or re-add to cold_stocks.")
        else:
            reason = (f"Physical {pc:.0f} ctn present but NO system record at all (no cold stock, no transfer/jobwork/"
                      f"direct-out). Inward likely never entered or was deleted. Action: create inward entry.")
    else:
        dbc, dbk, company = db["cartons"], db["kg"], db["company"]
        diff = dbc - pc   # +ve = system overstated
        if abs(diff) <= CARTON_TOL and abs(dbk - pk) <= max(KG_TOL, 0.01 * max(pk, 1)):
            status = "MATCH"
            reason = "OK — physical = system." + (f" (system also shows past movement of this lot — see Movements.)" if movsum else "")
        elif abs(diff) <= CARTON_TOL:
            status = "MATCH (kg diff)"
            reason = f"Cartons match ({pc:.0f}); only kg differs by {dbk - pk:+.1f} (per-carton weight/rate data variance)."
        elif diff > 0:
            status = "SYSTEM OVERSTATED"
            if evidence >= diff and evidence > 0:
                reason = (f"System has {diff:.0f} ctn MORE than physical. This lot has recorded outbound movements of "
                          f"~{evidence:.0f} ctn (see Movements) — likely these boxes were shipped but their cold_stocks "
                          f"rows were never deleted (phantom/leak). Action: verify the movements and deduct the shipped boxes.")
            elif evidence > 0:
                reason = (f"System overstated by {diff:.0f} ctn. Recorded outbound movements of this lot ~{evidence:.0f} ctn "
                          f"(see Movements) likely explain part of it; remaining ~{max(diff-evidence,0):.0f} ctn unexplained "
                          f"-> physical write-off after count.")
            else:
                reason = (f"System overstated by {diff:.0f} ctn but NO outbound movement found -> phantom stock / "
                          f"over-recorded inward. Action: write-off after physical count.")
        else:
            short = -diff
            status = "SYSTEM SHORT"
            base = f"System has {short:.0f} ctn LESS than physical (system under-counts the physically present stock)."
            if evidence > 0:
                reason = (base + f" Note: system also shows ~{evidence:.0f} ctn of this lot moved OUT (received elsewhere/"
                          f"jobwork/direct-out) yet physical at Savla still shows more -> physical may include already-shipped "
                          f"stock, transfer not physically executed, or duplicate physical count. Needs review.")
            else:
                reason = base + " No outbound movement -> inward under-recorded or over-deduction. Needs review."

    recon.append({
        "lot": lot, "company": company,
        "phys_unit": ", ".join(sorted(p["units"])), "phys_loc": ", ".join(sorted(p["locs"])),
        "item": (" | ".join(sorted(p["items"])))[:80],
        "inward_no": ", ".join(sorted(p["inwards"]))[:40],
        "phys_cartons": pc, "db_cartons": dbc, "diff_cartons": dbc - pc,
        "phys_kg": pk, "db_kg": dbk, "diff_kg": dbk - pk,
        "db_loc": db["locs"] if db else "", "db_unit": db["units"] if db else "",
        "status": status, "src_out": src_out, "ledger_out": ledger_out,
        "movements": movsum, "reason": reason,
    })

# order: problems first
order = {"SYSTEM OVERSTATED": 0, "NOT IN SYSTEM": 1, "SYSTEM SHORT": 2, "MATCH (kg diff)": 3, "MATCH": 4}
recon.sort(key=lambda x: (order.get(x["status"], 9), -abs(x["diff_cartons"])))

# ───────────────────────── write workbook ─────────────────────────
HFILL = PatternFill("solid", fgColor="1F4E78")
HFONT = Font(bold=True, color="FFFFFF", size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
STATUS_FILL = {
    "MATCH": "E2EFDA", "MATCH (kg diff)": "FFF2CC", "SYSTEM OVERSTATED": "FCE4D6",
    "NOT IN SYSTEM": "F8CBAD", "SYSTEM SHORT": "DDEBF7",
}
wbk = openpyxl.Workbook()

def style_header(ws, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HFILL; cell.font = HFONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{ws.max_row}"

# --- Summary sheet ---
wss = wbk.active; wss.title = "Summary"
cnt = Counter(r["status"] for r in recon)
over_ctn = sum(r["diff_cartons"] for r in recon if r["diff_cartons"] > 0)
short_ctn = sum(-r["diff_cartons"] for r in recon if r["diff_cartons"] < 0)
summ = [
    ["SAVLA / RISHI PHYSICAL RECONCILIATION", ""],
    ["Physical sheet", "Savla_Rishi_Inventory_4th June 2026.xlsx (as on 2026-06-03/04)"],
    ["Generated", "2026-06-04"],
    ["Matching key", "lot_no (verified unique per company & unit)"],
    ["", ""],
    ["Physical lots", len(lots)],
    ["Physical cartons", round(sum(d['cartons'] for d in phys.values()), 1)],
    ["Physical total kg", round(sum(d['total_kg'] for d in phys.values()), 1)],
    ["", ""],
    ["STATUS BREAKDOWN", "lots"],
    ["  MATCH", cnt.get("MATCH", 0)],
    ["  MATCH (kg diff only)", cnt.get("MATCH (kg diff)", 0)],
    ["  SYSTEM OVERSTATED (DB > physical)", cnt.get("SYSTEM OVERSTATED", 0)],
    ["  NOT IN SYSTEM (DB = 0)", cnt.get("NOT IN SYSTEM", 0)],
    ["  SYSTEM SHORT (DB < physical)", cnt.get("SYSTEM SHORT", 0)],
    ["", ""],
    ["Total system-OVERSTATED cartons", round(over_ctn, 1)],
    ["Total system-SHORT cartons", round(short_ctn, 1)],
    ["", ""],
    ["Reverse: DB Savla/Rishi lots NOT in physical sheet", len(reverse)],
    ["  -> their cartons", sum(r['cartons'] for r in reverse)],
    ["  -> their kg", round(sum(r['kg'] for r in reverse), 1)],
    ["", ""],
    ["NOTES", ""],
    ["", "DB cold_stocks = 1 row per carton; aggregated by lot_no."],
    ["", "Cartons compared exactly; kg tolerance 1kg/1%."],
    ["", "Movements pulled from transfers, jobwork OUT, direct-out snapshots, pending, disposition ledger."],
    ["", "Disposition ledger only live since 2026-05-29 — older moves come from source tables."],
]
for row in summ:
    wss.append(row)
wss["A1"].font = Font(bold=True, size=13, color="1F4E78")
for r in range(1, wss.max_row + 1):
    if wss.cell(row=r, column=1).value in ("STATUS BREAKDOWN", "NOTES"):
        wss.cell(row=r, column=1).font = Font(bold=True, color="1F4E78")
wss.column_dimensions["A"].width = 42
wss.column_dimensions["B"].width = 70

# --- Reconciliation sheet ---
wsr = wbk.create_sheet("Reconciliation")
cols = [("Lot No", 12), ("Company", 9), ("Item (physical)", 34), ("Unit", 10), ("Phys Loc", 12),
        ("Inward No", 16), ("Phys Cartons", 10), ("DB Cartons", 10), ("Diff Ctn", 9),
        ("Phys Kg", 11), ("DB Kg", 11), ("Diff Kg", 10), ("DB Location", 16), ("DB Unit", 10),
        ("Status", 18), ("Src-Out Ctn", 10), ("Ledger-Out Ctn", 11),
        ("Movements (Transfer / Jobwork / Direct-Out / Pending)", 70),
        ("Reason / Explanation", 78)]
wsr.append([c[0] for c in cols])
for r in recon:
    wsr.append([r["lot"], r["company"], r["item"], r["phys_unit"], r["phys_loc"], r["inward_no"],
                r["phys_cartons"], r["db_cartons"], r["diff_cartons"],
                round(r["phys_kg"], 1), round(r["db_kg"], 1), round(r["diff_kg"], 1),
                r["db_loc"], r["db_unit"], r["status"], r["src_out"], r["ledger_out"],
                r["movements"], r["reason"]])
for i, (_, w) in enumerate(cols, 1):
    wsr.column_dimensions[get_column_letter(i)].width = w
style_header(wsr, len(cols))
for ridx in range(2, wsr.max_row + 1):
    st = wsr.cell(row=ridx, column=15).value
    fill = STATUS_FILL.get(st)
    if fill:
        for c in range(1, len(cols) + 1):
            wsr.cell(row=ridx, column=c).fill = PatternFill("solid", fgColor=fill)
    for c in (18, 19):
        wsr.cell(row=ridx, column=c).alignment = Alignment(wrap_text=True, vertical="top")

# --- Reverse sheet ---
wsv = wbk.create_sheet("System not in Physical")
rcols = [("Lot No", 12), ("Company", 9), ("Item (system)", 40), ("Unit", 10),
         ("DB Location", 18), ("DB Cartons", 11), ("DB Kg", 12)]
wsv.append([c[0] for c in rcols])
for r in sorted(reverse, key=lambda x: -x["cartons"]):
    wsv.append([r["lot"], r["company"], r["items"], r["units"], r["locs"], r["cartons"], round(r["kg"], 1)])
for i, (_, w) in enumerate(rcols, 1):
    wsv.column_dimensions[get_column_letter(i)].width = w
style_header(wsv, len(rcols))

wbk.save(OUT_PATH)
conn.close()

# ───────────────────────── console summary ─────────────────────────
print("\n===== RECONCILIATION SUMMARY =====")
for k in ("MATCH", "MATCH (kg diff)", "SYSTEM OVERSTATED", "NOT IN SYSTEM", "SYSTEM SHORT"):
    print(f"  {k:22} {cnt.get(k,0)} lots")
print(f"  System overstated cartons: {over_ctn:.0f}")
print(f"  System short cartons:      {short_ctn:.0f}")
print(f"  Reverse (system not in physical): {len(reverse)} lots, "
      f"{sum(r['cartons'] for r in reverse)} ctn")
print(f"\nSaved -> {OUT_PATH}")
