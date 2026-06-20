"""Scan all double-dispatched boxes (same box_id+txn+lot listed in 2+ transfers)
and append a detailed sheet to Interunit_Transfer_Reconciliation.xlsx.

READ-ONLY on the DB. Writes only the Excel file (adds a new sheet, preserves
existing sheets).
"""
import os, io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
from sqlalchemy import create_engine, text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

raw = os.environ["DATABASE_URL"]
DB = raw.replace("postgresql://", "postgresql+psycopg://", 1) if raw.startswith("postgresql://") else raw
eng = create_engine(DB)

XLSX = "Interunit_Transfer_Reconciliation.xlsx"
SHEET = "Double-Dispatched Boxes"

DETAIL = text("""
    WITH dd AS (
        SELECT box_id, transaction_no, COALESCE(lot_number,'') lot
        FROM interunit_transfer_boxes
        WHERE box_id IS NOT NULL AND box_id <> ''
        GROUP BY box_id, transaction_no, COALESCE(lot_number,'')
        HAVING COUNT(DISTINCT header_id) > 1
    )
    SELECT ob.box_id, ob.transaction_no, ob.lot_number, ob.article,
           MAX(ob.net_weight) net_weight,
           ob.header_id AS listed_in, COUNT(*) AS rows_in_transfer,
           h.challan_no, h.status AS transfer_status, h.from_site, h.to_site,
           (SELECT ti.transfer_out_id FROM interunit_transfer_in_boxes ib
              JOIN interunit_transfer_in_header ti ON ti.id = ib.header_id
              WHERE ib.box_id = ob.box_id AND ib.transaction_no = ob.transaction_no
                AND TRIM(COALESCE(ib.lot_number,'')) = TRIM(COALESCE(ob.lot_number,''))
              LIMIT 1) AS received_by,
           (SELECT p.transfer_out_id FROM pending_transfer_stock p
              WHERE p.box_id = ob.box_id AND p.transaction_no = ob.transaction_no
                AND TRIM(COALESCE(p.lot_no,'')) = TRIM(COALESCE(ob.lot_number,''))
                AND p.status = 'In Transit' LIMIT 1) AS intransit_under,
           (SELECT COUNT(DISTINCT header_id) FROM interunit_transfer_boxes b2
              WHERE b2.box_id = ob.box_id AND b2.transaction_no = ob.transaction_no
                AND COALESCE(b2.lot_number,'') = COALESCE(ob.lot_number,'')) AS in_n_transfers,
           (CASE WHEN EXISTS(SELECT 1 FROM cfpl_cold_stocks s WHERE s.box_id=ob.box_id AND s.transaction_no=ob.transaction_no AND TRIM(COALESCE(s.lot_no,''))=TRIM(COALESCE(ob.lot_number,'')))
                  OR EXISTS(SELECT 1 FROM cdpl_cold_stocks s WHERE s.box_id=ob.box_id AND s.transaction_no=ob.transaction_no AND TRIM(COALESCE(s.lot_no,''))=TRIM(COALESCE(ob.lot_number,'')))
                 THEN 1 ELSE 0 END) AS in_cold
    FROM interunit_transfer_boxes ob
    JOIN dd ON dd.box_id = ob.box_id AND dd.transaction_no = ob.transaction_no
           AND dd.lot = COALESCE(ob.lot_number,'')
    JOIN interunit_transfers_header h ON h.id = ob.header_id
    GROUP BY ob.box_id, ob.transaction_no, ob.lot_number, ob.article,
             ob.header_id, h.challan_no, h.status, h.from_site, h.to_site
    ORDER BY ob.box_id, ob.transaction_no, ob.header_id
""")

with eng.connect() as c:
    c.execute(text("SET TRANSACTION READ ONLY"))
    rows = c.execute(DETAIL).fetchall()
    # transfer_out_id -> challan map for owner display
    chal = {r._mapping["id"]: r._mapping["challan_no"]
            for r in c.execute(text("SELECT id, challan_no FROM interunit_transfers_header")).fetchall()}

print(f"Double-dispatched box occurrences: {len(rows)}")

headers = ["Box ID", "Transaction No", "Lot No", "Article", "Net Wt (kg)",
           "Appears in (# transfers)", "Listed in Transfer", "Challan No",
           "From", "To", "Transfer Status", "Dup rows in transfer",
           "Role", "Owner Transfer", "Owner Challan", "Box Current Location"]

data_rows = []
n_spurious = 0
for r in rows:
    m = r._mapping
    owner = m["received_by"] or m["intransit_under"]
    listed = m["listed_in"]
    if owner is None:
        role = "UNRESOLVED (no receipt/in-transit)"
    elif owner == listed:
        role = "OWNER"
    else:
        role = "SPURIOUS (duplicate listing)"; n_spurious += 1
    if m["received_by"]:
        loc = f"Received by transfer {m['received_by']}"
    elif m["intransit_under"]:
        loc = f"In-Transit under transfer {m['intransit_under']}"
    elif m["in_cold"]:
        loc = "Still in COLD stock"
    else:
        loc = "Nowhere (not cold / not received / not in-transit)"
    data_rows.append([
        m["box_id"], m["transaction_no"], m["lot_number"] or "", m["article"] or "",
        float(m["net_weight"] or 0), int(m["in_n_transfers"] or 0),
        listed, m["challan_no"], m["from_site"], m["to_site"], m["transfer_status"],
        int(m["rows_in_transfer"] or 1), role, owner or "", chal.get(owner, "") if owner else "",
        loc,
    ])

wb = openpyxl.load_workbook(XLSX)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)

hdr_fill = PatternFill("solid", fgColor="1F4E78"); hdr_font = Font(bold=True, color="FFFFFF")
ws.append(headers)
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col); cell.fill = hdr_fill; cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
spur_fill = PatternFill("solid", fgColor="FCE4E4")
for dr in data_rows:
    ws.append(dr)
    if str(dr[12]).startswith("SPURIOUS"):
        for col in range(1, len(headers) + 1):
            ws.cell(row=ws.max_row, column=col).fill = spur_fill
ws.freeze_panes = "A2"
widths = [16, 20, 10, 34, 11, 12, 14, 22, 14, 10, 14, 12, 26, 13, 22, 40]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# summary at far right
ws.cell(row=1, column=len(headers)+2, value="SUMMARY").font = Font(bold=True)
summ = [
    ("Total double-dispatched occurrences", len(data_rows)),
    ("Distinct (box_id,txn,lot) keys", len({(d[0], d[1], d[2]) for d in data_rows})),
    ("SPURIOUS (duplicate) listings", n_spurious),
    ("Transfers involved", len({d[6] for d in data_rows})),
]
for i, (k, v) in enumerate(summ, start=2):
    ws.cell(row=i, column=len(headers)+2, value=k)
    ws.cell(row=i, column=len(headers)+3, value=v)

wb.save(XLSX)
print(f"Added sheet '{SHEET}' to {XLSX} — {len(data_rows)} rows ({n_spurious} spurious).")
print("Done.")
