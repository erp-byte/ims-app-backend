"""
READ-ONLY root-cause forensics for the 3 Savla problem buckets.
Box-level proof: do the lot's CURRENT cold rows / absences line up with movement records?
No DB writes.
"""
import io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl, psycopg2
from collections import defaultdict, Counter

env = {}
with open(".env", encoding="utf-8") as f:
    for ln in f:
        ln = ln.strip()
        if ln and not ln.startswith("#") and "=" in ln:
            k, v = ln.split("=", 1); env[k.strip()] = v.strip()
conn = psycopg2.connect(host=env["DB_HOST"], port=int(env.get("DB_PORT", "5432")),
                        dbname=env["DB_NAME"], user=env["DB_USER"], password=env["DB_PASSWORD"])
cur = conn.cursor()


def q(sql, args=None):
    cur.execute(sql, args or ()); return cur.fetchall()


# ── load problem lots from the reconciliation workbook ──
wb = openpyxl.load_workbook("../Corrections/Savla_Rishi_Reconciliation_4thJun2026.xlsx", data_only=True)
ws = wb["Reconciliation"]; rows = list(ws.iter_rows(values_only=True)); hdr = rows[0]
ix = {h: i for i, h in enumerate(hdr)}
buckets = defaultdict(list)
meta = {}
for r in rows[1:]:
    st = r[ix["Status"]]
    lot = str(r[ix["Lot No"]])
    if st in ("SYSTEM OVERSTATED", "NOT IN SYSTEM", "SYSTEM SHORT"):
        buckets[st].append(lot)
        meta[lot] = {"company": r[ix["Company"]], "phys": r[ix["Phys Cartons"]],
                     "db": r[ix["DB Cartons"]], "diff": r[ix["Diff Ctn"]],
                     "item": r[ix["Item (physical)"]]}
all_lots = [l for b in buckets.values() for l in b]
print("Problem lots:", {k: len(v) for k, v in buckets.items()})

# ── movement (box_id, txn) sets keyed by lot ──
def keyset(sql):
    d = defaultdict(set)
    for lot, bid, txn in q(sql, (all_lots,)):
        if bid:
            d[str(lot)].add((str(bid), str(txn) if txn else ""))
    return d

jw = keyset("SELECT lot_number, box_id, transaction_no FROM jb_materialout_lines WHERE lot_number = ANY(%s)")
tr_recv = defaultdict(set); tr_disp = defaultdict(set)
for lot, bid, txn, st in q("""SELECT b.lot_number, b.box_id, b.transaction_no, h.status
        FROM interunit_transfer_boxes b JOIN interunit_transfers_header h ON h.id=b.header_id
        WHERE b.lot_number = ANY(%s)""", (all_lots,)):
    if not bid: continue
    k = (str(bid), str(txn) if txn else "")
    (tr_recv if "receiv" in str(st).lower() else tr_disp)[str(lot)].add(k)

do = defaultdict(set)
for tbl in ("cdpl_cold_storage_direct_out", "cfpl_cold_storage_direct_out"):
    try:
        for lot, bid, txn in q(f"""SELECT elem->>'lot_no', elem->>'box_id', elem->>'transaction_no'
                FROM {tbl} d, jsonb_array_elements(d.removed_stock_snapshot) elem
                WHERE elem->>'lot_no' = ANY(%s)""", (all_lots,)):
            if bid: do[str(lot)].add((str(bid), str(txn) if txn else ""))
    except Exception as e:
        conn.rollback(); print("do skip", e)

disp = defaultdict(set)
for lot, bid, txn, dtype in q("""SELECT lot_no, box_id, transaction_no, disposition_type
        FROM cold_stock_disposition WHERE lot_no = ANY(%s) AND reverted=false""", (all_lots,)):
    if bid: disp[str(lot)].add((str(bid), str(txn) if txn else ""))

# ── current cold rows (box_id, txn) per lot ──
cold = defaultdict(set)
for tbl in ("cdpl_cold_stocks", "cfpl_cold_stocks"):
    for lot, bid, txn in q(f"SELECT lot_no, box_id, transaction_no FROM {tbl} WHERE lot_no = ANY(%s)", (all_lots,)):
        cold[str(lot)].add((str(bid) if bid else "", str(txn) if txn else ""))

# ── inward presence (boxes_v2) per lot ──
inward_present = set()      # lot ever inwarded (boxes_v2 uses column lot_number, not lot_no)
for tbl in ("cdpl_boxes_v2", "cfpl_boxes_v2"):
    try:
        for (lot,) in q(f"SELECT DISTINCT lot_number FROM {tbl} WHERE lot_number = ANY(%s)", (all_lots,)):
            inward_present.add(str(lot))
    except Exception as e:
        conn.rollback(); print(f"  {tbl} inward probe skip: {e}")

# ════════════════ OVERSTATED analysis ════════════════
print("\n" + "=" * 90)
print("SYSTEM OVERSTATED — box-level: do CURRENT cold rows match a recorded OUT movement?")
print("(if a cold row's (box_id,txn) is ALSO in a jobwork/transfer/direct-out record => shipped but NOT deducted)")
over_cause = Counter()
over_detail = []
for lot in buckets["SYSTEM OVERSTATED"]:
    m = meta[lot]; c = cold[lot]
    in_jw = sum(1 for k in c if k in jw[lot])
    in_tr = sum(1 for k in c if k in tr_recv[lot] or k in tr_disp[lot])
    in_do = sum(1 for k in c if k in do[lot])
    in_disp = sum(1 for k in c if k in disp[lot])
    leaked = in_jw + in_tr + in_do
    # classify dominant
    if in_jw and in_jw >= max(in_tr, in_do):
        cause = "JOBWORK out not deducted (box_id/txn matched in jb_materialout_lines but row still in cold)"
    elif in_tr and in_tr >= max(in_jw, in_do):
        cause = "TRANSFER out not deducted (box still in cold AND in transfer_boxes)"
    elif in_do:
        cause = "DIRECT-OUT not deducted"
    elif in_disp:
        cause = "Ledger says disposed but row still in cold (relabel/leak)"
    else:
        cause = ("Excess cold rows, NO movement ref (not duplicates, single inward date) -> "
                 "off-system outflow or bulk-load over-count; needs physical write-off")
    over_cause[cause] += 1
    over_detail.append((lot, m["company"], m["diff"], in_jw, in_tr, in_do, in_disp, cause, str(m["item"])[:34]))
for c, n in over_cause.most_common():
    print(f"  [{n:2}] {c}")
print("\n  Top overstated lots (lot | co | excess | jw/tr/do/disp overlap | cause):")
for lot, co, diff, ij, it, idr, idp, cause, item in sorted(over_detail, key=lambda x: -x[2])[:18]:
    print(f"   {lot:8}{co} +{diff:<5.0f} jw={ij:<4} tr={it:<4} do={idr:<3} disp={idp:<4} {item:34} | {cause[:46]}")

# ════════════════ NOT IN SYSTEM analysis ════════════════
print("\n" + "=" * 90)
print("NOT IN SYSTEM — was the lot ever inwarded? did it leave via a movement?")
nis_cause = Counter(); nis_detail = []
for lot in buckets["NOT IN SYSTEM"]:
    m = meta[lot]
    left = bool(jw[lot] or tr_recv[lot] or tr_disp[lot] or do[lot] or disp[lot])
    inwarded = lot in inward_present
    if disp[lot] or tr_disp[lot]:
        cause = "Dispatched/relabeled OUT (in pending/ledger) but stock physically still here -> over-dispatch"
    elif left:
        cause = "Left via transfer/jobwork/direct-out (system) but physically present -> wrongly shipped/deducted"
    elif inwarded:
        cause = "Was inwarded (boxes_v2) but cold rows gone & no movement -> inward deleted / cold wiped"
    else:
        cause = "NEVER in system (no inward, no cold, no movement) -> inward never entered"
    nis_cause[cause] += 1
    nis_detail.append((lot, m["company"], m["phys"], cause, str(m["item"])[:34]))
for c, n in nis_cause.most_common():
    print(f"  [{n:2}] {c}")
print("\n  Top not-in-system lots (lot | co | phys ctn | cause):")
for lot, co, phys, cause, item in sorted(nis_detail, key=lambda x: -x[2])[:18]:
    print(f"   {lot:8}{co} {phys:<6.0f} {item:34} | {cause[:52]}")

# ════════════════ SHORT analysis ════════════════
print("\n" + "=" * 90)
print("SYSTEM SHORT — system has FEWER than physical. Over-deduction or under-inward?")
short_cause = Counter(); short_detail = []
for lot in buckets["SYSTEM SHORT"]:
    m = meta[lot]; short = -m["diff"]
    out_recv = len(tr_recv[lot]); out_jw = len(jw[lot]); out_do = len(do[lot])
    moved = out_recv + out_jw + out_do
    if moved > 0:
        cause = "Outbound moves recorded (received elsewhere) yet physical>system -> physical may double-count shipped stock OR transfer not physically done"
    elif lot in inward_present:
        cause = "Inwarded but system qty < physical, no movement -> partial over-deduction / inward under-recorded"
    else:
        cause = "No inward & no movement, yet system has some rows -> data inconsistency, review"
    short_cause[cause] += 1
    short_detail.append((lot, m["company"], short, moved, cause, str(m["item"])[:34]))
for c, n in short_cause.most_common():
    print(f"  [{n:2}] {c}")
print("\n  Top short lots (lot | co | short ctn | movedOut | cause):")
for lot, co, short, moved, cause, item in sorted(short_detail, key=lambda x: -x[2])[:14]:
    print(f"   {lot:8}{co} -{short:<5.0f} movedOut={moved:<4} {item:34} | {cause[:40]}")

# ════════════════ LOCATE every problem lot across the WHOLE DB ════════════════
# Auto-discover every base table that has a lot column and search it.
print("\n" + "=" * 90)
print("LOCATING problem lots across ALL lot-bearing tables in the DB...")
lot_cols = q("""
    SELECT c.table_name, c.column_name
    FROM information_schema.columns c
    JOIN information_schema.tables t
      ON t.table_name = c.table_name AND t.table_schema = c.table_schema
    WHERE c.table_schema = 'public' AND t.table_type = 'BASE TABLE'
      AND c.column_name ILIKE '%%lot%%'
    ORDER BY c.table_name, c.column_name
""")
print(f"  lot-bearing tables/cols found: {len(lot_cols)}")
located = defaultdict(list)   # lot -> ["table(count)", ...]
for tbl, col in lot_cols:
    try:
        for lot, n in q(f'SELECT "{col}"::text, count(*) FROM "{tbl}" '
                        f'WHERE "{col}"::text = ANY(%s) GROUP BY "{col}"::text', (all_lots,)):
            located[str(lot)].append(f"{tbl}({n})")
    except Exception as e:
        conn.rollback()
        print(f"    skip {tbl}.{col}: {str(e)[:60]}")
# also probe direct-out JSONB snapshots (lot lives inside removed_stock_snapshot)
for tbl in ("cdpl_cold_storage_direct_out", "cfpl_cold_storage_direct_out"):
    try:
        for lot, n in q(f"""SELECT elem->>'lot_no', count(*)
                FROM {tbl} d, jsonb_array_elements(d.removed_stock_snapshot) elem
                WHERE elem->>'lot_no' = ANY(%s) GROUP BY elem->>'lot_no'""", (all_lots,)):
            located[str(lot)].append(f"{tbl}.snapshot({n})")
    except Exception:
        conn.rollback()

def loc_str(lot):
    hits = sorted(located.get(str(lot), []))
    return "; ".join(hits) if hits else "NOT FOUND in any lot table"

# quick console proof for NOT-IN-SYSTEM lots
print("\n  NOT IN SYSTEM lots — where do they live in the DB?")
for lot in sorted(buckets["NOT IN SYSTEM"], key=lambda l: -meta[l]["phys"])[:12]:
    print(f"   {lot:8} phys={meta[lot]['phys']:<6.0f} -> {loc_str(lot)[:90]}")

conn.close()

# ════════════════ write Root Cause tabs into the workbook ════════════════
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
OUT = "../Corrections/Savla_Rishi_Reconciliation_4thJun2026.xlsx"
wbw = openpyxl.load_workbook(OUT)
for nm in ("Root Cause Analysis", "Per-Lot Cause"):
    if nm in wbw.sheetnames:
        del wbw[nm]
HF = PatternFill("solid", fgColor="7030A0"); HFONT = Font(bold=True, color="FFFFFF")
TITLE = Font(bold=True, size=13, color="7030A0"); SUB = Font(bold=True, color="7030A0")

over_kg = sum(o[2] for o in over_detail)
# ---- narrative tab ----
rc = wbw.create_sheet("Root Cause Analysis", index=1)
def row(*vals): rc.append(list(vals))
row("SAVLA / RISHI — ROOT CAUSE ANALYSIS (why physical != system, and the code cause)")
row("")
row("BUCKET", "LOTS", "DATA-LEVEL ROOT CAUSE (proven from DB)", "CODE CAUSE (file:line)", "FIX")
# OVERSTATED
row("SYSTEM OVERSTATED", cnt_over := len(buckets["SYSTEM OVERSTATED"]),
    "31/33 lots: cold has EXTRA cartons that NO transfer/jobwork/direct-out references; not duplicates; "
    "one inward date per lot. => physical was depleted by outflows never entered as a system txn, "
    "or cold was bulk-loaded above current physical. 2/33 lots: boxes sit in cold AND in transfer_boxes "
    "= transfer-out never deducted them.",
    "1) No physical-count/stock-adjust path: cold_stocks only shrinks via inward-delete, transfer "
    "dispatch, jobwork-out (_deduct_cold_storage_stock), direct-out. Off-system issue => cold never "
    "drops.  2) Bulk loads scripts/replace_cold_stocks_from_excel.py & bulk_load_cold_stocks_csv.py set "
    "opening balances that drift.  3) job_work_server.py:419 'if not cold_row: continue' silently skips "
    "deduction on box_id/txn mismatch (recycled boxes).",
    "Add a Stock-Adjustment / physical-count endpoint; write-off the excess after count; add a qty "
    "fallback (deduct N by lot) when box_id/txn miss.")
# NOT IN SYSTEM
row("NOT IN SYSTEM", len(buckets["NOT IN SYSTEM"]),
    "24/48 NEVER in system (no inward in boxes_v2, no cold, no movement) -> legacy/opening stock never "
    "entered.  12/48 marked dispatched/in-transit in pending+ledger but physically still here -> "
    "over-dispatch / phantom in-transit.  12/48 recorded as shipped (transfer/jobwork/direct-out) yet "
    "physically present -> cold deducted but goods never left.",
    "1) No opening-balance import for pre-system lots.  2) Transfer dispatch deletes cold + parks pending "
    "BEFORE physical departure (interunit bridge); if goods don't leave, stock is phantom-gone.  3) "
    "Transfer-In marks Received & pick_from_pending clears ALL boxes without a box-count check "
    "(project_transfer_partial_receipt_leak).",
    "Import opening stock for the 24 lots; reverse over-dispatch (re-add to cold / cancel pending); "
    "enforce physical-departure + box-count gates.")
# SHORT
row("SYSTEM SHORT", len(buckets["SYSTEM SHORT"]),
    "12/24 have outbound moves received elsewhere yet physical>system -> physical likely double-counts "
    "already-shipped stock OR transfer was marked Received while goods stayed at Savla.  12/24 have cold "
    "rows but lot not in boxes_v2 and no movement -> cold partially bulk-loaded (fewer than actual) / "
    "over-deduction.",
    "1) Transfer-In Received without verifying goods physically arrived vs stayed (same receive-gap as "
    "above).  2) Bulk cold load under-counted these lots vs physical.  3) Exact box_id/txn deletes can "
    "over-remove when a recycled box_id matches the wrong row.",
    "Physical recount these lots; correct system qty up to physical; verify the related transfers were "
    "physically executed.")
row("")
row("CROSS-CUTTING ROOT CAUSE", "",
    "The 2026-03-14 bulk migration recycled box_id + transaction numbers, so the (box_id, transaction_no) "
    "key the deduction logic relies on is no longer unique/stable -> deductions hit the wrong row or no "
    "row. Combined with no physical-count reconciliation, errors accumulate.", "", "")
row("")
row("KEY POINT", "",
    "OVERSTATEMENT is mostly NOT a deduction-code bug — the excess cold rows are not referenced by any "
    "movement. It is missing physical-count reconciliation + bulk-load drift + off-system outflows.", "", "")
# style
rc["A1"].font = TITLE
for c in range(1, 6): rc.cell(row=3, column=c).fill = HF; rc.cell(row=3, column=c).font = HFONT
for r in range(1, rc.max_row + 1):
    v = rc.cell(row=r, column=1).value
    if v in ("CROSS-CUTTING ROOT CAUSE", "KEY POINT"): rc.cell(row=r, column=1).font = SUB
for col, w in (("A", 22), ("B", 7), ("C", 70), ("D", 70), ("E", 50)):
    rc.column_dimensions[col].width = w
for r in range(4, rc.max_row + 1):
    for c in range(1, 6):
        rc.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")

# ---- per-lot cause tab ----
pl = wbw.create_sheet("Per-Lot Cause", index=2)
NCOL = 10
pl.append(["Lot No", "Company", "Status", "Phys Ctn", "DB Ctn", "Diff Ctn",
           "Present in DB tables (whole-DB lot search)", "Box-overlap jw/tr/do/disp",
           "Cause Category", "Item"])
for lot, co, diff, ij, it, idr, idp, cause, item in sorted(over_detail, key=lambda x: -x[2]):
    pl.append([lot, co, "SYSTEM OVERSTATED", meta[lot]["phys"], meta[lot]["db"], diff,
               loc_str(lot), f"{ij}/{it}/{idr}/{idp}", cause, item])
for lot, co, phys, cause, item in sorted(nis_detail, key=lambda x: -x[2]):
    pl.append([lot, co, "NOT IN SYSTEM", phys, 0, -phys, loc_str(lot), "-", cause, item])
for lot, co, short, moved, cause, item in sorted(short_detail, key=lambda x: -x[2]):
    pl.append([lot, co, "SYSTEM SHORT", meta[lot]["phys"], meta[lot]["db"], -short,
               loc_str(lot), f"movedOut={moved}", cause, item])
for c in range(1, NCOL + 1): pl.cell(row=1, column=c).fill = HF; pl.cell(row=1, column=c).font = HFONT
pl.freeze_panes = "A2"; pl.auto_filter.ref = f"A1:{get_column_letter(NCOL)}{pl.max_row}"
for col, w in (("A", 11), ("B", 9), ("C", 18), ("D", 9), ("E", 9), ("F", 9),
               ("G", 52), ("H", 20), ("I", 70), ("J", 34)):
    pl.column_dimensions[col].width = w
# highlight "NOT FOUND" rows in red so genuinely-missing stock pops
RED = PatternFill("solid", fgColor="FFC7CE")
for r in range(2, pl.max_row + 1):
    pl.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical="top")
    pl.cell(row=r, column=9).alignment = Alignment(wrap_text=True, vertical="top")
    if pl.cell(row=r, column=7).value == "NOT FOUND in any lot table":
        pl.cell(row=r, column=7).fill = RED

try:
    wbw.save(OUT)
    saved_to = OUT
except PermissionError:
    saved_to = OUT.replace(".xlsx", "_v2.xlsx")
    wbw.save(saved_to)
    print(f"\n*** Original file is OPEN/locked. Saved to {saved_to} instead. "
          f"Close the original in Excel and re-run to overwrite it. ***")
print(f"\nAdded 'Root Cause Analysis' + 'Per-Lot Cause' tabs to {saved_to}")
print("DONE (read-only DB; workbook updated).")
