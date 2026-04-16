"""
Jobwork Dashboard API — queries actual jb_* tables.

Tables used:
  jb_materialout_header   — dispatch header (to_party=vendor, sub_category=process_type)
  jb_materialout_lines    — dispatched items (quantity_kgs, item_description)
  jb_work_inward_receipt  — inward receipt header (ir_number, receipt_date, receipt_type)
  jb_work_inward_lines    — receipt line items (finished_goods_kgs, waste_kgs, rejection_kgs)

Endpoints:
  GET /jobwork/dashboard/summary          — KPIs + grouped summary rows
  GET /jobwork/dashboard/filter-options   — dropdown options for filters
  GET /jobwork/dashboard/group-details    — expanded JWO rows within a group
  GET /jobwork/dashboard/jwo-receipts/{id}— IR receipts for a single JWO header
  GET /jobwork/dashboard/export-excel     — Excel export of filtered data
"""

import io
from datetime import date
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from shared.database import get_db
from shared.logger import get_logger

logger = get_logger("jobwork_dashboard")

router = APIRouter(
    prefix="/jobwork/dashboard",
    tags=["Jobwork Dashboard"],
)


def _f(v):
    """Round to 2 decimal places, default 0."""
    return round(float(v), 2) if v else 0.0


def _parse_csv(val: Optional[str]) -> List[str]:
    if not val:
        return []
    return [v.strip() for v in val.split(",") if v.strip()]


def _build_where_clauses(
    company: str,
    date_from: Optional[str],
    date_to: Optional[str],
    vendors: Optional[str],
    items: Optional[str],
    process_types: Optional[str],
    jwo_statuses: Optional[str],
):
    """Build WHERE clause fragments and params dict for jb_materialout_header (alias h)."""
    clauses = ["1=1"]
    params = {}

    # No company column in jb_materialout_header — skip if not applicable
    # If you have a company column, uncomment:
    # clauses.append("h.company = :company")
    # params["company"] = company.upper()

    if date_from:
        clauses.append("h.job_work_date >= :date_from")
        params["date_from"] = date_from
    if date_to:
        clauses.append("h.job_work_date <= :date_to")
        params["date_to"] = date_to

    vendor_list = _parse_csv(vendors)
    if vendor_list:
        placeholders = ", ".join(f":vendor_{i}" for i in range(len(vendor_list)))
        clauses.append(f"h.to_party IN ({placeholders})")
        for i, v in enumerate(vendor_list):
            params[f"vendor_{i}"] = v

    item_list = _parse_csv(items)
    if item_list:
        placeholders = ", ".join(f":item_{i}" for i in range(len(item_list)))
        clauses.append(f"h.sub_category IN ({placeholders})")
        for i, v in enumerate(item_list):
            params[f"item_{i}"] = v

    pt_list = _parse_csv(process_types)
    if pt_list:
        placeholders = ", ".join(f":pt_{i}" for i in range(len(pt_list)))
        clauses.append(f"h.purpose_of_work IN ({placeholders})")
        for i, v in enumerate(pt_list):
            params[f"pt_{i}"] = v

    status_list = _parse_csv(jwo_statuses)
    if status_list:
        placeholders = ", ".join(f":st_{i}" for i in range(len(status_list)))
        clauses.append(f"h.status IN ({placeholders})")
        for i, v in enumerate(status_list):
            params[f"st_{i}"] = v

    return " AND ".join(clauses), params


def _status_display(raw_status: str) -> str:
    """Map internal status values to display labels."""
    mapping = {
        "sent": "Open",
        "partially_received": "Partially Received",
        "fully_received": "Fully Received",
        "reconciled": "Reconciled",
        "closed": "Closed",
    }
    return mapping.get(raw_status, raw_status)


def _status_internal(display_status: str) -> str:
    """Map display labels back to internal status values."""
    mapping = {
        "Open": "sent",
        "Partially Received": "partially_received",
        "Fully Received": "fully_received",
        "Reconciled": "reconciled",
        "Closed": "closed",
    }
    return mapping.get(display_status, display_status)


def _group_expression(group_by: str) -> str:
    """Return SQL expression for GROUP BY."""
    if group_by == "vendor":
        return "h.to_party"
    elif group_by == "item":
        return "h.sub_category"
    elif group_by == "process_type":
        return "h.purpose_of_work"
    elif group_by == "month":
        return "TO_CHAR(h.job_work_date::date, 'YYYY-MM')"
    elif group_by == "jwo_status":
        return "h.status"
    return "h.to_party"


# ── Endpoint: Summary ────────────────────────────────────────────

@router.get("/summary")
async def get_dashboard_summary(
    company: str = Query("cdpl"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    vendors: Optional[str] = Query(None),
    items: Optional[str] = Query(None),
    process_types: Optional[str] = Query(None),
    jwo_statuses: Optional[str] = Query(None),
    loss_statuses: Optional[str] = Query(None),
    group_by: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Return KPI cards + grouped summary table based on filters."""
    # Map display statuses to internal values
    if jwo_statuses:
        mapped = ",".join(_status_internal(s.strip()) for s in jwo_statuses.split(",") if s.strip())
        jwo_statuses = mapped

    where_clause, params = _build_where_clauses(
        company, date_from, date_to, vendors, items, process_types, jwo_statuses,
    )

    effective_group_by = group_by or "vendor"
    grp_expr = _group_expression(effective_group_by)

    try:
        sql = text(f"""
            SELECT
                {grp_expr} AS group_label,
                COUNT(DISTINCT h.id) AS num_jwos,
                COALESCE(SUM(dispatched.total_kgs), 0) AS total_dispatched,
                COALESCE(SUM(dispatched.total_net_kgs), 0) AS total_net_dispatched,
                COALESCE(SUM(received.total_fg), 0) AS total_fg,
                COALESCE(SUM(received.total_waste), 0) AS total_waste,
                COALESCE(SUM(received.total_rejection), 0) AS total_rejection,
                COALESCE(SUM(dispatched.total_net_kgs), 0)
                    - COALESCE(SUM(received.total_fg), 0)
                    - COALESCE(SUM(received.total_waste), 0)
                    - COALESCE(SUM(received.total_rejection), 0) AS unaccounted_balance,
                CASE WHEN COALESCE(SUM(dispatched.total_net_kgs), 0) > 0
                    THEN (
                        COALESCE(SUM(dispatched.total_net_kgs), 0)
                        - COALESCE(SUM(received.total_fg), 0)
                        - COALESCE(SUM(received.total_waste), 0)
                        - COALESCE(SUM(received.total_rejection), 0)
                    ) / COALESCE(SUM(dispatched.total_net_kgs), 0) * 100
                    ELSE 0
                END AS avg_loss_pct,
                SUM(CASE WHEN h.status IN ('sent', 'partially_received') THEN 1 ELSE 0 END) AS open_jwos,
                SUM(CASE WHEN h.status IN ('sent', 'partially_received')
                    AND h.job_work_date::date < CURRENT_DATE - INTERVAL '30 days'
                    THEN 1 ELSE 0 END) AS overdue_jwos,
                SUM(CASE WHEN received.has_excess THEN 1 ELSE 0 END) AS excess_loss_flags
            FROM jb_materialout_header h
            LEFT JOIN LATERAL (
                SELECT
                    COALESCE(SUM(ml.quantity_kgs), 0) AS total_kgs,
                    COALESCE(SUM(CAST(ml.net_weight AS NUMERIC)), 0) AS total_net_kgs
                FROM jb_materialout_lines ml
                WHERE ml.header_id = h.id
            ) dispatched ON true
            LEFT JOIN LATERAL (
                SELECT
                    COALESCE(SUM(il.finished_goods_kgs), 0) AS total_fg,
                    COALESCE(SUM(il.waste_kgs), 0) AS total_waste,
                    COALESCE(SUM(il.rejection_kgs), 0) AS total_rejection,
                    MAX(ir.receipt_date) AS last_receipt_date,
                    CASE WHEN SUM(il.sent_kgs) > 0
                        AND (SUM(il.sent_kgs) - COALESCE(SUM(il.finished_goods_kgs), 0)
                             - COALESCE(SUM(il.waste_kgs), 0)
                             - COALESCE(SUM(il.rejection_kgs), 0))
                            / SUM(il.sent_kgs) * 100 > 10
                        THEN true ELSE false
                    END AS has_excess
                FROM jb_work_inward_receipt ir
                JOIN jb_work_inward_lines il ON il.inward_receipt_id = ir.id
                WHERE ir.header_id = h.id
            ) received ON true
            WHERE {where_clause}
            GROUP BY {grp_expr}
            ORDER BY {grp_expr}
        """)

        rows = db.execute(sql, params).fetchall()

        summary = []
        kpi_total_jwos = 0
        kpi_total_dispatched = 0.0
        kpi_total_fg = 0.0
        kpi_open_pending = 0
        kpi_excess_flags = 0
        loss_pcts = []

        for r in rows:
            label = str(r.group_label) if r.group_label else "Unknown"
            # Map internal status to display label for jwo_status grouping
            if effective_group_by == "jwo_status":
                label = _status_display(label)

            row_dict = {
                "group_label": label,
                "num_jwos": int(r.num_jwos or 0),
                "total_dispatched_kgs": _f(r.total_net_dispatched),
                "total_gross_dispatched_kgs": _f(r.total_dispatched),
                "total_fg_received_kgs": _f(r.total_fg),
                "total_waste_received_kgs": _f(r.total_waste),
                "total_rejection_kgs": _f(r.total_rejection),
                "unaccounted_balance_kgs": _f(r.unaccounted_balance),
                "avg_loss_pct": _f(r.avg_loss_pct),
                "open_jwos": int(r.open_jwos or 0),
                "overdue_jwos": int(r.overdue_jwos or 0),
                "excess_loss_flags": int(r.excess_loss_flags or 0),
                "avg_turnaround_days": 0,
            }
            summary.append(row_dict)

            kpi_total_jwos += row_dict["num_jwos"]
            kpi_total_dispatched += row_dict["total_dispatched_kgs"]
            kpi_total_fg += row_dict["total_fg_received_kgs"]
            kpi_open_pending += row_dict["open_jwos"]
            kpi_excess_flags += row_dict["excess_loss_flags"]
            if row_dict["avg_loss_pct"] > 0:
                loss_pcts.append(row_dict["avg_loss_pct"])

        avg_loss = round(sum(loss_pcts) / len(loss_pcts), 2) if loss_pcts else 0.0

        today = date.today()
        return {
            "kpis": {
                "total_jwos": kpi_total_jwos,
                "total_dispatched_kgs": _f(kpi_total_dispatched),
                "total_fg_received_kgs": _f(kpi_total_fg),
                "avg_loss_pct": avg_loss,
                "open_pending_jwos": kpi_open_pending,
                "excess_loss_flags": kpi_excess_flags,
            },
            "summary": summary,
            "group_by": effective_group_by,
            "as_of_date": today.isoformat(),
            "filters_applied": 0,
        }
    except Exception as exc:
        logger.error("Dashboard summary error: %s", exc, exc_info=True)
        raise HTTPException(500, f"Failed to load dashboard: {str(exc)}")


# ── Endpoint: Filter Options ────────────────────────────────────

@router.get("/filter-options")
async def get_filter_options(
    company: str = Query("cdpl"),
    db: Session = Depends(get_db),
):
    """Return distinct dropdown values for the filter panel."""
    try:
        # Vendors (to_party) with active JWO count
        vendor_rows = db.execute(text("""
            SELECT to_party AS name,
                   SUM(CASE WHEN status IN ('sent', 'partially_received') THEN 1 ELSE 0 END) AS active_count
            FROM jb_materialout_header
            GROUP BY to_party
            ORDER BY to_party
        """)).fetchall()
        vendors = [{"name": r.name, "active_jwo_count": int(r.active_count)} for r in vendor_rows if r.name]

        # Items (sub_category)
        item_rows = db.execute(text("""
            SELECT DISTINCT sub_category
            FROM jb_materialout_header
            WHERE sub_category IS NOT NULL AND sub_category != ''
            ORDER BY sub_category
        """)).fetchall()
        items_list = [r.sub_category for r in item_rows]

        # Process types (purpose_of_work)
        pt_rows = db.execute(text("""
            SELECT DISTINCT purpose_of_work
            FROM jb_materialout_header
            WHERE purpose_of_work IS NOT NULL AND purpose_of_work != ''
            ORDER BY purpose_of_work
        """)).fetchall()
        process_types = [r.purpose_of_work for r in pt_rows]

        return {
            "vendors": vendors,
            "items": items_list,
            "process_types": process_types,
        }
    except Exception as exc:
        logger.error("Filter options error: %s", exc, exc_info=True)
        raise HTTPException(500, f"Failed to load filter options: {str(exc)}")


# ── Endpoint: Group Details (expand a group row) ────────────────

@router.get("/group-details")
async def get_group_details(
    company: str = Query("cdpl"),
    group_by: str = Query("vendor"),
    group_label: str = Query(""),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    vendors: Optional[str] = Query(None),
    items: Optional[str] = Query(None),
    process_types: Optional[str] = Query(None),
    jwo_statuses: Optional[str] = Query(None),
    loss_statuses: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Return individual JWO rows within a specific group."""
    if jwo_statuses:
        mapped = ",".join(_status_internal(s.strip()) for s in jwo_statuses.split(",") if s.strip())
        jwo_statuses = mapped

    where_clause, params = _build_where_clauses(
        company, date_from, date_to, vendors, items, process_types, jwo_statuses,
    )

    # Add group filter
    if group_label:
        if group_by == "vendor":
            where_clause += " AND h.to_party = :grp_label"
            params["grp_label"] = group_label
        elif group_by == "item":
            where_clause += " AND h.sub_category = :grp_label"
            params["grp_label"] = group_label
        elif group_by == "process_type":
            where_clause += " AND h.purpose_of_work = :grp_label"
            params["grp_label"] = group_label
        elif group_by == "month":
            where_clause += " AND TO_CHAR(h.job_work_date::date, 'YYYY-MM') = :grp_label"
            params["grp_label"] = group_label
        elif group_by == "jwo_status":
            where_clause += " AND h.status = :grp_label"
            params["grp_label"] = _status_internal(group_label)

    try:
        sql = text(f"""
            SELECT
                h.id,
                h.challan_no AS jwo_id,
                h.job_work_date AS dispatch_date,
                h.to_party AS vendor_name,
                h.sub_category AS item_name,
                h.purpose_of_work AS process_type,
                COALESCE(dispatched.total_net_kgs, 0) AS qty_dispatched,
                h.status AS jwo_status,
                COALESCE(received.total_fg, 0) AS fg_received,
                COALESCE(received.total_waste, 0) AS waste_received,
                COALESCE(received.total_rejection, 0) AS rejection,
                COALESCE(dispatched.total_net_kgs, 0)
                    - COALESCE(received.total_fg, 0)
                    - COALESCE(received.total_waste, 0)
                    - COALESCE(received.total_rejection, 0) AS unaccounted_balance,
                CASE WHEN COALESCE(dispatched.total_net_kgs, 0) > 0
                    THEN (
                        COALESCE(dispatched.total_net_kgs, 0)
                        - COALESCE(received.total_fg, 0)
                        - COALESCE(received.total_waste, 0)
                        - COALESCE(received.total_rejection, 0)
                    ) / COALESCE(dispatched.total_net_kgs, 0) * 100
                    ELSE 0
                END AS actual_loss_pct,
                CASE
                    WHEN COALESCE(dispatched.total_net_kgs, 0) > 0
                        AND (
                            COALESCE(dispatched.total_net_kgs, 0)
                            - COALESCE(received.total_fg, 0)
                            - COALESCE(received.total_waste, 0)
                            - COALESCE(received.total_rejection, 0)
                        ) / COALESCE(dispatched.total_net_kgs, 0) * 100 > 10
                        THEN 'Excess Loss'
                    WHEN h.status IN ('sent', 'partially_received')
                        THEN 'Pending'
                    ELSE 'Normal'
                END AS loss_status
            FROM jb_materialout_header h
            LEFT JOIN LATERAL (
                SELECT
                    COALESCE(SUM(ml.quantity_kgs), 0) AS total_kgs,
                    COALESCE(SUM(CAST(ml.net_weight AS NUMERIC)), 0) AS total_net_kgs
                FROM jb_materialout_lines ml
                WHERE ml.header_id = h.id
            ) dispatched ON true
            LEFT JOIN LATERAL (
                SELECT
                    COALESCE(SUM(il.finished_goods_kgs), 0) AS total_fg,
                    COALESCE(SUM(il.waste_kgs), 0) AS total_waste,
                    COALESCE(SUM(il.rejection_kgs), 0) AS total_rejection
                FROM jb_work_inward_receipt ir
                JOIN jb_work_inward_lines il ON il.inward_receipt_id = ir.id
                WHERE ir.header_id = h.id
            ) received ON true
            WHERE {where_clause}
            ORDER BY h.job_work_date DESC
        """)

        rows = db.execute(sql, params).fetchall()

        return [
            {
                "id": r.id,
                "jwo_id": r.jwo_id or "",
                "dispatch_date": str(r.dispatch_date) if r.dispatch_date else None,
                "vendor_name": r.vendor_name or "",
                "item_name": r.item_name or "",
                "process_type": r.process_type or "",
                "qty_dispatched": _f(r.qty_dispatched),
                "fg_received": _f(r.fg_received),
                "waste_received": _f(r.waste_received),
                "rejection": _f(r.rejection),
                "unaccounted_balance": _f(r.unaccounted_balance),
                "actual_loss_pct": _f(r.actual_loss_pct),
                "loss_status": r.loss_status,
                "jwo_status": _status_display(r.jwo_status),
                "turnaround_days": None,
            }
            for r in rows
        ]
    except Exception as exc:
        logger.error("Group details error: %s", exc, exc_info=True)
        raise HTTPException(500, f"Failed to load group details: {str(exc)}")


# ── Endpoint: JWO Receipts (lazy-loaded IR expansion) ───────────

@router.get("/jwo-receipts/{header_id}")
async def get_jwo_receipts(
    header_id: int,
    company: str = Query("cdpl"),
    db: Session = Depends(get_db),
):
    """Return all Inward Receipts for a specific material-out header (lazy-loaded on row expand)."""
    try:
        sql = text("""
            SELECT
                ir.id,
                ir.ir_number,
                ir.receipt_date AS ir_date,
                ir.receipt_type,
                COALESCE(SUM(il.finished_goods_kgs), 0) AS fg_qty_received,
                COALESCE(SUM(il.waste_kgs), 0) AS waste_qty_received,
                COALESCE(SUM(il.rejection_kgs), 0) AS rejection_qty,
                CASE WHEN COALESCE(SUM(il.sent_kgs), 0) > 0
                    THEN (
                        COALESCE(SUM(il.sent_kgs), 0)
                        - COALESCE(SUM(il.finished_goods_kgs), 0)
                        - COALESCE(SUM(il.waste_kgs), 0)
                        - COALESCE(SUM(il.rejection_kgs), 0)
                    ) / COALESCE(SUM(il.sent_kgs), 0) * 100
                    ELSE 0
                END AS actual_loss_pct,
                CASE
                    WHEN COALESCE(SUM(il.sent_kgs), 0) > 0
                        AND (
                            COALESCE(SUM(il.sent_kgs), 0)
                            - COALESCE(SUM(il.finished_goods_kgs), 0)
                            - COALESCE(SUM(il.waste_kgs), 0)
                            - COALESCE(SUM(il.rejection_kgs), 0)
                        ) / COALESCE(SUM(il.sent_kgs), 0) * 100 > 10
                        THEN 'Excess Loss'
                    ELSE 'Normal'
                END AS loss_status
            FROM jb_work_inward_receipt ir
            LEFT JOIN jb_work_inward_lines il ON il.inward_receipt_id = ir.id
            WHERE ir.header_id = :header_id
            GROUP BY ir.id, ir.ir_number, ir.receipt_date, ir.receipt_type
            ORDER BY ir.receipt_date ASC
        """)

        rows = db.execute(sql, {"header_id": header_id}).fetchall()

        return [
            {
                "id": r.id,
                "ir_number": r.ir_number or "",
                "ir_date": str(r.ir_date) if r.ir_date else None,
                "receipt_type": r.receipt_type or "partial",
                "fg_qty_received": _f(r.fg_qty_received),
                "waste_qty_received": _f(r.waste_qty_received),
                "rejection_qty": _f(r.rejection_qty),
                "actual_loss_pct": _f(r.actual_loss_pct),
                "loss_status": r.loss_status,
            }
            for r in rows
        ]
    except Exception as exc:
        logger.error("JWO receipts error: %s", exc, exc_info=True)
        raise HTTPException(500, f"Failed to load receipts: {str(exc)}")


# ── Endpoint: Export Excel ──────────────────────────────────────

@router.get("/export-excel")
async def export_excel(
    company: str = Query("cdpl"),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    vendors: Optional[str] = Query(None),
    items: Optional[str] = Query(None),
    process_types: Optional[str] = Query(None),
    jwo_statuses: Optional[str] = Query(None),
    loss_statuses: Optional[str] = Query(None),
    group_by: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Export filtered dashboard data as .xlsx file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        summary_resp = await get_dashboard_summary(
            company, date_from, date_to, vendors, items,
            process_types, jwo_statuses, loss_statuses, group_by, db,
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        ws.merge_cells("A1:L1")
        ws["A1"] = f"Jobwork Summary — {company.upper()} — As of {summary_resp['as_of_date']}"
        ws["A1"].font = Font(bold=True, size=14)

        kpis = summary_resp["kpis"]
        kpi_labels = ["Total JWOs", "Total Dispatched (Kgs)", "Total FG Received (Kgs)",
                       "Avg Loss %", "Open/Pending JWOs", "Excess Loss Flags"]
        kpi_values = [kpis["total_jwos"], kpis["total_dispatched_kgs"], kpis["total_fg_received_kgs"],
                      kpis["avg_loss_pct"], kpis["open_pending_jwos"], kpis["excess_loss_flags"]]

        for i, (label, val) in enumerate(zip(kpi_labels, kpi_values)):
            ws.cell(row=3, column=i * 2 + 1, value=label).font = Font(bold=True)
            ws.cell(row=3, column=i * 2 + 2, value=val)

        headers = [
            "Group", "No. of JWOs", "Total Dispatched (Kgs)", "Total FG Received (Kgs)",
            "Total Waste (Kgs)", "Total Rejection (Kgs)", "Unaccounted Balance (Kgs)",
            "Avg Loss %", "Open JWOs", "Overdue JWOs", "Excess Loss Flags", "Avg Turnaround Days",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(summary_resp["summary"], 6):
            vals = [
                row_data["group_label"], row_data["num_jwos"],
                row_data["total_dispatched_kgs"], row_data["total_fg_received_kgs"],
                row_data["total_waste_received_kgs"], row_data["total_rejection_kgs"],
                row_data["unaccounted_balance_kgs"], row_data["avg_loss_pct"],
                row_data["open_jwos"], row_data["overdue_jwos"],
                row_data["excess_loss_flags"], row_data["avg_turnaround_days"],
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=v)
                cell.border = thin_border

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        today_str = date.today().strftime("%d%b%Y")
        filename = f"Jobwork_Summary_{company.upper()}_{today_str}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError:
        raise HTTPException(500, "openpyxl is not installed on the server. Run: pip install openpyxl")
    except Exception as exc:
        logger.error("Export Excel error: %s", exc, exc_info=True)
        raise HTTPException(500, f"Export failed: {str(exc)}")
