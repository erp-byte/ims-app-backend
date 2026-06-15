"""Delete 'Only in System' lots that are ABSENT from the real IMS inventory file
(IMS Savla_Rishi_Inventory_13th June 2026.xlsx), snapshotting each box into
cold_stock_disposition first.

Cross-check (user spec 14Jun): a lot flagged 'Only in System' in the comparison
sheet is SKIPPED if it is actually present in the IMS inventory file (false
positive). Only lots truly absent from the IMS file are deleted (all boxes).

Dry-run by default; --apply to write.
"""
import sys
import json
import openpyxl
import psycopg2
import psycopg2.extras

DB = dict(host="wms-postgres-db.cpis084golp7.ap-south-1.rds.amazonaws.com",
          port=5432, dbname="warehouse_db", user="wmsadmin", password="Candorfoods")
CMP = "../Corrections/Cold_and_IMS_Comparison_14Jun2026.xlsx"
IMS = "../Corrections/IMS Savla_Rishi_Inventory_13th June 2026.xlsx"
REF_NO = "IMS-ONLY-IN-SYSTEM-14JUN2026"
APPLY = "--apply" in sys.argv


def norm(v):
    if v is None:
        return None
    s = str(v).strip()
    return s[:-2] if s.endswith(".0") else s


def ims_lot_set():
    wb = openpyxl.load_workbook(IMS, data_only=True)
    ws = wb[wb.sheetnames[0]]
    lots = set()
    for r in range(7, ws.max_row + 1):
        lot = norm(ws.cell(r, 7).value)
        if lot:
            lots.add(lot)
    return lots


def only_in_system():
    wb = openpyxl.load_workbook(CMP, data_only=True)
    ws = wb["IMS_vs_System_Comparison"]
    out = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "Only in System":
            out.append((str(ws.cell(r, 2).value).strip().upper(), norm(ws.cell(r, 3).value)))
    return out


def main():
    ims = ims_lot_set()
    ois = only_in_system()
    conn = psycopg2.connect(**DB)
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    skipped_in_ims = skipped_absent = del_lots = err = 0
    del_boxes = 0
    print(f"MODE = {'APPLY' if APPLY else 'DRY-RUN'}   IMS-file lots={len(ims)}  only-in-system rows={len(ois)}\n")

    for ent, lot in ois:
        if lot in (None, "None"):
            continue
        if lot in ims:
            print(f"  [SKIP-IN-IMS] {ent} {lot}: present in IMS file")
            skipped_in_ims += 1
            continue
        tbl = "cdpl_cold_stocks" if ent == "CDPL" else "cfpl_cold_stocks"
        comp = "cdpl" if ent == "CDPL" else "cfpl"
        cur.execute(f"SELECT * FROM {tbl} WHERE lot_no = %s", (lot,))
        rows = cur.fetchall()
        if not rows:
            skipped_absent += 1
            continue
        try:
            if APPLY:
                for r in rows:
                    cur.execute("""
                        INSERT INTO cold_stock_disposition
                            (box_id, transaction_no, lot_no, item_description,
                             from_company, unit, from_site, source_table,
                             disposition_type, disposition_ref_no, disposed_by,
                             snapshot_data, notes)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'manual_correction',%s,'ims_recon',
                                CAST(%s AS JSONB), %s)
                    """, (r["box_id"], (r.get("transaction_no") or ""), lot, r.get("item_description"),
                          comp, r.get("unit"), r.get("storage_location"), tbl,
                          REF_NO, json.dumps(r, default=str),
                          "Only in System, absent from IMS Savla_Rishi 13Jun — deleted full lot"))
                cur.execute(f"DELETE FROM {tbl} WHERE lot_no = %s", (lot,))
                deleted = cur.rowcount
                cur.execute(f"SELECT COUNT(*) AS c FROM {tbl} WHERE lot_no = %s", (lot,))
                final = cur.fetchone()["c"]
                if deleted != len(rows) or final != 0:
                    conn.rollback()
                    print(f"  [ERR ] {ent} {lot}: deleted {deleted}/{len(rows)}, final {final}!=0 -> ROLLBACK")
                    err += 1
                    continue
                conn.commit()
            del_lots += 1
            del_boxes += len(rows)
            print(f"  DELETE {ent} {lot}: {len(rows)} boxes" + ("  [committed]" if APPLY else ""))
        except Exception as e:
            conn.rollback()
            print(f"  [ERR ] {ent} {lot}: {e!r} -> ROLLBACK")
            err += 1

    print(f"\n=== SUMMARY ({'APPLIED' if APPLY else 'DRY-RUN'}) ===")
    print(f"  skipped (present in IMS file) : {skipped_in_ims}")
    print(f"  skipped (already 0 boxes)     : {skipped_absent}")
    print(f"  DELETED lots={del_lots}  boxes={del_boxes}")
    print(f"  errors={err}")
    if not APPLY:
        print("\n  DRY-RUN — nothing written. Re-run with --apply.")
    conn.close()


if __name__ == "__main__":
    main()
