from io import BytesIO
from datetime import date
from html import escape
from typing import Literal, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.orm import Session

from shared.database import get_db
from services.ims_service.inward_models import Company
from services.ims_service.rtv_models import (
    RTVCreate,
    RTVHeaderUpdate,
    RTVWithDetails,
    RTVHeaderResponse,
    RTVListResponse,
    RTVDeleteResponse,
    RTVBoxUpsertRequest,
    RTVBoxUpsertResponse,
    RTVBulkBoxUpdateRequest,
    RTVBulkBoxUpdateResponse,
    RTVLinesUpdateRequest,
    RTVLinesUpdateResponse,
    RTVApprovalRequest,
    RTVApprovalResponse,
    RTVBoxEditLogRequest,
    RTVActionRequest,
    RTVActionResponse,
    SendForApprovalResponse,
)
from services.ims_service.rtv_tools import (
    create_rtv,
    list_rtvs,
    get_rtv,
    update_rtv,
    delete_rtv,
    upsert_rtv_box,
    bulk_save_boxes,
    update_rtv_lines,
    approve_rtv,
    log_rtv_box_edits,
    export_rtv_records,
    set_rtv_status,
    apply_rtv_email_action,
)
from services.ims_service.rtv_approval_token import verify_action_token, JWTError
from shared.email_notifier import (
    notify_rtv_created,
    notify_rtv_approved,
    notify_rtv_deleted,
    notify_rtv_header_updated,
    notify_rtv_lines_updated,
    notify_rtv_status_changed,
)

router = APIRouter(prefix="/rtv", tags=["rtv"])


# ── Export (must be before /{company} to avoid path conflict) ─────────


@router.get("/export")
def export_rtv_endpoint(
    company: Company = Query(..., description="Company code"),
    status: Optional[str] = Query(None),
    customer: Optional[str] = Query(None),
    factory_unit: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    sort_by: str = Query("created_ts"),
    sort_order: str = Query("desc"),
    db: Session = Depends(get_db),
):
    """Export filtered RTV records as an Excel (.xlsx) file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlalchemy import text as sa_text

    rows = export_rtv_records(
        company=company,
        status=status,
        customer=customer,
        factory_unit=factory_unit,
        from_date=from_date,
        to_date=to_date,
        sort_by=sort_by,
        sort_order=sort_order,
        db=db,
    )

    # Build edit-log lookup for highlighting
    rtv_ids = list({r["RTV ID"] for r in rows if r.get("RTV ID")})
    edited_cells: set[tuple[str, str]] = set()
    if rtv_ids:
        edit_rows = db.execute(
            sa_text("""
                SELECT box_id, field_name FROM box_edit_logs
                WHERE transaction_no = ANY(:rtv_ids)
            """),
            {"rtv_ids": rtv_ids},
        ).fetchall()
        for er in edit_rows:
            edited_cells.add((er.box_id, er.field_name))

    field_to_header = {
        "net_weight": "Box Net Weight",
        "gross_weight": "Box Gross Weight",
        "lot_number": "Box Lot Number",
        "count": "Box Count",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "RTV Records"

    headers = [
        "RTV ID", "RTV Date", "Factory Unit", "Customer",
        "Invoice Number", "Challan No", "DN No", "Conversion",
        "Sales POC", "Business Head", "Remark", "Status", "Created By", "Created At",
        "Material Type", "Item Category", "Sub Category",
        "Item Description", "UOM", "Qty", "Rate", "Value", "Line Net Weight", "Line Carton Weight",
        "Box ID", "Box Article", "Box Number", "Box UOM", "Box Conversion",
        "Box Net Weight", "Box Gross Weight", "Box Lot Number", "Box Count",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="29417A", end_color="29417A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    edited_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    box_header_cols = {h: i + 1 for i, h in enumerate(headers) if h in field_to_header.values()}

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        box_id = row_data.get("Box ID", "")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.border = thin_border

            if box_id and header in box_header_cols:
                for field_name, mapped_header in field_to_header.items():
                    if mapped_header == header and (box_id, field_name) in edited_cells:
                        cell.fill = edited_fill
                        break

    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(len(header) + 4, 14)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"rtv_{company}_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Box edit logging ─────────────────────────


@router.post("/box-edit-log")
def log_rtv_box_edit_endpoint(
    payload: RTVBoxEditLogRequest,
    db: Session = Depends(get_db),
):
    """Log audit entries for edits to a previously-printed RTV box."""
    return log_rtv_box_edits(payload, db)


# ── Email-button action (Approve / Reject / Hold) ─────────
# Registered before /{company} so FastAPI doesn't bind "action" as a company.


def _email_action_html(title: str, message: str, color: str) -> str:
    """Render the page that closes Gmail's in-app browser after a click."""
    return f"""<!DOCTYPE html>
<html lang="en"><head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{escape(title)}</title>
</head>
<body style="margin:0;padding:0;font-family:Arial,Helvetica,sans-serif;background:#f4f4f4;">
  <div style="max-width:480px;margin:60px auto;background:#fff;padding:36px 28px;border-radius:12px;text-align:center;box-shadow:0 4px 16px rgba(0,0,0,0.08);">
    <div style="width:64px;height:64px;border-radius:50%;background:{color};margin:0 auto 18px;line-height:64px;color:#fff;font-size:34px;font-weight:bold;">&#10003;</div>
    <h2 style="margin:0 0 10px;color:#222;font-size:20px;">{escape(title)}</h2>
    <p style="margin:0 0 22px;color:#555;font-size:14px;line-height:1.5;">{escape(message)}</p>
    <p style="margin:0;color:#999;font-size:12px;">This window will close automatically.</p>
  </div>
  <script>
    (function () {{
      // Best-effort close of mail-client in-app browser.
      function tryClose() {{
        try {{ window.close(); }} catch (_) {{}}
        try {{ self.close(); }} catch (_) {{}}
        try {{ window.history.back(); }} catch (_) {{}}
      }}
      tryClose();
      setTimeout(tryClose, 400);
      setTimeout(tryClose, 1200);
    }})();
  </script>
</body></html>"""


@router.get("/action", response_class=HTMLResponse)
def email_action_endpoint(
    rtv_id: str = Query(..., description="RTV id from the Created mail link"),
    bh_email: str = Query(..., description="Business head email from the link"),
    action: Literal["approve", "reject", "hold"] = Query(...),
    db: Session = Depends(get_db),
):
    """Email-button GET endpoint.

    Validates that bh_email matches the business_head stored on the RTV,
    applies the status transition (Approve/Reject/Hold), fires a
    confirmation mail naming who actioned the RTV, and returns an HTML page
    that closes the mail-client in-app browser.
    """
    try:
        result = apply_rtv_email_action(rtv_id=rtv_id, bh_email=bh_email, action=action, db=db)
    except HTTPException as exc:
        return HTMLResponse(
            _email_action_html(
                title="Action failed",
                message=str(exc.detail),
                color="#c0392b",
            ),
            status_code=exc.status_code,
        )

    if result["already_actioned"]:
        return HTMLResponse(
            _email_action_html(
                title="Already actioned",
                message=f"RTV {result['rtv_id']} is already {result['status']}.",
                color="#e67e22",
            )
        )

    # First successful state change → fire the confirmation mail.
    detail = result.get("detail")
    if isinstance(detail, dict):
        notify_rtv_status_changed(
            detail,
            new_status=result["status"],
            actioned_by=result["actioned_by"],
        )

    return HTMLResponse(
        _email_action_html(
            title=f"RTV {result['status']}",
            message=f"RTV {result['rtv_id']} is now {result['status']}.",
            color="#27ae60",
        )
    )


@router.post("/action", response_model=RTVActionResponse)
def set_rtv_status_endpoint(
    payload: RTVActionRequest,
    db: Session = Depends(get_db),
):
    """Programmatic status update (no token; for trusted internal callers).

    Email buttons use the GET endpoint above with a signed token. This POST
    remains for non-mail consumers and does NOT enforce BH-ownership.
    """
    return set_rtv_status(
        rtv_id=payload.rtv_id,
        business_head_email=payload.business_head_email,
        action=payload.action,
        db=db,
    )


# ── CRUD endpoints ───────────────────────────


@router.post("/{company}", response_model=RTVWithDetails, status_code=201)
def create_rtv_endpoint(
    company: Company,
    data: RTVCreate,
    created_by: str = Query("user@example.com"),
    db: Session = Depends(get_db),
):
    """Create a new RTV with header and line items."""
    result = create_rtv(data, created_by, db)
    result["_company"] = company
    return result


@router.get("/{company}", response_model=RTVListResponse)
def list_rtvs_endpoint(
    company: Company,
    page: int = Query(1, ge=1),
    per_page: int = Query(10, ge=1, le=100),
    status: Optional[str] = Query(None),
    factory_unit: Optional[str] = Query(None),
    customer: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    sort_by: str = Query("created_ts"),
    sort_order: str = Query("desc"),
    db: Session = Depends(get_db),
):
    """List RTVs with pagination and filters."""
    return list_rtvs(
        company, page, per_page, status, factory_unit, customer,
        from_date, to_date, sort_by, sort_order, db,
    )


@router.get("/{company}/{rtv_id}", response_model=RTVWithDetails)
def get_rtv_endpoint(
    company: Company,
    rtv_id: int,
    db: Session = Depends(get_db),
):
    """Get RTV detail with lines and boxes."""
    return get_rtv(company, rtv_id, db)


@router.post("/{company}/{rtv_id}/send-for-approval", response_model=SendForApprovalResponse)
def send_for_approval_endpoint(
    company: Company,
    rtv_id: int,
    db: Session = Depends(get_db),
):
    """Fire the Business-Head approval mail for an already-saved RTV.

    Separate from create so the FE can hard-save header+lines first, then send.
    Re-sendable; does not mutate lines/boxes."""
    detail = get_rtv(company, rtv_id, db)   # same builder GET /{company}/{rtv_id} uses
    detail["_company"] = company
    notify_rtv_created(detail)
    return {"status": "sent", "rtv_id": detail.get("rtv_id", "")}


@router.put("/{company}/{rtv_id}", response_model=RTVHeaderResponse)
def update_rtv_endpoint(
    company: Company,
    rtv_id: int,
    data: RTVHeaderUpdate,
    db: Session = Depends(get_db),
):
    """Update RTV header fields."""
    result = update_rtv(company, rtv_id, data, db)
    notify_rtv_header_updated(result)
    return result


@router.delete("/{company}/{rtv_id}", response_model=RTVDeleteResponse)
def delete_rtv_endpoint(
    company: Company,
    rtv_id: int,
    deleted_by: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Delete an RTV and all its lines and boxes."""
    result = delete_rtv(company, rtv_id, db)
    notify_rtv_deleted(
        rtv_id=result.get("rtv_id", ""),
        company=company,
        business_head=result.get("business_head"),
        created_by=result.get("created_by"),
        deleted_by=deleted_by,
    )
    return result


@router.put("/{company}/{rtv_id}/lines", response_model=RTVLinesUpdateResponse)
def update_rtv_lines_endpoint(
    company: Company,
    rtv_id: int,
    data: RTVLinesUpdateRequest,
    db: Session = Depends(get_db),
):
    """Replace all line items on an existing RTV."""
    result = update_rtv_lines(company, rtv_id, data, db)
    detail = get_rtv(company, rtv_id, db)
    notify_rtv_lines_updated(detail)
    return result


@router.put("/{company}/{rtv_id}/approve", response_model=RTVApprovalResponse)
def approve_rtv_endpoint(
    company: Company,
    rtv_id: int,
    payload: RTVApprovalRequest,
    notify: bool = Query(True, description="Send approval email (set false for create+approve combined flow)"),
    db: Session = Depends(get_db),
):
    """Approve an RTV with optional field completion."""
    result = approve_rtv(company, rtv_id, payload, db)
    if notify:
        detail = get_rtv(company, rtv_id, db)
        notify_rtv_approved(detail, payload.approved_by)
    return result


@router.put("/{company}/{rtv_id}/box", response_model=RTVBoxUpsertResponse)
def upsert_rtv_box_endpoint(
    company: Company,
    rtv_id: int,
    payload: RTVBoxUpsertRequest,
    db: Session = Depends(get_db),
):
    """Upsert a single RTV box (called at print time)."""
    return upsert_rtv_box(company, rtv_id, payload, db)


@router.put("/{company}/{rtv_id}/boxes", response_model=RTVBulkBoxUpdateResponse)
def bulk_save_boxes_endpoint(
    company: Company,
    rtv_id: int,
    payload: RTVBulkBoxUpdateRequest,
    notify_discrepancy: bool = True,
    db: Session = Depends(get_db),
):
    result = bulk_save_boxes(company, rtv_id, payload, db, notify_discrepancy=notify_discrepancy)
    return result


