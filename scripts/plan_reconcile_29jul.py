"""
READ-ONLY PLAN. Treats the 29-Jul-2026 closing sheet as EXACT and works out what
cold_stocks must become, lot by lot and box by box.

The sheet is a 29-Jul snapshot; the DB is live. So a raw diff would charge today's
real movement to the sheet. This computes the AS-OF-29-JUL position first:

    as_of_29jul = current_cold_boxes + departed_after_29jul - arrived_after_29jul

    departed_after_29jul : cold_stock_disposition rows created after the cutoff whose
                           source_table is a *_cold_stocks table and reverted IS NOT true
                           (a departure DELETEs the cold_stocks row, so add it back)
    arrived_after_29jul  : cold_stocks rows with created_at after the cutoff

The sheet is then compared to as_of_29jul (is the DB right as of the snapshot?), and the
action is computed against the CURRENT table:

    target_now = sheet_cartons - departed_after_29jul
    action     = target_now - current_cold_boxes

Actions never fabricate a real box id: inserts are proposed as marked synthetic ids
(RC29JUL-<lot>-NNNN, same convention as the executed RECON22JUL batch) and deletions are
costed against the safe pool only (never a box referenced by an in-flight movement table).

Run: python scripts/plan_reconcile_29jul.py [--xlsx OUT.xlsx]
"""
from __future__ import annotations
import os
import sys
import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl
import psycopg
from dotenv import load_dotenv

sys.path.insert(0, str(Path(__file__).resolve().parent))
from report_lots_in_closing_29jul import norm, _s, COL, HEADER_ROW, PREFIX

ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
EXCEL_PATH = ROOT / "data" / "Cold Storage Stock Details 29th July Closing.xlsx"
CUTOFF = "2026-07-29 23:59:59"   # sheet "As On Date" is 29-Jul-2026

# box_id columns meaning a box is in-flight -> never counted as safely deletable
REF_SOURCES = {
    "cold_transfer_inboxes": ["box_id"],
    "pending_transfer_stock": ["box_id", "original_box_id"],
    "cold_stock_disposition": ["box_id"],
    "interunit_transfer_boxes": ["box_id"],
    "transfer_box_reconciliation": ["actual_box_id", "original_box_id"],
    "interunit_transfer_in_boxes": ["box_id"],
}


def load_sheet():
    """Keyed on the LOT alone — the sheet's company label is an attribute, and for 5 lots
    it disagrees with the DB. Verified safe: 14570 is the only lot present under both
    companies in cold_stocks, and that is a confirmed duplicate, not two distinct piles."""
    ws = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)["Sheet1"]
    agg = {}
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        co, lot = _s(r[COL["company"]]).upper(), _s(r[COL["lot_no"]])
        if not lot or co not in PREFIX:
            continue
        k = norm(lot)
        d = agg.setdefault(k, {"lot": lot, "cart": 0.0, "company": co,
                               "desc": _s(r[COL["item_description"]])})
        try:
            d["cart"] += float(r[COL["cartons"]] or 0)
        except (TypeError, ValueError):
            pass
    for d in agg.values():
        d["cart"] = round(d["cart"])
    return agg


def load_db(conn):
    cur = conn.cursor()
    cold, arrived, boxes = {}, defaultdict(int), {}
    for co, p in PREFIX.items():
        cur.execute(f"""SELECT lot_no::text, count(*),
                          count(*) FILTER (WHERE created_at > %s),
                          count(*) FILTER (WHERE transaction_no = 'RECON22JUL'),
                          max(item_description)
                        FROM {p}_cold_stocks WHERE lot_no IS NOT NULL GROUP BY 1""",
                    (CUTOFF,))
        for lot, n, new, synth, desc in cur.fetchall():
            k = norm(lot)
            if not k:
                continue
            rec = cold.setdefault(k, {"lot": _s(lot), "n": 0, "synth": 0, "desc": desc,
                                      "by_co": defaultdict(int)})
            rec["n"] += n
            rec["synth"] += synth
            rec["by_co"][co] += n
            arrived[k] += new
    # safe-to-delete pool: boxes not referenced by any in-flight movement table.
    # ponytail: refset built in Python — the equivalent NOT IN over a 6-table UNION
    # never finished against 115k cold rows.
    refset = set()
    for t, cols in REF_SOURCES.items():
        for col in cols:
            cur.execute(f"SELECT DISTINCT {col}::text FROM {t} WHERE {col} IS NOT NULL")
            refset.update(b.strip() for (b,) in cur.fetchall() if b and b.strip())
    for co, p in PREFIX.items():
        cur.execute(f"SELECT lot_no::text, box_id FROM {p}_cold_stocks WHERE lot_no IS NOT NULL")
        for lot, box_id in cur.fetchall():
            k = norm(lot)
            if k and (not box_id or box_id.strip() not in refset):
                boxes[k] = boxes.get(k, 0) + 1

    nulls = {}
    for co, p in PREFIX.items():
        cur.execute(f"SELECT count(*) FROM {p}_cold_stocks WHERE lot_no IS NULL")
        nulls[co] = cur.fetchone()[0]

    departed = defaultdict(int)
    cur.execute("""SELECT lot_no::text, count(*)
                   FROM cold_stock_disposition
                   WHERE created_at > %s AND source_table LIKE '%%_cold_stocks'
                     AND reverted IS NOT TRUE
                   GROUP BY 1""", (CUTOFF,))
    for lot, n in cur.fetchall():
        k = norm(lot)
        if k:
            departed[k] += n
    return cold, arrived, departed, boxes, nulls


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=None)
    args = ap.parse_args()

    sheet = load_sheet()
    with psycopg.connect(os.environ["DATABASE_URL"]) as conn:
        cold, arrived, departed, safe, nulls = load_db(conn)

    rows = []
    for lk in sorted(set(sheet) | set(cold)):
        s, c = sheet.get(lk), cold.get(lk)
        cart = s["cart"] if s else 0
        now = c["n"] if c else 0
        dep, arr = departed.get(lk, 0), arrived.get(lk, 0)
        as_of = now + dep - arr
        # a lot absent from the sheet must end at 0 regardless of its movement
        target_now = (cart - dep) if s else 0
        act = target_now - now
        db_cos = sorted(c["by_co"]) if c else []
        wrong_co = bool(s and db_cos and s["company"] not in db_cos)

        if act == 0:
            # already at target — includes lots whose whole gap was today's departure
            kind = "MOVE_COMPANY" if wrong_co else "OK"
        elif not s:
            kind = "DELETE_ALL"
        elif not c:
            kind = "INSERT_NEW_LOT"
        elif act < 0:
            kind = "DELETE"
        else:
            kind = "INSERT"

        rows.append({
            "lot": (s or c)["lot"], "lot_key": lk,
            "sheet_company": s["company"] if s else "",
            "db_company": "+".join(db_cos),
            "company_mismatch": "YES" if wrong_co else "",
            "sheet_cartons": cart if s else None,
            "cold_now": now, "departed_after_29jul": dep, "arrived_after_29jul": arr,
            "cold_as_of_29jul": as_of,
            "as_of_delta": (as_of - cart) if s else None,
            "target_now": target_now if s else 0,
            "action": kind, "action_boxes": act,
            "safe_to_delete": safe.get(lk, 0),
            "synthetic_recon22jul": c["synth"] if c else 0,
            "item_description": (s or c).get("desc") or "",
        })

    tot_sheet = sum(r["sheet_cartons"] or 0 for r in rows)
    tot_now = sum(r["cold_now"] for r in rows)
    tot_target = sum(r["target_now"] for r in rows)
    print(f"sheet total          {tot_sheet:>9,d} cartons  (29-Jul-2026 as-on)")
    print(f"cold_stocks now      {tot_now:>9,d} boxes")
    print(f"departed after 29Jul {sum(departed.values()):>9,d} boxes  (backed out)")
    print(f"arrived  after 29Jul {sum(arrived.values()):>9,d} boxes  (backed out)")
    print(f"target cold_stocks   {tot_target:>9,d} boxes  = sheet - post-snapshot departures")
    print(f"net change required  {tot_target - tot_now:>+9,d} boxes")
    print(f"NOT counted above: cold_stocks rows with NO lot_no at all (the sheet lists "
          f"no such stock): {sum(nulls.values()):>6,d} boxes "
          f"({', '.join(f'{k} {v:,}' for k, v in sorted(nulls.items()))})\n")

    print("=== AS-OF-29-JUL accuracy (was the DB right at the snapshot?) ===")
    ss = [r for r in rows if r["sheet_cartons"] is not None]
    print(f"  lots on sheet                : {len(ss)}")
    print(f"  agreeing as of 29-Jul        : {sum(1 for r in ss if r['as_of_delta'] == 0)}")
    print(f"  overstated (DB > sheet)      : {sum(1 for r in ss if r['as_of_delta'] > 0)} "
          f"(+{sum(r['as_of_delta'] for r in ss if r['as_of_delta'] > 0):,} boxes)")
    print(f"  understated (DB < sheet)     : {sum(1 for r in ss if r['as_of_delta'] < 0)} "
          f"({sum(r['as_of_delta'] for r in ss if r['as_of_delta'] < 0):,} boxes)")
    fixed = [r for r in ss if r["departed_after_29jul"] or r["arrived_after_29jul"]]
    print(f"\n  lots where backing out today's movement CHANGED the verdict:")
    for r in sorted(fixed, key=lambda r: -abs(r["departed_after_29jul"] + r["arrived_after_29jul"])):
        raw = r["cold_now"] - r["sheet_cartons"]
        print(f"   {r['lot']:>8s} sheet={r['sheet_cartons']:>5d} "
              f"now={r['cold_now']:>5d} (raw delta {raw:+5d})  "
              f"dep={r['departed_after_29jul']:>4d} arr={r['arrived_after_29jul']:>5d} "
              f"-> as_of={r['cold_as_of_29jul']:>5d} (true delta {r['as_of_delta']:+5d})")

    print("\n=== PLAN by action ===")
    agg = defaultdict(lambda: [0, 0])
    for r in rows:
        a = agg[r["action"]]
        a[0] += 1
        a[1] += r["action_boxes"]
    for kind, (n, b) in sorted(agg.items(), key=lambda x: -abs(x[1][1])):
        print(f"  {kind:<16s} lots={n:<5d} boxes={b:>+9,d}")

    for kind in ("DELETE", "DELETE_ALL", "INSERT", "INSERT_NEW_LOT", "MOVE_COMPANY"):
        sub = [r for r in rows if r["action"] == kind]
        if not sub:
            continue
        print(f"\n-- {kind} ({len(sub)} lots, {sum(r['action_boxes'] for r in sub):+,} boxes) --")
        print(f"   {'LOT':>9s}{'SHEET':>7s}{'NOW':>7s}{'AS_OF':>7s}{'ACTION':>8s}"
              f"{'SAFE':>7s}{'SYNTH':>7s}  {'CO(sheet/db)':<14s}DESCRIPTION")
        for r in sorted(sub, key=lambda r: -abs(r["action_boxes"])):
            short = ("!! only %d safe" % r["safe_to_delete"]
                     if r["action_boxes"] < 0 and r["safe_to_delete"] < -r["action_boxes"] else "")
            co = f"{r['sheet_company'] or '-'}/{r['db_company'] or '-'}"
            print(f"   {r['lot']:>9s}"
                  f"{(r['sheet_cartons'] if r['sheet_cartons'] is not None else 0):>7d}"
                  f"{r['cold_now']:>7d}{r['cold_as_of_29jul']:>7d}{r['action_boxes']:>+8d}"
                  f"{r['safe_to_delete']:>7d}{r['synthetic_recon22jul']:>7d}  "
                  f"{co:<14s}{(r['item_description'] or '')[:30]:<30s}{short}")

    blocked = [r for r in rows if r["action_boxes"] < 0
               and r["safe_to_delete"] < -r["action_boxes"]]
    print(f"\n=== deletions blocked by in-flight boxes: {len(blocked)} lots, "
          f"{sum(-r['action_boxes'] - r['safe_to_delete'] for r in blocked):,} boxes short ===")
    for r in blocked:
        print(f"   {r['lot']:>8s} need to delete "
              f"{-r['action_boxes']:,} but only {r['safe_to_delete']:,} are unreferenced")

    if args.xlsx:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "plan_29jul"
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append(list(r.values()))
        ws.freeze_panes = "A2"
        wb.save(args.xlsx)
        print(f"\nwrote {args.xlsx}")


if __name__ == "__main__":
    main()
