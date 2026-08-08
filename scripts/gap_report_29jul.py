"""
READ-ONLY GAP REPORT. Every lot in the 29-Jul-2026 closing sheet, its carton quantity,
and its row count in EVERY lot-bearing IMS table — with the gaps computed.

The sheet is authoritative for the physical count. Tables are grouped by what they mean:

  COLD      current physical position ({p}_cold_stocks). A box is present IFF a live row
            exists; every departure DELETEs the row. This is the number the sheet grades.
  INWARD    how the stock was digitized on the way in (bulk entry / v2 / legacy boxes and
            their article headers). A lot with sheet cartons but no inward rows was never
            entered into the system at all.
  OUTBOUND  in-flight and completed departures (interunit transfer out/in, pending stock,
            disposition, cold transfer in, reconciliation).
  CONSUMED  jobwork, RTV, direct-out, floor.

Gaps reported:
  gap_vs_cold   = sheet_cartons - cold_total      (+ = sheet holds more than cold_stocks)
  gap_vs_inward = sheet_cartons - inward_boxes_max(+ = never digitized on the way in)

Article tables are counted as ROWS, never summed as quantities: their box_count /
quantity_units is TRANSACTION-grain (one row can cover many lots) and differencing it
against a per-lot carton figure manufactures fake gaps.

Run: python scripts/gap_report_29jul.py [--xlsx OUT.xlsx] [--all]
"""
from __future__ import annotations
import os
import sys
import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl
import psycopg
from dotenv import load_dotenv

sys.path.insert(0, str(Path(__file__).resolve().parent))
from report_lots_in_closing_29jul import norm, _s, COL, HEADER_ROW, PREFIX

ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
EXCEL_PATH = ROOT / "data" / "Cold Storage Stock Details 29th July Closing.xlsx"

# group -> [(column label, table, lot column, per_company)]
GROUPS = {
    "COLD": [
        ("cold", "{p}_cold_stocks", "lot_no", True),
    ],
    "INWARD": [
        ("bulk_boxes", "{p}_bulk_entry_boxes", "lot_number", True),
        ("boxes_v2", "{p}_boxes_v2", "lot_number", True),
        ("boxes", "{p}_boxes", "lot_number", True),
        ("bulk_articles", "{p}_bulk_entry_articles", "lot_number", True),
        ("articles_v2", "{p}_articles_v2", "lot_number", True),
        ("articles", "{p}_articles", "lot_number", True),
    ],
    "OUTBOUND": [
        ("itf_out", "interunit_transfer_boxes", "lot_number", False),
        ("itf_in", "interunit_transfer_in_boxes", "lot_number", False),
        ("itf_lines", "interunit_transfers_lines", "lot_number", False),
        ("pending", "pending_transfer_stock", "lot_no", False),
        ("disposition", "cold_stock_disposition", "lot_no", False),
        ("cold_in", "cold_transfer_inboxes", "lot_no", False),
        ("box_recon", "transfer_box_reconciliation", "lot_no", False),
    ],
    "CONSUMED": [
        ("jobwork_out", "jb_materialout_lines", "lot_number", False),
        ("jobwork_in", "jb_inward_boxes", "lot_no", False),
        ("rtv_boxes", "{p}_rtv_boxes", "lot_number", True),
        ("rtv_lines", "{p}_rtv_lines", "lot_number", True),
        ("direct_out", "{p}_cold_storage_direct_out", "lot_no", True),
        ("floor", "floor_inventory", "lot_number", False),
    ],
}
# inward tables counted in box terms (article tables are headers, not boxes)
INWARD_BOX_COLS = ["bulk_boxes", "boxes_v2", "boxes"]
ALL_COLS = [c for g in GROUPS.values() for c, _, _, _ in g]


def load_sheet():
    ws = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)["Sheet1"]
    agg, order = {}, []
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        co, lot = _s(r[COL["company"]]).upper(), _s(r[COL["lot_no"]])
        if not lot or co not in PREFIX:
            continue
        k = norm(lot)
        if k not in agg:
            agg[k] = {"lot": lot, "company": co, "cart": 0.0, "rows": 0,
                      "desc": _s(r[COL["item_description"]])}
            order.append(k)
        agg[k]["cart"] += float(r[COL["cartons"]] or 0)
        agg[k]["rows"] += 1
    for d in agg.values():
        d["cart"] = round(d["cart"])
    return agg, order


def load_counts(conn):
    """col -> {lot_key: {'CFPL': n, 'CDPL': n, 'ALL': n}}"""
    out = {c: defaultdict(lambda: defaultdict(int)) for c in ALL_COLS}
    cur = conn.cursor()
    for spec in (s for g in GROUPS.values() for s in g):
        col, tmpl, lotcol, per_co = spec
        targets = ([(co, tmpl.format(p=p)) for co, p in PREFIX.items()]
                   if per_co else [("ALL", tmpl)])
        for co, tbl in targets:
            cur.execute(f'SELECT {lotcol}::text, count(*) FROM "{tbl}" '
                        f"WHERE {lotcol} IS NOT NULL GROUP BY 1")
            for lot, n in cur.fetchall():
                k = norm(lot)
                if not k:
                    continue
                out[col][k][co] += n
                if per_co:
                    out[col][k]["ALL"] += n
    # extra breakdowns that change the meaning of a count
    extra = {}
    cur.execute("""SELECT lot_no::text, status, count(*) FROM pending_transfer_stock
                   WHERE lot_no IS NOT NULL GROUP BY 1,2""")
    p_intransit, p_other = defaultdict(int), defaultdict(int)
    for lot, st, n in cur.fetchall():
        k = norm(lot)
        if k:
            (p_intransit if (st or "").lower() == "in transit" else p_other)[k] += n
    cur.execute("""SELECT lot_no::text, reverted, count(*) FROM cold_stock_disposition
                   WHERE lot_no IS NOT NULL GROUP BY 1,2""")
    d_active, d_rev = defaultdict(int), defaultdict(int)
    for lot, rev, n in cur.fetchall():
        k = norm(lot)
        if k:
            (d_rev if rev else d_active)[k] += n
    # synthetic vs real split of the cold position. A lot holding BOTH is double-counted:
    # the RECON22JUL placeholder was never removed when the real boxes were scanned in.
    synth, real = defaultdict(int), defaultdict(int)
    for co, p in PREFIX.items():
        cur.execute(f"""SELECT lot_no::text,
                          count(*) FILTER (WHERE transaction_no = 'RECON22JUL'),
                          count(*) FILTER (WHERE transaction_no <> 'RECON22JUL'
                                              OR transaction_no IS NULL)
                        FROM {p}_cold_stocks WHERE lot_no IS NOT NULL GROUP BY 1""")
        for lot, s, r in cur.fetchall():
            k = norm(lot)
            if k:
                synth[k] += s
                real[k] += r
    extra = {"pending_in_transit": p_intransit, "pending_other": p_other,
             "disposition_active": d_active, "disposition_reverted": d_rev,
             "cold_synthetic": synth, "cold_real": real}
    return out, extra


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=None)
    ap.add_argument("--all", action="store_true", help="print every lot, not just gaps")
    args = ap.parse_args()

    sheet, order = load_sheet()
    with psycopg.connect(os.environ["DATABASE_URL"]) as conn:
        cnt, extra = load_counts(conn)

    rows = []
    for k in order:
        s = sheet[k]
        r = {"lot": s["lot"], "sheet_company": s["company"], "sheet_cartons": s["cart"],
             "sheet_rows": s["rows"]}
        for col in ALL_COLS:
            c = cnt[col].get(k, {})
            if col == "cold":
                r["cold_cfpl"] = c.get("CFPL", 0)
                r["cold_cdpl"] = c.get("CDPL", 0)
                r["cold_total"] = c.get("ALL", 0)
            else:
                r[col] = c.get("ALL", 0)
        for name, d in extra.items():
            r[name] = d.get(k, 0)
        inward_max = max(r[c] for c in INWARD_BOX_COLS)
        r["inward_boxes_max"] = inward_max
        r["gap_vs_cold"] = s["cart"] - r["cold_total"]
        r["gap_vs_inward"] = s["cart"] - inward_max
        r["cold_company_mismatch"] = (
            "YES" if r["cold_total"] and not c_has(cnt, k, s["company"]) else "")
        r["verdict"] = verdict(r)
        r["item_description"] = s["desc"]
        rows.append(r)

    print(f"29-Jul-2026 closing: {len(rows)} lots, "
          f"{sum(r['sheet_cartons'] for r in rows):,} cartons\n")

    print("=== GAP SUMMARY ===")
    agg = defaultdict(lambda: [0, 0, 0])
    for r in rows:
        a = agg[r["verdict"]]
        a[0] += 1
        a[1] += r["sheet_cartons"]
        a[2] += r["gap_vs_cold"]
    for v, (n, cart, gap) in sorted(agg.items(), key=lambda x: -abs(x[1][2])):
        print(f"  {v:<26s} lots={n:<5d} sheet={cart:>8,d}  gap_vs_cold={gap:>+8,d}")
    print(f"\n  total sheet cartons  {sum(r['sheet_cartons'] for r in rows):>9,d}")
    print(f"  total cold_stocks    {sum(r['cold_total'] for r in rows):>9,d}")
    print(f"  net gap_vs_cold      {sum(r['gap_vs_cold'] for r in rows):>+9,d}")
    # NOT "never digitized" — most of these hold RECON22JUL synthetic cold rows.
    noin = [r for r in rows if r["inward_boxes_max"] == 0]
    never = [r for r in rows if r["verdict"] == "NEVER_IN_ANY_TABLE"]
    dbl = [r for r in rows if r["cold_synthetic"] and r["cold_real"]]
    print(f"  no inward-box rows   {len(noin):>9,d}  ({sum(r['sheet_cartons'] for r in noin):,} "
          f"cartons) - most DO hold synthetic cold rows, so this is not 'missing'")
    print(f"  absent from ALL IMS  {len(never):>9,d}  ({sum(r['sheet_cartons'] for r in never):,} cartons)")
    print(f"  cold_stocks: {sum(r['cold_synthetic'] for r in rows):,} synthetic RECON22JUL "
          f"+ {sum(r['cold_real'] for r in rows):,} real")
    print(f"  DOUBLE-COUNTED (lot holds synthetic AND real): {len(dbl)} lots, "
          f"{sum(min(r['cold_synthetic'], r['cold_real']) for r in dbl):,} surplus boxes")
    for r in dbl:
        print(f"     lot {r['lot']:>8s} synthetic={r['cold_synthetic']} real={r['cold_real']} "
              f"sheet={r['sheet_cartons']}")

    show = rows if args.all else [r for r in rows if r["gap_vs_cold"] or r["verdict"] != "MATCH"]
    show.sort(key=lambda r: (-abs(r["gap_vs_cold"]), -r["sheet_cartons"]))
    print(f"\n=== PER-LOT GAPS ({len(show)} lots) ===")
    hdr = (f"{'LOT':>9s}{'CO':>6s}{'SHEET':>7s}{'COLD':>7s}{'GAP':>7s}"
           f"{'SYNTH':>7s}{'REAL':>7s}{'INWARD':>7s}{'OUT':>6s}{'PEND':>6s}{'DISP':>6s}{'JW':>5s}{'RTV':>5s}"
           f"  {'VERDICT':<24s}DESCRIPTION")
    print(hdr)
    print("-" * len(hdr))
    for r in show:
        print(f"{r['lot']:>9s}{r['sheet_company']:>6s}{r['sheet_cartons']:>7d}"
              f"{r['cold_total']:>7d}{r['gap_vs_cold']:>+7d}"
              f"{r['cold_synthetic']:>7d}{r['cold_real']:>7d}"
              f"{r['inward_boxes_max']:>7d}{r['itf_out']:>6d}"
              f"{r['pending_in_transit']:>6d}{r['disposition_active']:>6d}"
              f"{r['jobwork_out']:>5d}{r['rtv_boxes']:>5d}  "
              f"{r['verdict']:<24s}{(r['item_description'] or '')[:28]}")

    if args.xlsx:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "gap_29jul"
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append(list(r.values()))
        ws.freeze_panes = "D2"
        wb.save(args.xlsx)
        print(f"\nwrote {args.xlsx}")


def c_has(cnt, k, company):
    return cnt["cold"].get(k, {}).get(company, 0) > 0


def verdict(r):
    """Why the sheet and the DB differ, using the other tables as evidence."""
    gap = r["gap_vs_cold"]
    if gap == 0:
        return "MATCH" if not r["cold_company_mismatch"] else "MATCH_WRONG_COMPANY"
    if r["cold_total"] == 0:
        if r["inward_boxes_max"] == 0 and not any(
                r[c] for c in ("itf_out", "itf_in", "pending", "disposition",
                               "jobwork_out", "rtv_boxes", "floor")):
            return "NEVER_IN_ANY_TABLE"
        if r["cold_company_mismatch"]:
            return "COLD_UNDER_OTHER_CO"
        if r["disposition_active"] or r["itf_out"] or r["pending_in_transit"]:
            return "DEPARTED_NOT_IN_COLD"
        return "IN_DB_NOT_IN_COLD"
    if gap > 0:
        return "COLD_SHORT_OF_SHEET"
    return "COLD_EXCEEDS_SHEET"


if __name__ == "__main__":
    main()
