"""
READ-ONLY. List every lot in 'Cold Storage Stock Details 29th July Closing.xlsx'
and say where it exists in the DB: cold_stocks (the physical cold position),
inward/box tables, and movement tables.

Lot keys are matched on a normalised form (strip, drop a trailing '.0', drop
leading zeros) because the sheet carries lots as numbers and the DB as varchar.

Run: python scripts/report_lots_in_closing_29jul.py [--xlsx OUT.xlsx]
"""
from __future__ import annotations
import os
import re
import sys
import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl
import psycopg
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
EXCEL_PATH = ROOT / "data" / "Cold Storage Stock Details 29th July Closing.xlsx"

HEADER_ROW = 6  # 1-based; data starts at row 7
COL = {"inward_no": 2, "cold_item_mark": 4, "vakkal": 5, "lot_no": 6, "cartons": 7,
       "item_mark": 12, "item_description": 14, "company": 15}

PREFIX = {"CFPL": "cfpl", "CDPL": "cdpl"}

# per-company tables: (label, table_suffix, lot_column)
COMPANY_TABLES = [
    ("cold", "cold_stocks", "lot_no"),
    ("boxes_v2", "boxes_v2", "lot_number"),
    ("boxes", "boxes", "lot_number"),
    ("bulk_boxes", "bulk_entry_boxes", "lot_number"),
    ("articles_v2", "articles_v2", "lot_number"),
    ("articles", "articles", "lot_number"),
    ("bulk_articles", "bulk_entry_articles", "lot_number"),
    ("rtv_boxes", "rtv_boxes", "lot_number"),
    ("direct_out", "cold_storage_direct_out", "lot_no"),
]

# company-agnostic movement tables: (label, table, lot_column)
SHARED_TABLES = [
    ("itf_out", "interunit_transfer_boxes", "lot_number"),
    ("itf_in", "interunit_transfer_in_boxes", "lot_number"),
    ("itf_lines", "interunit_transfers_lines", "lot_number"),
    ("pending", "pending_transfer_stock", "lot_no"),
    ("disposition", "cold_stock_disposition", "lot_no"),
    ("cold_in", "cold_transfer_inboxes", "lot_no"),
    ("jobwork_out", "jb_materialout_lines", "lot_number"),
    ("jobwork_in", "jb_inward_boxes", "lot_no"),
    ("floor", "floor_inventory", "lot_number"),
]


def norm(v) -> str | None:
    """Normalised lot key: '079952.0' -> '79952'. Non-numeric lots keep their text."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    if not s:
        return None
    s = re.sub(r"\.0+$", "", s)
    return s.lstrip("0") or "0" if s.lstrip("0").isdigit() or s.isdigit() else s.upper()


def _s(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def load_excel():
    """[(company, lot_raw, lot_key, cartons, rows, desc...)] aggregated per company+lot."""
    wb = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)
    ws = wb["Sheet1"]
    agg, order, skipped = {}, [], []
    for rn, r in enumerate(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True),
                           start=HEADER_ROW + 1):
        company, lot = _s(r[COL["company"]]).upper(), _s(r[COL["lot_no"]])
        if not lot or company not in PREFIX:
            if any(x is not None for x in r[:22]):
                skipped.append((rn, company, lot))
            continue
        key = (company, norm(lot))
        if key not in agg:
            agg[key] = {"company": company, "lot_raw": lot, "lot_key": key[1],
                        "cart": 0.0, "rows": 0, "raws": set(),
                        "desc": _s(r[COL["item_description"]]),
                        "mark": _s(r[COL["cold_item_mark"]]),
                        "item_mark": _s(r[COL["item_mark"]]),
                        "vakkal": _s(r[COL["vakkal"]]),
                        "inward_no": set()}
            order.append(key)
        d = agg[key]
        try:
            d["cart"] += float(r[COL["cartons"]] or 0)
        except (TypeError, ValueError):
            pass
        d["rows"] += 1
        d["raws"].add(lot)
        if _s(r[COL["inward_no"]]):
            d["inward_no"].add(_s(r[COL["inward_no"]]))
    return [agg[k] for k in order], skipped


def load_presence(conn):
    """label -> {company: {lot_key: count}}; shared tables land under company '*'."""
    pres = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    cur = conn.cursor()
    for company, p in PREFIX.items():
        for label, suffix, lotcol in COMPANY_TABLES:
            tbl = f"{p}_{suffix}"
            cur.execute(
                f'SELECT {lotcol}::text, count(*) FROM "{tbl}" '
                f"WHERE {lotcol} IS NOT NULL GROUP BY 1")
            for lot, n in cur.fetchall():
                k = norm(lot)
                if k:
                    pres[label][company][k] += n
    for label, tbl, lotcol in SHARED_TABLES:
        cur.execute(
            f'SELECT {lotcol}::text, count(*) FROM "{tbl}" '
            f"WHERE {lotcol} IS NOT NULL GROUP BY 1")
        for lot, n in cur.fetchall():
            k = norm(lot)
            if k:
                pres[label]["*"][k] += n
    return pres


def recon_marked(conn):
    """lot_key -> synthetic RECON22JUL box count (last reconciliation's inserts)."""
    out = defaultdict(int)
    cur = conn.cursor()
    for company, p in PREFIX.items():
        cur.execute(f'SELECT lot_no::text, count(*) FROM "{p}_cold_stocks" '
                    "WHERE transaction_no = 'RECON22JUL' GROUP BY 1")
        for lot, n in cur.fetchall():
            k = norm(lot)
            if k:
                out[(company, k)] += n
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=None)
    args = ap.parse_args()

    lots, skipped = load_excel()
    with psycopg.connect(os.environ["DATABASE_URL"]) as conn:
        pres = load_presence(conn)
        marked = recon_marked(conn)

    inward_labels = ["boxes_v2", "boxes", "bulk_boxes",
                     "articles_v2", "articles", "bulk_articles"]
    move_labels = [lbl for lbl, _, _ in SHARED_TABLES] + ["rtv_boxes", "direct_out"]

    rows = []
    for d in lots:
        c, k = d["company"], d["lot_key"]
        other = "CDPL" if c == "CFPL" else "CFPL"
        cold = pres["cold"][c].get(k, 0)
        cold_other = pres["cold"][other].get(k, 0)
        inward = {lbl: pres[lbl][c].get(k, 0) for lbl in inward_labels}
        inward_other = sum(pres[lbl][other].get(k, 0) for lbl in inward_labels)
        move = {lbl: (pres[lbl]["*"].get(k, 0) if lbl in pres and "*" in pres[lbl]
                      else pres[lbl][c].get(k, 0)) for lbl in move_labels}
        in_inward = any(inward.values())
        in_move = any(move.values())

        if cold:
            status = "IN_COLD"
        elif cold_other:
            status = "COLD_WRONG_COMPANY"
        elif in_inward:
            status = "IN_DB_NOT_COLD"
        elif inward_other:
            status = "DB_WRONG_COMPANY"
        elif in_move:
            status = "MOVEMENT_ONLY"
        else:
            status = "NOT_IN_DB"

        rows.append({
            "company": c, "lot": d["lot_raw"], "lot_key": k,
            "sheet_cartons": round(d["cart"], 3), "sheet_rows": d["rows"],
            "cold_boxes": cold, "delta": cold - round(d["cart"]),
            "recon22jul_synthetic": marked.get((c, k), 0),
            "status": status,
            "cold_other_company": cold_other,
            "inward_other_company": inward_other,
            **{f"in_{lbl}": inward[lbl] for lbl in inward_labels},
            **{f"mv_{lbl}": move[lbl] for lbl in move_labels},
            "inward_no": ",".join(sorted(d["inward_no"])),
            "item_description": d["desc"], "cold_item_mark": d["mark"],
            "item_mark": d["item_mark"], "vakkal": d["vakkal"],
            "lot_raw_variants": ",".join(sorted(d["raws"])) if len(d["raws"]) > 1 else "",
        })

    rows.sort(key=lambda r: (r["company"], r["status"] != "IN_COLD",
                             -r["sheet_cartons"]))

    # ---- console listing ----
    hdr = (f"{'CO':5s}{'LOT':>10s}{'SHEET':>8s}{'COLD':>8s}{'DELTA':>8s}"
           f"{'SYNTH':>7s}  {'STATUS':<19s}{'INWARD(v2/b/bulk)':>19s}  DESCRIPTION")
    print(f"29-Jul-2026 closing: {len(rows)} distinct (company, lot) over "
          f"{sum(r['sheet_rows'] for r in rows)} sheet rows, "
          f"{sum(r['sheet_cartons'] for r in rows):,.0f} cartons\n")
    print(hdr)
    print("-" * len(hdr))
    for r in rows:
        inw = f"{r['in_boxes_v2']}/{r['in_boxes']}/{r['in_bulk_boxes']}"
        print(f"{r['company']:5s}{r['lot']:>10s}{r['sheet_cartons']:>8.0f}"
              f"{r['cold_boxes']:>8d}{r['delta']:>8d}"
              f"{r['recon22jul_synthetic']:>7d}  {r['status']:<19s}{inw:>19s}  "
              f"{(r['item_description'] or '')[:26]:<26s} {r['cold_item_mark'][:28]}")

    print("\n--- summary by status ---")
    by = defaultdict(lambda: [0, 0, 0])
    for r in rows:
        s = by[r["status"]]
        s[0] += 1
        s[1] += r["sheet_cartons"]
        s[2] += r["cold_boxes"]
    for st, (n, cart, cold) in sorted(by.items(), key=lambda x: -x[1][0]):
        print(f"{st:<20s} lots={n:<5d} sheet_cartons={cart:>10,.0f} cold_boxes={cold:>10,d}")
    matched = [r for r in rows if r["status"] == "IN_COLD" and r["delta"] == 0]
    print(f"\nexact carton match (sheet == cold_stocks): {len(matched)} lots")
    print(f"synthetic RECON22JUL boxes still standing in these lots: "
          f"{sum(r['recon22jul_synthetic'] for r in rows):,}")
    if skipped:
        print(f"\nsheet rows skipped (no lot / unknown company): {len(skipped)}")
        for rn, co, lot in skipped[:20]:
            print(f"  row {rn}: company={co!r} lot={lot!r}")

    if args.xlsx:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "lots_29jul"
        ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append(list(r.values()))
        ws.freeze_panes = "A2"
        wb.save(args.xlsx)
        print(f"\nwrote {args.xlsx}")


if __name__ == "__main__":
    main()
