"""Relabel ACTIVE double-use boxes so blocked transfers can receive.

Target: an OUT box in a Dispatch/Partial transfer B whose (box_id, txn, lot) is
held In-Transit by ANOTHER transfer (so B 409s on receive) and B does NOT itself
own a pending row for it.

Action per target:
  1. new_box_id = "<orig>-R<transfer_out_id>"  (unique, SAME transaction_no)
  2. UPDATE interunit_transfer_boxes -> new_box_id (all of B's rows for that box)
  3. INSERT a TRACKING pending_transfer_stock row for B (status In Transit) with the
     new box_id. source_table='' and destination_table='' => pick/restore touch NO
     real stock (cold/boxes_v2 untouched; box already deducted by the owner).

So B becomes receivable, records stay (history safe), inventory tables unchanged.
Old completed double-dispatches are NOT touched.

DRY-RUN by default; --apply commits. Always writes/updates an Excel sheet listing
what was (or would be) relabeled.
"""
import os, io, sys, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
from sqlalchemy import create_engine, text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def _exists_box_id(c, box_id):
    for tbl in ("interunit_transfer_boxes", "pending_transfer_stock"):
        if c.execute(text(f"SELECT 1 FROM {tbl} WHERE box_id = :b LIMIT 1"), {"b": box_id}).fetchone():
            return True
    return False

raw = os.environ["DATABASE_URL"]
DB = raw.replace("postgresql://", "postgresql+psycopg://", 1) if raw.startswith("postgresql://") else raw
APPLY = "--apply" in sys.argv
eng = create_engine(DB)
XLSX = "Interunit_Transfer_Reconciliation.xlsx"
SHEET = "Relabeled Active Double-Use"
COLD_SITES = {"cold storage", "rishi cold", "savla d-39 cold", "savla d-514 cold"}
def is_cold(s): return (s or "").strip().lower() in COLD_SITES

TARGETS = text("""
    SELECT DISTINCT ob.header_id, h.challan_no, h.from_site, h.to_site, h.status,
           h.created_by, h.created_ts,
           ob.box_id, ob.transaction_no, ob.lot_number, ob.article,
           ob.net_weight, ob.gross_weight,
           (SELECT p.transfer_out_id FROM pending_transfer_stock p
              WHERE p.box_id=ob.box_id AND p.transaction_no=ob.transaction_no
                AND TRIM(COALESCE(p.lot_no,''))=TRIM(COALESCE(ob.lot_number,''))
                AND p.status='In Transit' AND p.transfer_out_id<>ob.header_id LIMIT 1) AS held_by
    FROM interunit_transfer_boxes ob
    JOIN interunit_transfers_header h ON h.id = ob.header_id
    WHERE ob.box_id IS NOT NULL AND ob.box_id <> ''
      AND h.status IN ('Dispatch','Partial')
      AND EXISTS (SELECT 1 FROM pending_transfer_stock p
            WHERE p.box_id=ob.box_id AND p.transaction_no=ob.transaction_no
              AND TRIM(COALESCE(p.lot_no,''))=TRIM(COALESCE(ob.lot_number,''))
              AND p.status='In Transit' AND p.transfer_out_id<>ob.header_id)
      AND NOT EXISTS (SELECT 1 FROM pending_transfer_stock p2
            WHERE p2.box_id=ob.box_id AND p2.transaction_no=ob.transaction_no
              AND p2.status='In Transit' AND p2.transfer_out_id=ob.header_id)
    ORDER BY ob.header_id, ob.box_id
""")

INSERT_PENDING = text("""
    INSERT INTO pending_transfer_stock
        (transfer_type, transfer_out_id, transfer_out_challan_no, box_id, transaction_no,
         from_site, to_site, from_storage_type, to_storage_type,
         source_table, source_row_id, destination_table,
         item_description, lot_no, weight_kg, no_of_cartons,
         gross_weight, net_weight, status, dispatched_at, dispatched_by)
    VALUES
        ('INTERUNIT', :tid, :challan, :box_id, :txn,
         :from_site, :to_site, :fst, :tst,
         '', NULL, '',
         :item, :lot, :wkg, 1,
         :gross, :net, 'In Transit', :ts, :by)
    ON CONFLICT (box_id, transaction_no) DO NOTHING
""")

out_rows = []
with eng.begin() as c:
    targets = c.execute(TARGETS).fetchall()
    print(f"Active double-use boxes to relabel: {len(targets)}  APPLY={APPLY}\n")
    # Box-ID format mirrors the inward module: {last 8 digits of epoch ms}-{n}
    base = str(int(time.time() * 1000))[-8:]
    seq = 0
    for r in targets:
        m = r._mapping
        seq += 1
        new_id = f"{base}-{seq}"
        while _exists_box_id(c, new_id):   # guarantee uniqueness
            seq += 1
            new_id = f"{base}-{seq}"
        out_rows.append([
            m["header_id"], m["challan_no"], m["article"], m["lot_number"] or "",
            m["transaction_no"], m["box_id"], new_id, m["held_by"],
            float(m["net_weight"] or 0), "DONE" if APPLY else "PLANNED",
        ])
        if APPLY:
            c.execute(text("""
                UPDATE interunit_transfer_boxes SET box_id=:new
                WHERE header_id=:h AND box_id=:old AND transaction_no=:txn
                  AND COALESCE(lot_number,'')=COALESCE(:lot,'')
            """), {"new": new_id, "h": m["header_id"], "old": m["box_id"],
                   "txn": m["transaction_no"], "lot": m["lot_number"]})
            c.execute(INSERT_PENDING, {
                "tid": m["header_id"], "challan": m["challan_no"],
                "box_id": new_id, "txn": m["transaction_no"],
                "from_site": m["from_site"], "to_site": m["to_site"],
                "fst": "cold" if is_cold(m["from_site"]) else "warehouse",
                "tst": "cold" if is_cold(m["to_site"]) else "warehouse",
                "item": m["article"], "lot": m["lot_number"],
                "wkg": float(m["net_weight"] or 0),
                "gross": m["gross_weight"], "net": m["net_weight"],
                "ts": m["created_ts"], "by": m["created_by"] or "relabel_fix",
            })

    # per-transfer summary print
    by_t = {}
    for row in out_rows:
        by_t.setdefault((row[0], row[1]), 0)
        by_t[(row[0], row[1])] += 1
    for (tid, ch), n in sorted(by_t.items(), key=lambda kv: -kv[1]):
        print(f"  transfer {tid} {ch}: relabel {n} box(es)")
    print(f"\nTOTAL relabeled: {len(out_rows)} boxes across {len(by_t)} transfers.")
    if not APPLY:
        print("DRY-RUN only — no DB writes. (Excel preview written.)")

# Write Excel sheet
wb = openpyxl.load_workbook(XLSX)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)
headers = ["Transfer ID", "Challan No", "Article", "Lot No", "Transaction No",
           "Old Box ID", "New Box ID", "Held In-Transit By (transfer)", "Net Wt (kg)", "Status"]
ws.append(headers)
for col in range(1, len(headers)+1):
    cc = ws.cell(row=1, column=col); cc.fill = PatternFill("solid", fgColor="1F4E78")
    cc.font = Font(bold=True, color="FFFFFF"); cc.alignment = Alignment(horizontal="center", wrap_text=True)
for row in out_rows:
    ws.append(row)
ws.freeze_panes = "A2"
for i, w in enumerate([12, 22, 30, 10, 20, 18, 22, 16, 11, 10], start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
try:
    wb.save(XLSX)
    saved_to = XLSX
except PermissionError:
    alt = "Interunit_Transfer_Reconciliation_RELABEL.xlsx"
    wb.save(alt)
    saved_to = f"{alt}  (main file was OPEN/locked — close Excel & re-run to merge into main)"
print(f"Excel sheet '{SHEET}' written to: {saved_to} — {len(out_rows)} rows.")
print("Done.")
