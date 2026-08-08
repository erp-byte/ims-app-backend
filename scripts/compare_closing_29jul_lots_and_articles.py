"""
READ-ONLY. Two-way comparison of the 29-Jul-2026 Savla closing sheet against the DB:

  1. LOTS  — union of sheet lots and cfpl_/cdpl_cold_stocks lots, both directions
             (sheet-only, DB-only, both + carton delta).
  2. ARTICLES — per (company, lot), the sheet's article attributes vs what the DB
             holds, on two independent sources:
               a) cold_stocks   (item_description / group_name / item_subgroup /
                                 item_mark / vakkal / cold_item_mark)
               b) the article headers ({p}_articles, {p}_articles_v2,
                                 {p}_bulk_entry_articles) + their box_count
             plus whether the sheet's article name is a known SKU
             ({p}sku.item_description / all_sku.particulars).

Comparison is on a normalised form (casefold, collapse whitespace, drop
punctuation/non-ascii) so 'Medjoul-Jumbo' == 'MEDJOUL JUMBO'.

Run: python scripts/compare_closing_29jul_lots_and_articles.py [--xlsx OUT.xlsx]
"""
from __future__ import annotations
import os
import re
import sys
import argparse
from collections import defaultdict
from pathlib import Path

import openpyxl
import psycopg
from dotenv import load_dotenv

sys.path.insert(0, str(Path(__file__).resolve().parent))
from report_lots_in_closing_29jul import norm, _s, COL, HEADER_ROW, PREFIX  # reuse

ROOT = Path(__file__).resolve().parents[1]
load_dotenv(ROOT / ".env")
EXCEL_PATH = ROOT / "data" / "Cold Storage Stock Details 29th July Closing.xlsx"

# sheet column -> logical article attribute
SHEET_ATTRS = {
    "article": COL["item_description"],       # col 14 'Tally Name'
    "group": 10,                              # 'Group Name*'
    "subgroup": 11,                           # 'Sub Group Name*'
    "item_mark": COL["item_mark"],            # col 12
    "cold_item_mark": COL["cold_item_mark"],  # col 4
    "vakkal": COL["vakkal"],                  # col 5
    "cold_item_desc": 3,                      # 'Cold Item Description'
}

# cold_stocks column holding each logical attribute
COLD_ATTRS = {
    "article": "item_description", "group": "group_name", "subgroup": "item_subgroup",
    "item_mark": "item_mark", "cold_item_mark": "cold_item_mark", "vakkal": "vakkal",
}

# article-header tables: table suffix -> (attr -> column), qty column
ART_TABLES = {
    "articles": ({"article": "item_description", "group": "item_category",
                  "subgroup": "sub_category"}, "quantity_units"),
    "articles_v2": ({"article": "item_description", "group": "item_category",
                     "subgroup": "sub_category", "item_mark": "item_mark"},
                    "quantity_units"),
    "bulk_entry_articles": ({"article": "item_description", "group": "item_category",
                             "subgroup": "sub_category", "item_mark": "item_mark",
                             "vakkal": "vakkal"}, "box_count"),
}

ATTRS = ["article", "group", "subgroup", "item_mark", "cold_item_mark", "vakkal"]


def nz(v) -> str:
    """Normalise a descriptive value for comparison: casefold, ascii-only, no punct."""
    s = _s(v)
    if not s:
        return ""
    s = re.sub(r"[^\x20-\x7e]", " ", s)          # kill mojibake / non-ascii
    s = re.sub(r"[^0-9a-zA-Z]+", " ", s)         # punctuation -> space
    return " ".join(s.split()).casefold()


def fmt_set(vals) -> str:
    return " | ".join(sorted(v for v in vals if v)) or "-"


def cmp_attr(sheet_vals: set[str], db_vals: set[str]) -> str:
    """MATCH / MISMATCH / DB_BLANK / SHEET_BLANK / n-a."""
    s = {v for v in sheet_vals if v}
    d = {v for v in db_vals if v}
    if not s and not d:
        return "both_blank"
    if not d:
        return "DB_BLANK"
    if not s:
        return "SHEET_BLANK"
    return "MATCH" if s & d else "MISMATCH"


# ---------------- sheet ----------------
def load_sheet():
    ws = openpyxl.load_workbook(str(EXCEL_PATH), read_only=True, data_only=True)["Sheet1"]
    agg = {}
    for r in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        co, lot = _s(r[COL["company"]]).upper(), _s(r[COL["lot_no"]])
        if not lot or co not in PREFIX:
            continue
        key = (co, norm(lot))
        d = agg.setdefault(key, {"company": co, "lot": lot, "cart": 0.0, "rows": 0,
                                 "raw": defaultdict(set), "nrm": defaultdict(set)})
        try:
            d["cart"] += float(r[COL["cartons"]] or 0)
        except (TypeError, ValueError):
            pass
        d["rows"] += 1
        for attr, ci in SHEET_ATTRS.items():
            d["raw"][attr].add(_s(r[ci]))
            d["nrm"][attr].add(nz(r[ci]))
    return agg


# ---------------- db ----------------
def load_cold(conn):
    """(co, lotkey) -> {'n':boxes, 'lot':raw, 'synth':n, 'raw':{attr:set}, 'nrm':{attr:set}}"""
    out = {}
    cur = conn.cursor()
    cols = list(dict.fromkeys(COLD_ATTRS.values()))
    for co, p in PREFIX.items():
        cur.execute(f"""SELECT lot_no::text, count(*),
                        count(*) FILTER (WHERE transaction_no = 'RECON22JUL'),
                        {', '.join(f'array_agg(DISTINCT {c})' for c in cols)}
                        FROM {p}_cold_stocks WHERE lot_no IS NOT NULL GROUP BY 1""")
        for row in cur.fetchall():
            lot, n, synth = row[0], row[1], row[2]
            arrays = dict(zip(cols, row[3:]))
            k = (co, norm(lot))
            if not k[1]:
                continue
            rec = out.setdefault(k, {"company": co, "lot": _s(lot), "n": 0, "synth": 0,
                                     "raw": defaultdict(set), "nrm": defaultdict(set)})
            rec["n"] += n
            rec["synth"] += synth
            for attr, c in COLD_ATTRS.items():
                for v in (arrays[c] or []):
                    rec["raw"][attr].add(_s(v))
                    rec["nrm"][attr].add(nz(v))
    return out


def load_articles(conn):
    """(co, lotkey) -> {'src':set(table), 'qty':float, 'rows':n,
                        'raw':{attr:set}, 'nrm':{attr:set}, 'sku':set(sku_id)}"""
    out = {}
    cur = conn.cursor()
    for co, p in PREFIX.items():
        for suffix, (attrmap, qtycol) in ART_TABLES.items():
            tbl = f"{p}_{suffix}"
            sel = ", ".join(sorted(set(attrmap.values())))
            cur.execute(f"""SELECT lot_number::text, sku_id, {sel},
                                   COALESCE({qtycol}, 0)
                            FROM {tbl} WHERE lot_number IS NOT NULL""")
            names = [d.name for d in cur.description]
            for row in cur.fetchall():
                rec_row = dict(zip(names, row))
                k = (co, norm(rec_row["lot_number"]))
                if not k[1]:
                    continue
                rec = out.setdefault(k, {"src": set(), "qty": 0.0, "rows": 0,
                                         "raw": defaultdict(set), "nrm": defaultdict(set),
                                         "sku": set()})
                rec["src"].add(suffix)
                rec["rows"] += 1
                rec["qty"] += float(row[-1] or 0)
                if rec_row.get("sku_id"):
                    rec["sku"].add(rec_row["sku_id"])
                for attr, c in attrmap.items():
                    rec["raw"][attr].add(_s(rec_row[c]))
                    rec["nrm"][attr].add(nz(rec_row[c]))
    return out


def load_sku(conn):
    """company -> {normalised sku description}, plus the shared all_sku particulars."""
    cur = conn.cursor()
    per = {}
    for co, p in PREFIX.items():
        cur.execute(f'SELECT item_description FROM "{p}sku"')
        per[co] = {nz(v) for (v,) in cur.fetchall() if nz(v)}
    cur.execute("SELECT particulars FROM all_sku")
    allsku = {nz(v) for (v,) in cur.fetchall() if nz(v)}
    return per, allsku


# ---------------- main ----------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default=None)
    args = ap.parse_args()

    sheet = load_sheet()
    with psycopg.connect(os.environ["DATABASE_URL"]) as conn:
        cold = load_cold(conn)
        arts = load_articles(conn)
        skus, allsku = load_sku(conn)

    # ---------- 1. LOTS, both directions ----------
    lot_rows = []
    for k in sorted(set(sheet) | set(cold), key=lambda x: (x[0], x[1])):
        co, lk = k
        s, c = sheet.get(k), cold.get(k)
        scart = round(s["cart"]) if s else None
        cn = c["n"] if c else None
        if s and c:
            side = "BOTH"
        elif s:
            side = "SHEET_ONLY"
        else:
            side = "DB_ONLY"
        lot_rows.append({
            "company": co, "lot": (s or c)["lot"], "lot_key": lk, "side": side,
            "sheet_cartons": scart, "cold_boxes": cn,
            "delta": (cn - scart) if (s and c) else None,
            "synthetic_recon22jul": c["synth"] if c else 0,
            "sheet_article": fmt_set((s or {"raw": {}})["raw"].get("article", set())) if s else "",
            "cold_article": fmt_set(c["raw"].get("article", set())) if c else "",
            "has_article_row": "yes" if k in arts else "no",
        })

    # ---------- 2. ARTICLES ----------
    art_rows = []
    for k in sorted(set(sheet), key=lambda x: (x[0], x[1])):
        co, lk = k
        s = sheet[k]
        c, a = cold.get(k), arts.get(k)
        row = {"company": co, "lot": s["lot"], "sheet_cartons": round(s["cart"])}
        for attr in ATTRS:
            row[f"sheet_{attr}"] = fmt_set(s["raw"][attr])
        for attr in ATTRS:
            row[f"cold_{attr}"] = fmt_set(c["raw"].get(attr, set())) if c else ""
            row[f"cold_{attr}_cmp"] = (cmp_attr(s["nrm"][attr], c["nrm"].get(attr, set()))
                                       if c else "NO_COLD_ROW")
        row["article_src"] = ",".join(sorted(a["src"])) if a else ""
        row["article_rows"] = a["rows"] if a else 0
        # ponytail: TRANSACTION-grain, not per-lot. Verified: TR-20260702172629 has one
        # article row of box_count=15279 covering 6 lots; TR-20260624143934 has 3 rows
        # summing 4190 over 20 lots. So this is NOT comparable to a per-lot closing
        # quantity — carried for reference only, never differenced against sheet cartons.
        row["article_txn_qty"] = round(a["qty"], 3) if a else None
        row["article_sku_ids"] = ",".join(str(x) for x in sorted(a["sku"])) if a else ""
        for attr in ("article", "group", "subgroup", "item_mark", "vakkal"):
            row[f"art_{attr}"] = fmt_set(a["raw"].get(attr, set())) if a else ""
            row[f"art_{attr}_cmp"] = (cmp_attr(s["nrm"][attr], a["nrm"].get(attr, set()))
                                      if a else "NO_ARTICLE_ROW")
        sa = {v for v in s["nrm"]["article"] if v}
        row["sku_master"] = ("MATCH" if sa & skus[co] else
                             "IN_OTHER_SKU_MASTER" if sa & skus["CDPL" if co == "CFPL" else "CFPL"]
                             else "IN_ALL_SKU" if sa & allsku else "NOT_A_KNOWN_SKU")
        art_rows.append(row)

    # ---------- console ----------
    print(f"=== 1. LOT COMPARISON (two-way) ===")
    by = defaultdict(lambda: [0, 0, 0])
    for r in lot_rows:
        b = by[r["side"]]
        b[0] += 1
        b[1] += r["sheet_cartons"] or 0
        b[2] += r["cold_boxes"] or 0
    for side in ("BOTH", "SHEET_ONLY", "DB_ONLY"):
        n, sc, cb = by[side]
        print(f"  {side:<12s} lots={n:<5d} sheet_cartons={sc:>9,d}  cold_boxes={cb:>9,d}")
    both = [r for r in lot_rows if r["side"] == "BOTH"]
    print(f"  of BOTH: {sum(1 for r in both if r['delta'] == 0)} exact, "
          f"{sum(1 for r in both if r['delta'] != 0)} carton mismatch")

    print(f"\n  -- DB_ONLY (in cold_stocks, absent from the 29-Jul sheet) --")
    dbo = sorted((r for r in lot_rows if r["side"] == "DB_ONLY"),
                 key=lambda r: -r["cold_boxes"])
    print(f"  {'CO':5s}{'LOT':>9s}{'COLD':>7s}{'SYNTH':>7s}  ARTICLE (cold_stocks)")
    for r in dbo:
        print(f"  {r['company']:5s}{r['lot']:>9s}{r['cold_boxes']:>7d}"
              f"{r['synthetic_recon22jul']:>7d}  {r['cold_article'][:60]}")
    print(f"  total DB-only boxes: {sum(r['cold_boxes'] for r in dbo):,}")

    print(f"\n=== 2. ARTICLE COMPARISON ({len(art_rows)} sheet lots) ===")
    print("  -- vs cold_stocks --")
    for attr in ATTRS:
        cnt = defaultdict(int)
        for r in art_rows:
            cnt[r[f"cold_{attr}_cmp"]] += 1
        print(f"   {attr:<16s}" + "  ".join(f"{k}={v}" for k, v in sorted(cnt.items())))
    print("  -- vs article headers --")
    for attr in ("article", "group", "subgroup", "item_mark", "vakkal"):
        cnt = defaultdict(int)
        for r in art_rows:
            cnt[r[f"art_{attr}_cmp"]] += 1
        print(f"   {attr:<16s}" + "  ".join(f"{k}={v}" for k, v in sorted(cnt.items())))
    cnt = defaultdict(int)
    for r in art_rows:
        cnt[r["sku_master"]] += 1
    print("  -- sheet article name vs SKU master --")
    print("   " + "  ".join(f"{k}={v}" for k, v in sorted(cnt.items())))

    print(f"\n  -- article-name MISMATCH vs cold_stocks --")
    mm = [r for r in art_rows if r["cold_article_cmp"] == "MISMATCH"]
    for r in sorted(mm, key=lambda r: -r["sheet_cartons"]):
        print(f"   {r['company']} {r['lot']:>8s} {r['sheet_cartons']:>6d}  "
              f"sheet={r['sheet_article'][:38]:<38s} cold={r['cold_article'][:38]}")
    print(f"   ({len(mm)} lots)")

    for attr in ("group", "subgroup", "item_mark"):
        mm = [r for r in art_rows if r[f"cold_{attr}_cmp"] == "MISMATCH"]
        print(f"\n  -- {attr} MISMATCH vs cold_stocks ({len(mm)} lots) --")
        for r in sorted(mm, key=lambda r: -r["sheet_cartons"])[:40]:
            print(f"   {r['company']} {r['lot']:>8s} {r['sheet_cartons']:>6d}  "
                  f"sheet={r[f'sheet_{attr}'][:34]:<34s} cold={r[f'cold_{attr}'][:34]}")

    print(f"\n  -- article-header coverage --")
    have = [r for r in art_rows if r["article_rows"]]
    print(f"   {len(have)} of {len(art_rows)} sheet lots have any article header row "
          f"({len(art_rows) - len(have)} have NONE)")
    print("   article box_count/quantity_units is TRANSACTION-grain (one row can cover "
          "many lots),\n   so it is NOT differenced against per-lot sheet cartons:")
    for r in sorted(have, key=lambda r: -(r["article_txn_qty"] or 0)):
        print(f"   {r['company']} {r['lot']:>8s} sheet_cartons={r['sheet_cartons']:>6d} "
              f"txn_qty={r['article_txn_qty']:>9.0f}  src={r['article_src']}")

    print(f"\n  -- sheet article name NOT in any SKU master --")
    for r in sorted((x for x in art_rows if x["sku_master"] == "NOT_A_KNOWN_SKU"),
                    key=lambda r: -r["sheet_cartons"]):
        print(f"   {r['company']} {r['lot']:>8s} {r['sheet_cartons']:>6d}  "
              f"{r['sheet_article'][:70]}")

    if args.xlsx:
        wb = openpyxl.Workbook()
        for title, rows in (("lots_two_way", lot_rows), ("articles", art_rows)):
            ws = wb.create_sheet(title)
            ws.append(list(rows[0].keys()))
            for r in rows:
                ws.append(list(r.values()))
            ws.freeze_panes = "A2"
        del wb["Sheet"]
        wb.save(args.xlsx)
        print(f"\nwrote {args.xlsx}")


if __name__ == "__main__":
    main()
