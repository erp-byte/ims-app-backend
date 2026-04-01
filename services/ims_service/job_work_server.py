import base64
import io
import json
import os
import re
from datetime import datetime

import anthropic
import openpyxl
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends, Query
from fastapi.responses import JSONResponse
from sqlalchemy import text
from sqlalchemy.orm import Session

from shared.database import get_db

router = APIRouter(prefix="/job-work", tags=["Job Work"])


def _parse_date_to_iso(date_str: str) -> str:
    """Convert DD-MM-YYYY to YYYY-MM-DD for PostgreSQL. Pass through if already ISO."""
    if not date_str:
        return date_str
    parts = date_str.split("-")
    if len(parts) == 3 and len(parts[0]) == 2 and len(parts[2]) == 4:
        return f"{parts[2]}-{parts[1]}-{parts[0]}"
    return date_str


# ════════════════════════════════════════════════════════════
#  DATABASE TABLE CREATION (runs once on first request)
# ════════════════════════════════════════════════════════════

_tables_created = False


def _ensure_tables(db: Session):
    global _tables_created
    if _tables_created:
        return

    # ── JOB WORK HEADER (Material Out) ──
    # Stores exactly what the material-out form submits
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS jb_materialout_header (
            id SERIAL PRIMARY KEY,
            challan_no VARCHAR(100) NOT NULL,
            job_work_date VARCHAR(20),
            from_warehouse VARCHAR(100),
            to_party VARCHAR(255),
            party_address TEXT,
            party_state VARCHAR(100),
            party_city VARCHAR(100),
            party_pin_code VARCHAR(10),
            party_contact_company VARCHAR(255),
            party_contact_mobile VARCHAR(50),
            party_email VARCHAR(255),
            sub_category VARCHAR(100),
            contact_person VARCHAR(255),
            contact_number VARCHAR(50),
            purpose_of_work TEXT,
            expected_return_date VARCHAR(20),
            vehicle_no VARCHAR(50),
            driver_name VARCHAR(100),
            authorized_person VARCHAR(255),
            remarks TEXT,
            e_way_bill_no VARCHAR(100),
            dispatched_through VARCHAR(255),
            type VARCHAR(10) NOT NULL DEFAULT 'OUT',
            status VARCHAR(30) NOT NULL DEFAULT 'sent',
            dispatch_to JSONB,
            payload JSONB,
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW()
        )
    """))

    # ── JOB WORK LINE ITEMS (Items dispatched) ──
    # Matches exactly what line_items[] in payload contains
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS jb_materialout_lines (
            id SERIAL PRIMARY KEY,
            header_id INTEGER REFERENCES jb_materialout_header(id) ON DELETE CASCADE,
            sl_no INTEGER,
            item_description VARCHAR(500),
            material_type VARCHAR(50),
            item_category VARCHAR(100),
            sub_category VARCHAR(100),
            quantity_kgs NUMERIC(12,3) DEFAULT 0,
            quantity_boxes INTEGER DEFAULT 0,
            rate_per_kg NUMERIC(12,2) DEFAULT 0,
            amount NUMERIC(12,2) DEFAULT 0,
            uom VARCHAR(20),
            case_pack VARCHAR(20),
            net_weight VARCHAR(20),
            total_weight VARCHAR(20),
            batch_number VARCHAR(100),
            lot_number VARCHAR(100),
            manufacturing_date VARCHAR(20),
            expiry_date VARCHAR(20),
            line_remarks TEXT,
            cold_unit VARCHAR(50),
            item_mark VARCHAR(255),
            box_id VARCHAR(100),
            transaction_no VARCHAR(100),
            cold_storage_snapshot JSONB
        )
    """))
    # Add columns if table already exists without them (covers all columns in CREATE TABLE above)
    for col, col_type in [
        ("sl_no", "INTEGER"),
        ("item_description", "VARCHAR(500)"),
        ("material_type", "VARCHAR(50)"),
        ("item_category", "VARCHAR(100)"),
        ("sub_category", "VARCHAR(100)"),
        ("quantity_kgs", "NUMERIC(12,3) DEFAULT 0"),
        ("quantity_boxes", "INTEGER DEFAULT 0"),
        ("rate_per_kg", "NUMERIC(12,2) DEFAULT 0"),
        ("amount", "NUMERIC(12,2) DEFAULT 0"),
        ("uom", "VARCHAR(20)"),
        ("case_pack", "VARCHAR(20)"),
        ("net_weight", "VARCHAR(20)"),
        ("total_weight", "VARCHAR(20)"),
        ("batch_number", "VARCHAR(100)"),
        ("lot_number", "VARCHAR(100)"),
        ("manufacturing_date", "VARCHAR(20)"),
        ("expiry_date", "VARCHAR(20)"),
        ("line_remarks", "TEXT"),
        ("cold_unit", "VARCHAR(50)"),
        ("item_mark", "VARCHAR(255)"),
        ("box_id", "VARCHAR(100)"),
        ("transaction_no", "VARCHAR(100)"),
        ("cold_storage_snapshot", "JSONB"),
    ]:
        try:
            db.execute(text(f"ALTER TABLE jb_materialout_lines ADD COLUMN IF NOT EXISTS {col} {col_type}"))
        except Exception:
            pass

    # ── INWARD RECEIPT HEADER (each partial/final receive event) ──
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS jb_work_inward_receipt (
            id SERIAL PRIMARY KEY,
            ir_number VARCHAR(100) NOT NULL,
            challan_no VARCHAR(100),
            header_id INTEGER REFERENCES jb_materialout_header(id) ON DELETE CASCADE,
            receipt_date VARCHAR(20),
            receipt_type VARCHAR(10) NOT NULL DEFAULT 'partial',
            vehicle_no VARCHAR(50),
            driver_name VARCHAR(100),
            remarks TEXT,
            created_by VARCHAR(255),
            created_at TIMESTAMP DEFAULT NOW()
        )
    """))
    # Add columns if table already exists without them
    try:
        db.execute(text("ALTER TABLE jb_work_inward_receipt ADD COLUMN IF NOT EXISTS challan_no VARCHAR(100)"))
        db.execute(text("ALTER TABLE jb_work_inward_receipt ADD COLUMN IF NOT EXISTS inward_warehouse VARCHAR(255)"))
    except Exception:
        pass

    # ── INWARD RECEIPT LINE ITEMS (FG + Waste + Rejection per item per IR) ──
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS jb_work_inward_lines (
            id SERIAL PRIMARY KEY,
            inward_receipt_id INTEGER REFERENCES jb_work_inward_receipt(id) ON DELETE CASCADE,
            sl_no INTEGER,
            item_description VARCHAR(500),
            sent_kgs NUMERIC(12,3) DEFAULT 0,
            sent_boxes INTEGER DEFAULT 0,
            finished_goods_kgs NUMERIC(12,3) DEFAULT 0,
            finished_goods_boxes INTEGER DEFAULT 0,
            waste_kgs NUMERIC(12,3) DEFAULT 0,
            waste_type VARCHAR(100),
            rejection_kgs NUMERIC(12,3) DEFAULT 0,
            rejection_boxes INTEGER DEFAULT 0,
            line_remarks TEXT,
            process_type VARCHAR(100),
            min_loss_pct NUMERIC(5,2) DEFAULT 0,
            max_loss_pct NUMERIC(5,2) DEFAULT 0,
            waste_with_partial BOOLEAN DEFAULT TRUE,
            single_shot BOOLEAN DEFAULT FALSE
        )
    """))


    # ── JB INWARD BOXES — table created manually in DB ──

    db.commit()
    _tables_created = True


# ════════════════════════════════════════════════════════════
#  ALL_SKU DROPDOWN (for Article Entry in Material-In)
# ════════════════════════════════════════════════════════════

@router.get("/all-sku-dropdown")
def all_sku_dropdown(
    item_category: str = Query(""),
    sub_category: str = Query(""),
    search: str = Query(""),
    db: Session = Depends(get_db),
):
    """Cascading dropdown from public.all_sku: item_group -> sub_group -> particulars."""
    tbl = "public.all_sku"

    # 1) All item groups
    item_categories = db.execute(
        text(f"""
            SELECT DISTINCT UPPER(item_group) AS grp FROM {tbl}
            WHERE item_group IS NOT NULL AND item_group != ''
            ORDER BY grp ASC
        """)
    ).scalars().all()

    # 2) Sub groups (filtered by item_group)
    sub_categories = []
    if item_category.strip():
        sub_categories = db.execute(
            text(f"""
                SELECT DISTINCT UPPER(sub_group) AS sg FROM {tbl}
                WHERE UPPER(item_group) = UPPER(:ic)
                  AND sub_group IS NOT NULL AND sub_group != ''
                ORDER BY sg ASC
            """),
            {"ic": item_category.strip()},
        ).scalars().all()

    # 3) Particulars (filtered by item_group + sub_group)
    item_descriptions = []
    if item_category.strip() and sub_category.strip():
        where = "UPPER(item_group) = UPPER(:ic) AND UPPER(sub_group) = UPPER(:sc) AND particulars IS NOT NULL AND particulars != ''"
        params: dict = {"ic": item_category.strip(), "sc": sub_category.strip()}
        if search.strip():
            where += " AND LOWER(particulars) LIKE :search"
            params["search"] = f"%{search.strip().lower()}%"
        item_descriptions = db.execute(
            text(f"""
                SELECT DISTINCT UPPER(particulars) AS item_desc FROM {tbl}
                WHERE {where}
                ORDER BY item_desc ASC
            """),
            params,
        ).scalars().all()

    return {
        "options": {
            "item_categories": list(item_categories),
            "sub_categories": list(sub_categories),
            "item_descriptions": list(item_descriptions),
        }
    }


@router.get("/all-sku-search")
def all_sku_search(
    search: str = Query(""),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
):
    """Global search on public.all_sku by particulars — returns item with group/sub_group."""
    tbl = "public.all_sku"
    search_term = search.strip()
    if not search_term:
        return {"items": []}

    rows = db.execute(
        text(f"""
            SELECT DISTINCT ON (UPPER(particulars))
                   UPPER(particulars) AS item_desc,
                   UPPER(item_group) AS grp,
                   UPPER(sub_group) AS sg
            FROM {tbl}
            WHERE LOWER(particulars) LIKE :search
              AND particulars IS NOT NULL AND particulars != ''
            ORDER BY UPPER(particulars) ASC
            LIMIT :limit
        """),
        {"search": f"%{search_term.lower()}%", "limit": limit},
    ).fetchall()

    return {
        "items": [
            {"item_description": r[0], "item_group": r[1] or "", "sub_group": r[2] or ""}
            for r in rows
        ]
    }


# ════════════════════════════════════════════════════════════
#  COLD STORAGE STOCK HELPERS
# ════════════════════════════════════════════════════════════

COLD_UNIT_TO_TABLE = {
    "cfpl": "cfpl_cold_stocks",
    "cdpl": "cdpl_cold_stocks",
}


def _resolve_cold_table(cold_unit: str):
    """Map cold_unit value to the correct cold storage table name.
    Accepts codes like 'cfpl'/'cdpl' or display names like 'Savla D-39'/'Rishi'."""
    if not cold_unit:
        return None
    cu = cold_unit.strip().lower()
    if cu in COLD_UNIT_TO_TABLE:
        return COLD_UNIT_TO_TABLE[cu]
    # Map display names to table
    if "savla" in cu or "d-39" in cu or "d39" in cu:
        return "cfpl_cold_stocks"
    if "rishi" in cu:
        return "cdpl_cold_stocks"
    return None


def _deduct_cold_storage_stock(db: Session, header_id: int, line_items: list):
    """Delete rows from cold storage tables for each dispatched box.
    Before deleting, snapshot the full cold row into jb_materialout_lines.cold_storage_snapshot."""
    for item in line_items:
        box_id = item.get("box_id", "")
        transaction_no = item.get("transaction_no", "")
        cold_unit = item.get("cold_unit", "")
        if not box_id or not transaction_no or not cold_unit:
            continue
        table = _resolve_cold_table(cold_unit)
        if not table:
            continue

        # Fetch the full cold storage row before deleting
        cold_row = db.execute(
            text(f"SELECT * FROM {table} WHERE box_id = :box_id AND transaction_no = :transaction_no"),
            {"box_id": box_id, "transaction_no": transaction_no},
        ).mappings().fetchone()

        if not cold_row:
            # Row not found by box_id+transaction_no — save frontend snapshot if available
            frontend_snapshot = item.get("cold_stock_snapshot")
            if frontend_snapshot and isinstance(frontend_snapshot, dict):
                snapshot = {k: (str(v) if v is not None else None) for k, v in frontend_snapshot.items() if k != "id"}
                db.execute(
                    text("""
                        UPDATE jb_materialout_lines
                        SET cold_storage_snapshot = :snapshot
                        WHERE header_id = :header_id AND box_id = :box_id AND transaction_no = :transaction_no
                    """),
                    {"snapshot": json.dumps(snapshot), "header_id": header_id, "box_id": box_id, "transaction_no": transaction_no},
                )
            continue

        # Build snapshot: prefer DB row (has all columns), merge with frontend snapshot for any missing fields
        snapshot = {k: (str(v) if v is not None else None) for k, v in dict(cold_row).items() if k != "id"}
        frontend_snapshot = item.get("cold_stock_snapshot")
        if frontend_snapshot and isinstance(frontend_snapshot, dict):
            for k, v in frontend_snapshot.items():
                if k != "id" and (snapshot.get(k) is None) and v is not None:
                    snapshot[k] = str(v)

        db.execute(
            text("""
                UPDATE jb_materialout_lines
                SET cold_storage_snapshot = :snapshot
                WHERE header_id = :header_id AND box_id = :box_id AND transaction_no = :transaction_no
            """),
            {"snapshot": json.dumps(snapshot), "header_id": header_id, "box_id": box_id, "transaction_no": transaction_no},
        )

        # Delete from cold storage
        db.execute(
            text(f"DELETE FROM {table} WHERE box_id = :box_id AND transaction_no = :transaction_no"),
            {"box_id": box_id, "transaction_no": transaction_no},
        )


def _restore_cold_storage_stock(db: Session, header_id: int):
    """Re-insert rows into cold storage tables from saved snapshots (reversal on delete)."""
    lines = db.execute(text("""
        SELECT box_id, transaction_no, cold_unit, cold_storage_snapshot,
               item_description, item_mark, lot_number, net_weight, item_category
        FROM jb_materialout_lines
        WHERE header_id = :header_id
    """), {"header_id": header_id}).mappings().all()

    for l in lines:
        box_id = l["box_id"] or ""
        transaction_no = l["transaction_no"] or ""
        cold_unit = l["cold_unit"] or ""
        if not box_id or not transaction_no or not cold_unit:
            continue
        table = _resolve_cold_table(cold_unit)
        if not table:
            continue

        # Check if row already exists (avoid duplicates)
        existing = db.execute(
            text(f"SELECT id FROM {table} WHERE box_id = :box_id AND transaction_no = :transaction_no"),
            {"box_id": box_id, "transaction_no": transaction_no},
        ).fetchone()
        if existing:
            continue

        # Try to restore from snapshot first (has all original fields)
        snapshot_raw = l["cold_storage_snapshot"]
        if snapshot_raw:
            snapshot = json.loads(snapshot_raw) if isinstance(snapshot_raw, str) else snapshot_raw
            cols = [k for k in snapshot.keys() if k != "id" and snapshot[k] is not None]
            if cols:
                col_names = ", ".join(cols)
                col_params = ", ".join(f":{c}" for c in cols)
                db.execute(text(f"INSERT INTO {table} ({col_names}) VALUES ({col_params})"), snapshot)
                continue

        # Fallback: restore with minimal fields from line item data
        weight_kg = float(l["net_weight"] or 0)
        db.execute(
            text(f"""
                INSERT INTO {table}
                    (item_description, item_mark, lot_no, no_of_cartons, weight_kg,
                     total_inventory_kgs, group_name, box_id, transaction_no)
                VALUES
                    (:item_description, :item_mark, :lot_no, 1, :weight_kg,
                     :weight_kg, :group_name, :box_id, :transaction_no)
            """),
            {
                "item_description": l["item_description"] or "",
                "item_mark": l["item_mark"] or "",
                "lot_no": l["lot_number"] or "",
                "weight_kg": weight_kg,
                "group_name": l["item_category"] or "",
                "box_id": box_id,
                "transaction_no": transaction_no,
            },
        )


# ════════════════════════════════════════════════════════════
#  POST /job-work/out  — Submit Material Out
# ════════════════════════════════════════════════════════════

@router.post("/out")
def submit_material_out(
    payload: dict,
    created_by: str = Query(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    header = payload.get("header", {})
    dispatch_to = payload.get("dispatch_to", {})
    line_items = payload.get("line_items", [])

    result = db.execute(text("""
        INSERT INTO jb_materialout_header
            (challan_no, job_work_date, from_warehouse, to_party, party_address,
             party_state, party_city, party_pin_code, party_contact_company,
             party_contact_mobile, party_email, sub_category,
             contact_person, contact_number, purpose_of_work, expected_return_date,
             vehicle_no, driver_name, authorized_person, remarks,
             e_way_bill_no, dispatched_through, type, status, dispatch_to, payload, created_by)
        VALUES
            (:challan_no, :job_work_date, :from_warehouse, :to_party, :party_address,
             :party_state, :party_city, :party_pin_code, :party_contact_company,
             :party_contact_mobile, :party_email, :sub_category,
             :contact_person, :contact_number, :purpose_of_work, :expected_return_date,
             :vehicle_no, :driver_name, :authorized_person, :remarks,
             :e_way_bill_no, :dispatched_through, 'OUT', 'sent', :dispatch_to, :payload, :created_by)
        RETURNING id
    """), {
        "challan_no": header.get("challan_no") or payload.get("challan_no", ""),
        "job_work_date": header.get("job_work_date") or payload.get("dated", ""),
        "from_warehouse": header.get("from_warehouse", ""),
        "to_party": header.get("to_party") or dispatch_to.get("name", ""),
        "party_address": header.get("party_address") or dispatch_to.get("address", ""),
        "party_state": dispatch_to.get("state", ""),
        "party_city": dispatch_to.get("city", ""),
        "party_pin_code": dispatch_to.get("pin_code", ""),
        "party_contact_company": dispatch_to.get("contact_company", ""),
        "party_contact_mobile": dispatch_to.get("contact_mobile", ""),
        "party_email": dispatch_to.get("email", ""),
        "sub_category": dispatch_to.get("sub_category", ""),
        "contact_person": header.get("contact_person", ""),
        "contact_number": header.get("contact_number", ""),
        "purpose_of_work": header.get("purpose_of_work", "") or dispatch_to.get("sub_category", ""),
        "expected_return_date": header.get("expected_return_date", ""),
        "vehicle_no": header.get("vehicle_no") or payload.get("motor_vehicle_no", ""),
        "driver_name": header.get("driver_name", ""),
        "authorized_person": header.get("authorized_person", ""),
        "remarks": header.get("remarks") or payload.get("remarks", ""),
        "e_way_bill_no": payload.get("e_way_bill_no", ""),
        "dispatched_through": payload.get("dispatched_through", ""),
        "dispatch_to": json.dumps(dispatch_to),
        "payload": json.dumps(payload),
        "created_by": created_by,
    })
    header_id = result.fetchone()[0]

    for item in line_items:
        qty = item.get("quantity", {})
        kgs = qty.get("kgs", 0) if isinstance(qty, dict) else 0
        boxes = qty.get("boxes", 0) if isinstance(qty, dict) else 0

        db.execute(text("""
            INSERT INTO jb_materialout_lines
                (header_id, sl_no, item_description, material_type, item_category, sub_category,
                 quantity_kgs, quantity_boxes, rate_per_kg, amount,
                 uom, case_pack, net_weight, total_weight,
                 batch_number, lot_number, manufacturing_date, expiry_date, line_remarks,
                 box_id, transaction_no, cold_unit, item_mark)
            VALUES
                (:header_id, :sl_no, :item_description, :material_type, :item_category, :sub_category,
                 :quantity_kgs, :quantity_boxes, :rate_per_kg, :amount,
                 :uom, :case_pack, :net_weight, :total_weight,
                 :batch_number, :lot_number, :manufacturing_date, :expiry_date, :line_remarks,
                 :box_id, :transaction_no, :cold_unit, :item_mark)
        """), {
            "header_id": header_id,
            "sl_no": item.get("sl_no", 0),
            "item_description": item.get("item_description") or item.get("description", ""),
            "material_type": item.get("material_type", ""),
            "item_category": item.get("item_category", ""),
            "sub_category": item.get("sub_category", ""),
            "quantity_kgs": kgs,
            "quantity_boxes": boxes,
            "rate_per_kg": item.get("rate_per_kg", 0),
            "amount": item.get("amount", 0),
            "uom": item.get("uom", ""),
            "case_pack": str(item.get("case_pack", "")),
            "net_weight": str(item.get("net_weight", "")),
            "total_weight": str(item.get("total_weight", "")),
            "batch_number": item.get("batch_number", ""),
            "lot_number": item.get("lot_number", ""),
            "manufacturing_date": item.get("manufacturing_date", ""),
            "expiry_date": item.get("expiry_date", ""),
            "line_remarks": item.get("remarks") or item.get("line_remarks", ""),
            "box_id": item.get("box_id", ""),
            "transaction_no": item.get("transaction_no", ""),
            "cold_unit": item.get("cold_unit", ""),
            "item_mark": item.get("item_mark", ""),
        })

    # ── Subtract from cold storage tables ──
    # If material is dispatched from cold storage, delete the picked box rows
    _deduct_cold_storage_stock(db, header_id, line_items)

    db.commit()
    return {"status": "success", "id": header_id, "challan_no": header.get("challan_no") or payload.get("challan_no", "")}


# ════════════════════════════════════════════════════════════
#  PUT /job-work/out/{record_id}  — Update Material Out
# ════════════════════════════════════════════════════════════

@router.put("/out/{record_id}")
def update_material_out(
    record_id: int,
    payload: dict,
    created_by: str = Query(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    # Verify record exists
    existing = db.execute(text("SELECT id FROM jb_materialout_header WHERE id = :id"), {"id": record_id}).fetchone()
    if not existing:
        raise HTTPException(status_code=404, detail="Record not found.")

    header = payload.get("header", {})
    dispatch_to = payload.get("dispatch_to", {})
    line_items = payload.get("line_items", [])

    # Update header
    db.execute(text("""
        UPDATE jb_materialout_header SET
            challan_no = :challan_no, job_work_date = :job_work_date, from_warehouse = :from_warehouse,
            to_party = :to_party, party_address = :party_address,
            party_state = :party_state, party_city = :party_city, party_pin_code = :party_pin_code,
            party_contact_company = :party_contact_company, party_contact_mobile = :party_contact_mobile,
            party_email = :party_email, sub_category = :sub_category,
            contact_person = :contact_person, contact_number = :contact_number,
            purpose_of_work = :purpose_of_work, expected_return_date = :expected_return_date,
            vehicle_no = :vehicle_no, driver_name = :driver_name, authorized_person = :authorized_person,
            remarks = :remarks, e_way_bill_no = :e_way_bill_no, dispatched_through = :dispatched_through,
            dispatch_to = :dispatch_to, payload = :payload, updated_at = NOW()
        WHERE id = :id
    """), {
        "id": record_id,
        "challan_no": header.get("challan_no") or payload.get("challan_no", ""),
        "job_work_date": header.get("job_work_date") or payload.get("dated", ""),
        "from_warehouse": header.get("from_warehouse", ""),
        "to_party": header.get("to_party") or dispatch_to.get("name", ""),
        "party_address": header.get("party_address") or dispatch_to.get("address", ""),
        "party_state": dispatch_to.get("state", ""),
        "party_city": dispatch_to.get("city", ""),
        "party_pin_code": dispatch_to.get("pin_code", ""),
        "party_contact_company": dispatch_to.get("contact_company", ""),
        "party_contact_mobile": dispatch_to.get("contact_mobile", ""),
        "party_email": dispatch_to.get("email", ""),
        "sub_category": dispatch_to.get("sub_category", ""),
        "contact_person": header.get("contact_person", ""),
        "contact_number": header.get("contact_number", ""),
        "purpose_of_work": header.get("purpose_of_work", "") or dispatch_to.get("sub_category", ""),
        "expected_return_date": header.get("expected_return_date", ""),
        "vehicle_no": header.get("vehicle_no") or payload.get("motor_vehicle_no", ""),
        "driver_name": header.get("driver_name", ""),
        "authorized_person": header.get("authorized_person", ""),
        "remarks": header.get("remarks") or payload.get("remarks", ""),
        "e_way_bill_no": payload.get("e_way_bill_no", ""),
        "dispatched_through": payload.get("dispatched_through", ""),
        "dispatch_to": json.dumps(dispatch_to),
        "payload": json.dumps(payload),
    })

    # Restore old cold storage stock before deleting old lines
    _restore_cold_storage_stock(db, record_id)

    # Delete old lines and re-insert
    db.execute(text("DELETE FROM jb_materialout_lines WHERE header_id = :id"), {"id": record_id})

    for item in line_items:
        qty = item.get("quantity", {})
        kgs = qty.get("kgs", 0) if isinstance(qty, dict) else 0
        boxes = qty.get("boxes", 0) if isinstance(qty, dict) else 0

        db.execute(text("""
            INSERT INTO jb_materialout_lines
                (header_id, sl_no, item_description, material_type, item_category, sub_category,
                 quantity_kgs, quantity_boxes, rate_per_kg, amount,
                 uom, case_pack, net_weight, total_weight,
                 batch_number, lot_number, manufacturing_date, expiry_date, line_remarks,
                 box_id, transaction_no, cold_unit, item_mark)
            VALUES
                (:header_id, :sl_no, :item_description, :material_type, :item_category, :sub_category,
                 :quantity_kgs, :quantity_boxes, :rate_per_kg, :amount,
                 :uom, :case_pack, :net_weight, :total_weight,
                 :batch_number, :lot_number, :manufacturing_date, :expiry_date, :line_remarks,
                 :box_id, :transaction_no, :cold_unit, :item_mark)
        """), {
            "header_id": record_id,
            "sl_no": item.get("sl_no", 0),
            "item_description": item.get("item_description") or item.get("description", ""),
            "material_type": item.get("material_type", ""),
            "item_category": item.get("item_category", ""),
            "sub_category": item.get("sub_category", ""),
            "quantity_kgs": kgs,
            "quantity_boxes": boxes,
            "rate_per_kg": item.get("rate_per_kg", 0),
            "amount": item.get("amount", 0),
            "uom": item.get("uom", ""),
            "case_pack": str(item.get("case_pack", "")),
            "net_weight": str(item.get("net_weight", "")),
            "total_weight": str(item.get("total_weight", "")),
            "batch_number": item.get("batch_number", ""),
            "lot_number": item.get("lot_number", ""),
            "manufacturing_date": item.get("manufacturing_date", ""),
            "expiry_date": item.get("expiry_date", ""),
            "line_remarks": item.get("remarks") or item.get("line_remarks", ""),
            "box_id": item.get("box_id", ""),
            "transaction_no": item.get("transaction_no", ""),
            "cold_unit": item.get("cold_unit", ""),
            "item_mark": item.get("item_mark", ""),
        })

    # Deduct new cold storage stock
    _deduct_cold_storage_stock(db, record_id, line_items)

    db.commit()
    return {"status": "success", "id": record_id, "challan_no": header.get("challan_no") or payload.get("challan_no", "")}


# ════════════════════════════════════════════════════════════
#  GET /job-work/list  — List all job work records
# ════════════════════════════════════════════════════════════

@router.get("/list")
def list_job_work_records(
    page: int = Query(1, ge=1),
    per_page: int = Query(15, ge=1, le=100),
    challan_no: str = Query("", description="Search by challan number"),
    status: str = Query("", description="Filter by status"),
    date: str = Query("", description="Filter by date (YYYY-MM-DD)"),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    where_clauses = []
    params: dict = {"limit": per_page}
    if challan_no.strip():
        where_clauses.append("h.challan_no ILIKE :challan_no")
        params["challan_no"] = f"%{challan_no.strip()}%"
    if status.strip():
        where_clauses.append("h.status = :status")
        params["status"] = status.strip()
    if date.strip():
        where_clauses.append("h.job_work_date = :date")
        params["date"] = date.strip()

    where_sql = (" WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    count_result = db.execute(text(f"SELECT COUNT(*) FROM jb_materialout_header h{where_sql}"), params)
    total = count_result.scalar() or 0
    total_pages = max(1, (total + per_page - 1) // per_page)
    offset = (page - 1) * per_page
    params["offset"] = offset

    rows = db.execute(text(f"""
        SELECT h.id, h.challan_no, h.job_work_date, h.from_warehouse, h.to_party,
               h.party_address, h.status, h.type, h.vehicle_no, h.driver_name,
               h.authorized_person, h.remarks, h.created_by, h.created_at,
               (SELECT COUNT(*) FROM jb_materialout_lines l WHERE l.header_id = h.id) as items_count,
               (SELECT string_agg(l2.item_description, ', ') FROM jb_materialout_lines l2 WHERE l2.header_id = h.id) as item_descriptions,
               (SELECT COALESCE(SUM(l3.quantity_boxes), 0) FROM jb_materialout_lines l3 WHERE l3.header_id = h.id) as total_qty,
               (SELECT COALESCE(SUM(l4.quantity_kgs), 0) FROM jb_materialout_lines l4 WHERE l4.header_id = h.id) as total_weight
        FROM jb_materialout_header h
        {where_sql}
        ORDER BY h.created_at DESC
        LIMIT :limit OFFSET :offset
    """), params).fetchall()

    records = []
    for r in rows:
        records.append({
            "id": r[0],
            "challan_no": r[1],
            "job_work_date": r[2] or "",
            "from_warehouse": r[3] or "",
            "to_party": r[4] or "",
            "party_address": r[5] or "",
            "status": r[6] or "sent",
            "type": r[7] or "OUT",
            "vehicle_no": r[8] or "",
            "driver_name": r[9] or "",
            "authorized_person": r[10] or "",
            "remarks": r[11] or "",
            "created_by": r[12] or "",
            "created_at": str(r[13]) if r[13] else "",
            "items_count": r[14] or 0,
            "item_descriptions": r[15] or "",
            "total_qty": int(r[16] or 0),
            "total_weight": float(r[17] or 0),
        })

    return {"records": records, "total": total, "total_pages": total_pages, "page": page}


# ════════════════════════════════════════════════════════════
#  GET /job-work/out/search  — Search outward record (for Material In)
# ════════════════════════════════════════════════════════════

@router.get("/out/search")
def search_material_out(
    challan_no: str = Query(""),
    vendor_name: str = Query(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    if not challan_no and not vendor_name:
        raise HTTPException(status_code=400, detail="Provide challan_no or vendor_name to search.")

    conditions = ["h.type = 'OUT'"]
    params = {}
    if challan_no:
        conditions.append("LOWER(h.challan_no) = LOWER(:challan_no)")
        params["challan_no"] = challan_no.strip()
    if vendor_name:
        conditions.append("LOWER(h.to_party) LIKE LOWER(:vendor_name)")
        params["vendor_name"] = f"%{vendor_name.strip()}%"

    where = " AND ".join(conditions)
    row = db.execute(text(f"""
        SELECT h.id, h.challan_no, h.job_work_date, h.from_warehouse, h.to_party,
               h.status, h.vehicle_no, h.sub_category, h.dispatch_to, h.payload
        FROM jb_materialout_header h
        WHERE {where}
        ORDER BY h.created_at DESC
        LIMIT 1
    """), params).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="No matching material out record found.")

    header_id = row[0]
    lines = db.execute(text("""
        SELECT sl_no, item_description, quantity_kgs, quantity_boxes, net_weight
        FROM jb_materialout_lines
        WHERE header_id = :header_id
        ORDER BY sl_no
    """), {"header_id": header_id}).fetchall()

    # Get cumulative received from inward_receipt + inward_lines tables
    prev_lines = db.execute(text("""
        SELECT il.sl_no,
               COALESCE(SUM(il.finished_goods_kgs), 0),
               COALESCE(SUM(il.finished_goods_boxes), 0),
               COALESCE(SUM(il.waste_kgs), 0),
               COALESCE(SUM(il.rejection_kgs), 0)
        FROM jb_work_inward_lines il
        JOIN jb_work_inward_receipt ir ON ir.id = il.inward_receipt_id
        WHERE ir.header_id = :header_id
        GROUP BY il.sl_no
    """), {"header_id": header_id}).fetchall()

    cumulative = {}
    for pl in prev_lines:
        cumulative[pl[0]] = {
            "fg_kgs": float(pl[1]), "fg_boxes": int(pl[2]),
            "waste_kgs": float(pl[3]), "rejection_kgs": float(pl[4]),
        }

    ir_count_row = db.execute(text(
        "SELECT COUNT(*) FROM jb_work_inward_receipt WHERE header_id = :header_id"
    ), {"header_id": header_id}).fetchone()
    ir_count = ir_count_row[0] if ir_count_row else 0

    # Get prior IR history (dates, challan, receipt type, totals)
    prior_irs = []
    ir_rows = db.execute(text("""
        SELECT ir.ir_number, ir.challan_no, ir.receipt_date, ir.receipt_type, ir.created_at,
               COALESCE(SUM(il.finished_goods_kgs), 0) as total_fg,
               COALESCE(SUM(il.waste_kgs), 0) as total_waste,
               COALESCE(SUM(il.rejection_kgs), 0) as total_rejection
        FROM jb_work_inward_receipt ir
        LEFT JOIN jb_work_inward_lines il ON il.inward_receipt_id = ir.id
        WHERE ir.header_id = :header_id
        GROUP BY ir.id, ir.ir_number, ir.challan_no, ir.receipt_date, ir.receipt_type, ir.created_at
        ORDER BY ir.created_at ASC
    """), {"header_id": header_id}).fetchall()
    for ir_row in ir_rows:
        prior_irs.append({
            "ir_number": ir_row[0] or "",
            "challan_no": ir_row[1] or "",
            "receipt_date": ir_row[2] or "",
            "receipt_type": ir_row[3] or "partial",
            "total_fg_kgs": round(float(ir_row[5]), 3),
            "total_waste_kgs": round(float(ir_row[6]), 3),
            "total_rejection_kgs": round(float(ir_row[7]), 3),
        })

    # Get sub_category (process type) from the already-fetched row (index 7)
    process_type = row[7] or ""

    # Try to get loss config from most recent inward line for this header
    loss_config = None
    if header_id:
        lc_row = db.execute(text("""
            SELECT il.process_type, il.min_loss_pct, il.max_loss_pct, il.waste_with_partial, il.single_shot
            FROM jb_work_inward_lines il
            JOIN jb_work_inward_receipt ir ON ir.id = il.inward_receipt_id
            WHERE ir.header_id = :header_id AND il.process_type IS NOT NULL AND il.process_type != ''
            ORDER BY il.id DESC LIMIT 1
        """), {"header_id": header_id}).fetchone()
        if lc_row:
            loss_config = {
                "min_loss_pct": float(lc_row[1]), "max_loss_pct": float(lc_row[2]),
                "loss_component": "",
                "waste_with_partial": bool(lc_row[3]), "single_shot": bool(lc_row[4]),
            }
    line_items = []
    for l in lines:
        sl_no = l[0]
        prev = cumulative.get(sl_no, {"fg_kgs": 0, "fg_boxes": 0, "waste_kgs": 0, "rejection_kgs": 0})
        line_items.append({
            "sl_no": sl_no,
            "item_description": l[1] or "",
            "quantity_kgs": float(l[2] or 0),
            "quantity_boxes": int(l[3] or 0),
            "net_weight": str(l[4] or "0"),
            "prev_fg_kgs": round(prev["fg_kgs"], 3),
            "prev_fg_boxes": prev["fg_boxes"],
            "prev_waste_kgs": round(prev["waste_kgs"], 3),
            "prev_rejection_kgs": round(prev["rejection_kgs"], 3),
        })

    return {
        "record": {
            "id": row[0],
            "challan_no": row[1],
            "job_work_date": row[2] or "",
            "from_warehouse": row[3] or "",
            "to_party": row[4] or "",
            "status": row[5] or "sent",
            "vehicle_no": row[6] or "",
            "sub_category": process_type,
        },
        "line_items": line_items,
        "receive_count": ir_count,
        "loss_config": loss_config,
        "prior_irs": prior_irs,
    }


# ════════════════════════════════════════════════════════════
#  GET /job-work/out-by-id/{record_id}  — Get single record by ID (for edit)
# ════════════════════════════════════════════════════════════

@router.get("/out-by-id/{record_id}")
def get_material_out_by_id(
    record_id: int,
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    # Use column_name based access for safety
    row = db.execute(text("""
        SELECT * FROM jb_materialout_header WHERE id = :id
    """), {"id": record_id}).mappings().first()

    if not row:
        raise HTTPException(status_code=404, detail="Record not found.")

    dispatch_to = {}
    raw_dt = row.get("dispatch_to", None)
    if raw_dt:
        try:
            dispatch_to = json.loads(raw_dt) if isinstance(raw_dt, str) else raw_dt
        except Exception:
            dispatch_to = {}

    # Also try payload JSON for extra fields
    payload_data = {}
    raw_payload = row.get("payload", None)
    if raw_payload:
        try:
            payload_data = json.loads(raw_payload) if isinstance(raw_payload, str) else raw_payload
        except Exception:
            payload_data = {}

    lines = db.execute(text("""
        SELECT * FROM jb_materialout_lines
        WHERE header_id = :header_id ORDER BY sl_no
    """), {"header_id": row["id"]}).mappings().all()

    def safe(d, key, default=""):
        try:
            v = d.get(key, default)
            return v if v is not None else default
        except Exception:
            return default

    return {
        "id": row["id"],
        "challan_no": safe(row, "challan_no"),
        "job_work_date": safe(row, "job_work_date"),
        "from_warehouse": safe(row, "from_warehouse"),
        "to_party": safe(row, "to_party"),
        "status": safe(row, "status", "sent"),
        "vehicle_no": safe(row, "vehicle_no"),
        "sub_category": safe(row, "sub_category"),
        "dispatch_to": dispatch_to,
        "driver_name": safe(row, "driver_name"),
        "authorized_person": safe(row, "authorized_person"),
        "remarks": safe(row, "remarks"),
        "party_address": safe(row, "party_address"),
        "e_way_bill_no": safe(row, "e_way_bill_no"),
        "dispatched_through": safe(row, "dispatched_through"),
        "items": [
            {
                "sl_no": safe(l, "sl_no", 0),
                "item_description": safe(l, "item_description"),
                "sub_category": safe(l, "sub_category"),
                "uom": safe(l, "uom", "KG"),
                "quantity_boxes": int(safe(l, "quantity_boxes", 0) or 0),
                "net_weight": float(safe(l, "net_weight", 0) or 0),
                "total_weight": float(safe(l, "total_weight", 0) or 0),
                "lot_number": safe(l, "lot_number"),
                "remarks": safe(l, "line_remarks") or safe(l, "remarks"),
                "material_type": safe(l, "material_type"),
                "item_category": safe(l, "item_category"),
                "case_pack": str(safe(l, "case_pack", "0")),
                "batch_number": safe(l, "batch_number"),
                "manufacturing_date": safe(l, "manufacturing_date"),
                "expiry_date": safe(l, "expiry_date"),
                "rate_per_kg": float(safe(l, "rate_per_kg", 0) or 0),
                "amount": float(safe(l, "amount", 0) or 0),
                "cold_unit": safe(l, "cold_unit"),
                "item_mark": safe(l, "item_mark"),
                "box_id": safe(l, "box_id"),
                "transaction_no": safe(l, "transaction_no"),
            }
            for l in lines
        ],
    }


# ════════════════════════════════════════════════════════════
#  GET /job-work/out/{challan_no}  — Get single record by challan (for DC print)
# ════════════════════════════════════════════════════════════

@router.get("/out/{challan_no}")
def get_material_out_by_challan(
    challan_no: str,
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    row = db.execute(text("""
        SELECT h.id, h.challan_no, h.job_work_date, h.from_warehouse, h.to_party,
               h.status, h.vehicle_no, h.sub_category, h.dispatch_to, h.payload,
               h.driver_name, h.authorized_person, h.remarks, h.party_address,
               h.purpose_of_work, h.contact_person, h.contact_number, h.expected_return_date
        FROM jb_materialout_header h
        WHERE LOWER(h.challan_no) = LOWER(:challan_no)
        ORDER BY h.created_at DESC
        LIMIT 1
    """), {"challan_no": challan_no.strip()}).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="No matching material out record found.")

    header_id = row[0]

    # Parse dispatch_to JSON
    dispatch_to = {}
    if row[8]:
        try:
            dispatch_to = json.loads(row[8]) if isinstance(row[8], str) else row[8]
        except Exception:
            dispatch_to = {}

    lines = db.execute(text("""
        SELECT sl_no, item_description, sub_category, uom,
               quantity_boxes, net_weight, total_weight, lot_number, line_remarks,
               rate_per_kg, amount, material_type, item_category,
               batch_number, manufacturing_date, expiry_date,
               box_id, transaction_no, cold_unit, item_mark
        FROM jb_materialout_lines
        WHERE header_id = :header_id
        ORDER BY sl_no
    """), {"header_id": header_id}).fetchall()

    return {
        "id": row[0],
        "challan_no": row[1],
        "job_work_date": row[2] or "",
        "from_warehouse": row[3] or "",
        "to_party": row[4] or "",
        "status": row[5] or "sent",
        "vehicle_no": row[6] or "",
        "sub_category": row[7] or "",
        "dispatch_to": dispatch_to,
        "driver_name": row[10] or "",
        "authorized_person": row[11] or "",
        "remarks": row[12] or "",
        "party_address": row[13] or "",
        "purpose_of_work": row[14] or "",
        "contact_person": row[15] or "",
        "contact_number": row[16] or "",
        "expected_return_date": row[17] or "",
        "items": [
            {
                "sl_no": l[0],
                "item_description": l[1] or "",
                "sub_category": l[2] or "",
                "uom": l[3] or "KG",
                "quantity_boxes": int(l[4] or 0),
                "net_weight": float(l[5] or 0),
                "total_weight": float(l[6] or 0),
                "lot_number": l[7] or "",
                "remarks": l[8] or "",
                "rate_per_kg": float(l[9] or 0),
                "amount": float(l[10] or 0),
                "material_type": l[11] or "",
                "item_category": l[12] or "",
                "batch_number": l[13] or "",
                "manufacturing_date": l[14] or "",
                "expiry_date": l[15] or "",
                "box_id": l[16] or "",
                "transaction_no": l[17] or "",
                "cold_unit": l[18] or "",
                "item_mark": l[19] or "",
            }
            for l in lines
        ],
    }


# ════════════════════════════════════════════════════════════
#  POST /job-work/material-in  — Submit Material In
# ════════════════════════════════════════════════════════════

@router.post("/material-in")
def submit_material_in(
    payload: dict,
    created_by: str = Query(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    header_id = payload.get("original_record_id")
    original_challan_no = payload.get("original_challan_no", "")
    receipt_type = payload.get("receipt_type", "partial")
    items = payload.get("items", [])

    # Generate IR number: IR-{challan}-{sequence}
    ir_seq = db.execute(text(
        "SELECT COUNT(*) + 1 FROM jb_work_inward_receipt WHERE header_id = :id"
    ), {"id": header_id}).fetchone()
    seq = ir_seq[0] if ir_seq else 1
    ir_number = f"IR-{original_challan_no}-{str(seq).zfill(2)}"

    # Insert inward receipt header
    challan_no = payload.get("challan_no", "").strip()
    ir_result = db.execute(text("""
        INSERT INTO jb_work_inward_receipt
            (ir_number, challan_no, header_id, receipt_date, receipt_type, vehicle_no, driver_name, remarks, created_by, inward_warehouse)
        VALUES
            (:ir_number, :challan_no, :header_id, :receipt_date, :receipt_type, :vehicle_no, :driver_name, :remarks, :created_by, :inward_warehouse)
        RETURNING id
    """), {
        "ir_number": ir_number,
        "challan_no": challan_no,
        "header_id": header_id,
        "receipt_date": _parse_date_to_iso(payload.get("received_date", "")),
        "receipt_type": receipt_type,
        "vehicle_no": payload.get("vehicle_no", ""),
        "driver_name": payload.get("driver_name", ""),
        "remarks": payload.get("remarks", ""),
        "created_by": created_by,
        "inward_warehouse": payload.get("inward_warehouse", ""),
    })
    inward_receipt_id = ir_result.fetchone()[0]

    # Fetch sub_category (process type) from the material-out header
    header_process_type = ""
    if header_id:
        pt_row = db.execute(text("SELECT sub_category FROM jb_materialout_header WHERE id = :id"), {"id": header_id}).fetchone()
        if pt_row:
            header_process_type = pt_row[0] or ""

    # Insert inward receipt line items (including loss config per line)
    loss_config = payload.get("loss_config") or {}
    for item in items:
        db.execute(text("""
            INSERT INTO jb_work_inward_lines
                (inward_receipt_id, sl_no, item_description, sent_kgs, sent_boxes,
                 finished_goods_kgs, finished_goods_boxes, waste_kgs, waste_type,
                 rejection_kgs, rejection_boxes, line_remarks,
                 process_type, min_loss_pct, max_loss_pct, waste_with_partial, single_shot)
            VALUES
                (:ir_id, :sl_no, :description, :sent_kgs, :sent_boxes,
                 :fg_kgs, :fg_boxes, :waste_kgs, :waste_type,
                 :rejection_kgs, :rejection_boxes, :remarks,
                 :process_type, :min_loss_pct, :max_loss_pct, :waste_with_partial, :single_shot)
        """), {
            "ir_id": inward_receipt_id,
            "sl_no": item.get("sl_no", 0),
            "description": item.get("description", ""),
            "sent_kgs": float(item.get("sent_kgs", 0)),
            "sent_boxes": int(item.get("sent_boxes", 0)),
            "fg_kgs": float(item.get("finished_goods_kgs", 0)),
            "fg_boxes": int(item.get("finished_goods_boxes", 0)),
            "waste_kgs": float(item.get("waste_kgs", 0)),
            "waste_type": item.get("waste_type", ""),
            "rejection_kgs": float(item.get("rejection_kgs", 0)),
            "rejection_boxes": int(item.get("rejection_boxes", 0)),
            "remarks": item.get("line_remarks", ""),
            "process_type": item.get("process_type") or loss_config.get("process_type", "") or header_process_type,
            "min_loss_pct": float(item.get("min_loss_pct") or loss_config.get("min_loss_pct", 0)),
            "max_loss_pct": float(item.get("max_loss_pct") or loss_config.get("max_loss_pct", 0)),
            "waste_with_partial": bool(item.get("waste_with_partial") if item.get("waste_with_partial") is not None else loss_config.get("waste_with_partial", True)),
            "single_shot": bool(item.get("single_shot") if item.get("single_shot") is not None else loss_config.get("single_shot", False)),
        })

    # Insert box entries (from Article Entry section)
    boxes = payload.get("boxes", [])
    inward_warehouse = payload.get("inward_warehouse", "")
    for box in boxes:
        db.execute(text("""
            INSERT INTO jb_inward_boxes
                (inward_receipt_id, transaction_no, box_id, box_number,
                 item_description, item_group, sub_group,
                 net_weight, gross_weight, inward_warehouse,
                 vakkal, lot_no, item_mark, storage_location, exporter, rate, spl_remarks)
            VALUES
                (:ir_id, :transaction_no, :box_id, :box_number,
                 :item_description, :item_group, :sub_group,
                 :net_weight, :gross_weight, :inward_warehouse,
                 :vakkal, :lot_no, :item_mark, :storage_location, :exporter, :rate, :spl_remarks)
            ON CONFLICT (transaction_no, box_id) DO NOTHING
        """), {
            "ir_id": inward_receipt_id,
            "transaction_no": box.get("transaction_no", ""),
            "box_id": box.get("box_id", ""),
            "box_number": int(box.get("box_number", 0)),
            "item_description": box.get("item_description", ""),
            "item_group": box.get("item_group", ""),
            "sub_group": box.get("sub_group", ""),
            "net_weight": float(box.get("net_weight", 0)),
            "gross_weight": float(box.get("gross_weight", 0)),
            "inward_warehouse": inward_warehouse,
            "vakkal": box.get("vakkal", ""),
            "lot_no": box.get("lot_no", ""),
            "item_mark": box.get("item_mark", ""),
            "storage_location": box.get("storage_location", ""),
            "exporter": box.get("exporter", ""),
            "rate": float(box.get("rate", 0)),
            "spl_remarks": box.get("spl_remarks", ""),
        })

    # Insert into cfpl_cold_stocks / cdpl_cold_stocks if cold storage material-in
    cold_company = payload.get("cold_company", "")
    cold_inward_date = payload.get("cold_inward_date", "")
    if cold_company and boxes:
        cold_table = "cfpl_cold_stocks" if cold_company == "cfpl" else "cdpl_cold_stocks"
        for box in boxes:
            weight = float(box.get("net_weight", 0))
            rate_val = float(box.get("rate", 0))
            value_val = weight * rate_val if rate_val > 0 else 0
            db.execute(text(f"""
                INSERT INTO {cold_table}
                    (inward_dt, unit, inward_no, item_description, item_mark,
                     vakkal, lot_no, no_of_cartons, weight_kg,
                     total_inventory_kgs, group_name, item_subgroup, storage_location,
                     exporter, last_purchase_rate, value,
                     box_id, transaction_no, spl_remarks)
                VALUES
                    (:inward_dt, :unit, :inward_no, :item_description, :item_mark,
                     :vakkal, :lot_no, 1, :weight_kg,
                     :weight_kg, :group_name, :item_subgroup, :storage_location,
                     :exporter, :last_purchase_rate, :value,
                     :box_id, :transaction_no, :spl_remarks)
            """), {
                "inward_dt": _parse_date_to_iso(cold_inward_date or payload.get("received_date", "")),
                "unit": box.get("storage_location", ""),
                "inward_no": payload.get("challan_no", ""),
                "item_description": box.get("item_description", ""),
                "item_mark": box.get("item_mark", ""),
                "vakkal": box.get("vakkal", ""),
                "lot_no": box.get("lot_no", ""),
                "weight_kg": weight,
                "group_name": box.get("item_group", ""),
                "item_subgroup": box.get("sub_group", ""),
                "storage_location": box.get("storage_location", ""),
                "exporter": box.get("exporter", ""),
                "last_purchase_rate": rate_val,
                "value": value_val,
                "box_id": box.get("box_id", ""),
                "transaction_no": box.get("transaction_no", ""),
                "spl_remarks": box.get("spl_remarks", ""),
            })

    # Update JWO status based on receipt_type and cumulative totals
    if header_id:
        if receipt_type == "final":
            new_status = "fully_received"
        else:
            new_status = "partially_received"

        db.execute(text("""
            UPDATE jb_materialout_header SET status = :status, updated_at = NOW() WHERE id = :id
        """), {"id": header_id, "status": new_status})

    db.commit()
    return {
        "status": "success",
        "ir_number": ir_number,
        "message": f"Inward Receipt {ir_number} recorded against {original_challan_no}",
    }


# ════════════════════════════════════════════════════════════
#  GET /job-work/material-in/list  — List all inward receipts
# ════════════════════════════════════════════════════════════

@router.get("/material-in/list")
def list_material_in(
    page: int = Query(1),
    per_page: int = Query(15),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    offset = (page - 1) * per_page

    total_row = db.execute(text("SELECT COUNT(*) FROM jb_work_inward_receipt")).fetchone()
    total = total_row[0] if total_row else 0

    rows = db.execute(text("""
        SELECT ir.id, ir.ir_number, ir.challan_no, ir.receipt_date, ir.receipt_type,
               ir.vehicle_no, ir.driver_name, ir.remarks, ir.created_by, ir.created_at,
               h.challan_no as jwo_challan, h.to_party, h.sub_category,
               COALESCE(SUM(il.finished_goods_kgs), 0) as total_fg,
               COALESCE(SUM(il.waste_kgs), 0) as total_waste,
               COALESCE(SUM(il.rejection_kgs), 0) as total_rejection,
               COALESCE(SUM(il.sent_kgs), 0) as total_sent,
               (SELECT string_agg(DISTINCT il2.item_description, ', ')
                FROM jb_work_inward_lines il2 WHERE il2.inward_receipt_id = ir.id) as item_descriptions
        FROM jb_work_inward_receipt ir
        LEFT JOIN jb_materialout_header h ON h.id = ir.header_id
        LEFT JOIN jb_work_inward_lines il ON il.inward_receipt_id = ir.id
        GROUP BY ir.id, ir.ir_number, ir.challan_no, ir.receipt_date, ir.receipt_type,
                 ir.vehicle_no, ir.driver_name, ir.remarks, ir.created_by, ir.created_at,
                 h.challan_no, h.to_party, h.sub_category
        ORDER BY ir.created_at DESC
        LIMIT :limit OFFSET :offset
    """), {"limit": per_page, "offset": offset}).fetchall()

    records = []
    for r in rows:
        records.append({
            "id": r[0],
            "ir_number": r[1] or "",
            "challan_no": r[2] or "",
            "receipt_date": r[3] or "",
            "receipt_type": r[4] or "partial",
            "vehicle_no": r[5] or "",
            "driver_name": r[6] or "",
            "remarks": r[7] or "",
            "created_by": r[8] or "",
            "created_at": str(r[9] or ""),
            "jwo_challan": r[10] or "",
            "to_party": r[11] or "",
            "sub_category": r[12] or "",
            "total_fg_kgs": round(float(r[13]), 3),
            "total_waste_kgs": round(float(r[14]), 3),
            "total_rejection_kgs": round(float(r[15]), 3),
            "total_sent_kgs": round(float(r[16]), 3),
            "item_descriptions": r[17] or "",
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": max(1, -(-total // per_page)),
    }


# ════════════════════════════════════════════════════════════
#  GET /job-work/material-in/{id}  — View inward receipt details
# ════════════════════════════════════════════════════════════

@router.get("/material-in/{ir_id}")
def get_material_in_detail(
    ir_id: int,
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    # Fetch receipt header
    row = db.execute(text("""
        SELECT ir.id, ir.ir_number, ir.challan_no, ir.header_id, ir.receipt_date,
               ir.receipt_type, ir.vehicle_no, ir.driver_name, ir.remarks, ir.created_by, ir.created_at,
               h.challan_no as jwo_challan, h.to_party, h.sub_category, h.from_warehouse
        FROM jb_work_inward_receipt ir
        LEFT JOIN jb_materialout_header h ON h.id = ir.header_id
        WHERE ir.id = :id
    """), {"id": ir_id}).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Inward receipt not found")

    # Fetch line items
    lines = db.execute(text("""
        SELECT sl_no, item_description, sent_kgs, sent_boxes,
               finished_goods_kgs, finished_goods_boxes, waste_kgs, waste_type,
               rejection_kgs, rejection_boxes, line_remarks, process_type
        FROM jb_work_inward_lines
        WHERE inward_receipt_id = :ir_id
        ORDER BY sl_no
    """), {"ir_id": ir_id}).fetchall()

    return {
        "receipt": {
            "id": row[0],
            "ir_number": row[1],
            "challan_no": row[2] or "",
            "header_id": row[3],
            "receipt_date": row[4] or "",
            "receipt_type": row[5] or "partial",
            "vehicle_no": row[6] or "",
            "driver_name": row[7] or "",
            "remarks": row[8] or "",
            "created_by": row[9] or "",
            "created_at": str(row[10]) if row[10] else "",
            "jwo_challan": row[11] or "",
            "to_party": row[12] or "",
            "process_type": row[13] or "",
            "from_warehouse": row[14] or "",
        },
        "lines": [{
            "sl_no": l[0],
            "item_description": l[1] or "",
            "sent_kgs": float(l[2] or 0),
            "sent_boxes": int(l[3] or 0),
            "finished_goods_kgs": float(l[4] or 0),
            "finished_goods_boxes": int(l[5] or 0),
            "waste_kgs": float(l[6] or 0),
            "waste_type": l[7] or "",
            "rejection_kgs": float(l[8] or 0),
            "rejection_boxes": int(l[9] or 0),
            "line_remarks": l[10] or "",
            "process_type": l[11] or "",
        } for l in lines],
    }


# ════════════════════════════════════════════════════════════
#  DELETE /job-work/material-in/{id}  — Delete an inward receipt
# ════════════════════════════════════════════════════════════

DELETE_ALLOWED_EMAILS = ["b.hrithik@candorfoods.in", "yash@candorfoods.in"]

@router.delete("/material-in/{ir_id}")
def delete_material_in(
    ir_id: int,
    user_email: str = Query(default=""),
    db: Session = Depends(get_db),
):
    if user_email not in DELETE_ALLOWED_EMAILS:
        raise HTTPException(status_code=403, detail="You do not have permission to delete records.")
    _ensure_tables(db)

    row = db.execute(text(
        "SELECT id, ir_number, header_id FROM jb_work_inward_receipt WHERE id = :id"
    ), {"id": ir_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Inward receipt not found.")

    header_id = row[2]

    # Delete matching rows from cold storage tables (material-in inserts into these)
    cold_boxes = db.execute(text(
        "SELECT box_id, transaction_no FROM jb_inward_boxes WHERE inward_receipt_id = :id"
    ), {"id": ir_id}).fetchall()
    for cb in cold_boxes:
        box_id, txn_no = cb[0], cb[1]
        if box_id and txn_no:
            for cold_tbl in ("cfpl_cold_stocks", "cdpl_cold_stocks"):
                db.execute(text(
                    f"DELETE FROM {cold_tbl} WHERE box_id = :box_id AND transaction_no = :txn_no"
                ), {"box_id": box_id, "txn_no": txn_no})

    db.execute(text("DELETE FROM jb_inward_boxes WHERE inward_receipt_id = :id"), {"id": ir_id})
    db.execute(text("DELETE FROM jb_work_inward_lines WHERE inward_receipt_id = :id"), {"id": ir_id})
    db.execute(text("DELETE FROM jb_work_inward_receipt WHERE id = :id"), {"id": ir_id})

    # Recalculate JWO status
    if header_id:
        remaining = db.execute(text(
            "SELECT COUNT(*) FROM jb_work_inward_receipt WHERE header_id = :id"
        ), {"id": header_id}).fetchone()
        new_status = "partially_received" if remaining and remaining[0] > 0 else "sent"
        db.execute(text(
            "UPDATE jb_materialout_header SET status = :status, updated_at = NOW() WHERE id = :id"
        ), {"id": header_id, "status": new_status})

    db.commit()
    return {"status": "success", "message": f"Inward receipt {row[1]} deleted."}


# ════════════════════════════════════════════════════════════
#  DELETE /job-work/{id}  — Delete a job work record
# ════════════════════════════════════════════════════════════

@router.delete("/{record_id}")
def delete_job_work_record(
    record_id: int,
    user_email: str = Query(default=""),
    db: Session = Depends(get_db),
):
    if user_email not in DELETE_ALLOWED_EMAILS:
        raise HTTPException(status_code=403, detail="You do not have permission to delete records.")
    _ensure_tables(db)

    row = db.execute(text("SELECT id, challan_no FROM jb_materialout_header WHERE id = :id"), {"id": record_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Record not found.")

    # Restore cold storage stock before deleting line items
    _restore_cold_storage_stock(db, record_id)

    # Delete cold storage rows added by material-in for this header
    inward_boxes = db.execute(text("""
        SELECT b.box_id, b.transaction_no FROM jb_inward_boxes b
        JOIN jb_work_inward_receipt r ON r.id = b.inward_receipt_id
        WHERE r.header_id = :id
    """), {"id": record_id}).fetchall()
    for cb in inward_boxes:
        box_id, txn_no = cb[0], cb[1]
        if box_id and txn_no:
            for cold_tbl in ("cfpl_cold_stocks", "cdpl_cold_stocks"):
                db.execute(text(
                    f"DELETE FROM {cold_tbl} WHERE box_id = :box_id AND transaction_no = :txn_no"
                ), {"box_id": box_id, "txn_no": txn_no})

    # Delete inward boxes, inward lines, inward receipts, then lines, then header
    db.execute(text("""
        DELETE FROM jb_inward_boxes WHERE inward_receipt_id IN
        (SELECT id FROM jb_work_inward_receipt WHERE header_id = :id)
    """), {"id": record_id})
    db.execute(text("""
        DELETE FROM jb_work_inward_lines WHERE inward_receipt_id IN
        (SELECT id FROM jb_work_inward_receipt WHERE header_id = :id)
    """), {"id": record_id})
    db.execute(text("DELETE FROM jb_work_inward_receipt WHERE header_id = :id"), {"id": record_id})
    db.execute(text("DELETE FROM jb_materialout_lines WHERE header_id = :id"), {"id": record_id})
    db.execute(text("DELETE FROM jb_materialout_header WHERE id = :id"), {"id": record_id})
    db.commit()

    return {"status": "success", "message": f"Record {row[1]} deleted successfully"}


# ════════════════════════════════════════════════════════════
#  GET /job-work/{id}  — Get a single job work record (for edit)
# ════════════════════════════════════════════════════════════

@router.get("/{record_id}")
def get_job_work_record(
    record_id: int,
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    row = db.execute(text("""
        SELECT id, challan_no, job_work_date, from_warehouse, to_party, party_address,
               contact_person, contact_number, purpose_of_work, expected_return_date,
               vehicle_no, driver_name, authorized_person, remarks,
               e_way_bill_no, dispatched_through, type, status, dispatch_to, payload,
               created_by, created_at
        FROM jb_materialout_header WHERE id = :id
    """), {"id": record_id}).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Record not found.")

    lines = db.execute(text("""
        SELECT sl_no, item_description, material_type, item_category, sub_category,
               quantity_kgs, quantity_boxes, rate_per_kg, amount,
               uom, case_pack, net_weight, total_weight,
               batch_number, lot_number, manufacturing_date, expiry_date, line_remarks
        FROM jb_materialout_lines WHERE header_id = :id ORDER BY sl_no
    """), {"id": record_id}).fetchall()

    dispatch_to = row[18]
    if isinstance(dispatch_to, str):
        try:
            dispatch_to = json.loads(dispatch_to)
        except Exception:
            dispatch_to = {}

    return {
        "id": row[0],
        "challan_no": row[1],
        "job_work_date": row[2] or "",
        "from_warehouse": row[3] or "",
        "to_party": row[4] or "",
        "party_address": row[5] or "",
        "vehicle_no": row[10] or "",
        "driver_name": row[11] or "",
        "remarks": row[13] or "",
        "e_way_bill_no": row[14] or "",
        "dispatched_through": row[15] or "",
        "type": row[16] or "OUT",
        "status": row[17] or "sent",
        "dispatch_to": dispatch_to or {},
        "line_items": [
            {
                "sl_no": l[0],
                "item_description": l[1] or "",
                "material_type": l[2] or "",
                "item_category": l[3] or "",
                "sub_category": l[4] or "",
                "quantity_kgs": float(l[5] or 0),
                "quantity_boxes": int(l[6] or 0),
                "rate_per_kg": float(l[7] or 0),
                "amount": float(l[8] or 0),
                "uom": l[9] or "",
                "case_pack": str(l[10] or ""),
                "net_weight": str(l[11] or ""),
                "total_weight": str(l[12] or ""),
                "batch_number": l[13] or "",
                "lot_number": l[14] or "",
                "manufacturing_date": l[15] or "",
                "expiry_date": l[16] or "",
                "line_remarks": l[17] or "",
            }
            for l in lines
        ],
    }


def _cell(ws, row, col):
    """Get cell value safely."""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def _parse_material_out_excel(file_bytes: bytes) -> dict:
    """
    Parse a Material Out Excel file (openpyxl) and return the same
    JSON structure as the PDF AI extractor — no Anthropic API needed.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    max_row = ws.max_row

    # ── Header fields ──
    # Company is in Row 3 Col A (we don't need it, it's our own company)
    challan_no = _cell(ws, 4, 7)          # Row 4, Col G
    e_way_bill_no = _cell(ws, 4, 9)       # Row 4, Col I
    dated = _cell(ws, 4, 10)              # Row 4, Col J
    mode_terms = _cell(ws, 6, 10)         # Row 6, Col J
    motor_vehicle_no = _cell(ws, 12, 10)  # Row 12, Col J

    # Reference No & Date — Row 8, Col G (e.g., "SR. NO.: 1170  dt. 11-Mar-26")
    ref_raw = _cell(ws, 8, 7)
    reference_no = ""
    reference_date = ""
    if ref_raw:
        # Try to split "SR. NO.: 1170  dt. 11-Mar-26"
        dt_match = re.search(r'dt\.\s*(.+)', ref_raw)
        if dt_match:
            reference_date = dt_match.group(1).strip()
            reference_no = ref_raw[:dt_match.start()].strip().rstrip('.')
        else:
            reference_no = ref_raw

    dispatch_doc_no_raw = _cell(ws, 7, 10)
    dispatch_doc_no = "" if dispatch_doc_no_raw in ("Other References", "Destination") else dispatch_doc_no_raw
    dispatched_through = _cell(ws, 11, 7) if _cell(ws, 11, 7) != "Dispatched through" else ""
    other_references = _cell(ws, 8, 10) if _cell(ws, 8, 10) else ""
    destination = _cell(ws, 10, 10) if _cell(ws, 10, 10) else ""
    duration_of_process = _cell(ws, 13, 10) if _cell(ws, 13, 10) != "Duration of Process" else ""
    bill_of_lading = _cell(ws, 13, 7) if _cell(ws, 13, 7) != "Bill of Lading/LR-RR No." else ""

    # Date & Time of Issue — search rows 14-16 Col G
    date_time_of_issue = ""
    for r in range(14, 17):
        v = _cell(ws, r, 7)
        if v and re.search(r'\d{1,2}-\w{3}-\d{2}.*at', v):
            date_time_of_issue = v
            break

    # Nature of Processing — row after "Nature of Processing" label
    nature_of_processing = ""
    for r in range(15, 20):
        if "Nature of Processing" in _cell(ws, r, 7):
            nature_of_processing = _cell(ws, r + 1, 7)
            break

    # ── Helper: parse name + address + state block starting at a given row ──
    def _parse_address_block(start_row, max_rows=8):
        name = _cell(ws, start_row, 1)
        addr_lines = []
        state = ""
        state_code = ""
        gstin = ""
        for r in range(start_row + 1, start_row + max_rows):
            v = _cell(ws, r, 1)
            if not v or "Sl" == v:
                break
            if "GSTIN" in v:
                gstin = _cell(ws, r, 4) or v.replace("GSTIN/UIN:", "").replace("GSTIN/UIN :", "").strip()
                continue
            if v.startswith("State Name"):
                state_val = _cell(ws, r, 4) or v.replace("State Name :", "").replace("State Name:", "").strip()
                code_match = re.search(r'Code\s*:\s*(\d+)', state_val)
                if code_match:
                    state_code = code_match.group(1)
                    state = state_val[:code_match.start()].strip().rstrip(',')
                else:
                    state = state_val
                break
            # Stop if we hit another label
            if v in ("Party", "Dispatch To", "Bill of Lading/LR-RR No."):
                break
            addr_lines.append(v)
        return name, "\n".join(addr_lines), state or "Maharashtra", state_code or "27", gstin

    # ── Dispatch To (consignee) — search for "Dispatch To" label ──
    dispatch_name = ""
    dispatch_addr = ""
    dispatch_state = "Maharashtra"
    dispatch_state_code = "27"
    dispatch_gstin = ""
    for r in range(10, 20):
        if _cell(ws, r, 1) == "Dispatch To":
            dispatch_name, dispatch_addr, dispatch_state, dispatch_state_code, dispatch_gstin = _parse_address_block(r + 1)
            break

    # ── Motor Vehicle No — search for label ──
    for r in range(10, 16):
        if "Motor Vehicle" in _cell(ws, r, 10):
            mv = _cell(ws, r + 1, 10)
            if mv:
                motor_vehicle_no = mv
            break

    # ── Party details — search for "Party" label ──
    party_name = ""
    party_addr = ""
    party_state = "Maharashtra"
    party_state_code = "27"
    party_gstin = ""
    for r in range(14, 26):
        if _cell(ws, r, 1) == "Party":
            party_name, party_addr, party_state, party_state_code, party_gstin = _parse_address_block(r + 1)
            break

    # ── Find item table header row (Sl / No.) ──
    header_row = 0
    for r in range(20, max_row):
        if _cell(ws, r, 1) == "Sl" and "Description" in _cell(ws, r, 2):
            header_row = r
            break
    if not header_row:
        # Fallback: look for "No." in col A
        for r in range(20, max_row):
            if _cell(ws, r, 1) == "No.":
                header_row = r - 1
                break

    # ── Parse line items ──
    line_items = []
    remarks_text = ""
    if header_row:
        data_start = header_row + 2  # Skip header + "No." row
        r = data_start
        while r <= max_row:
            sl_val = ws.cell(row=r, column=1).value
            # Check if this is a "Total" row — stop parsing
            desc_val = _cell(ws, r, 2)
            if desc_val == "Total":
                break

            # Item row starts with a number in Col A
            if isinstance(sl_val, (int, float)) and sl_val > 0:
                description = _cell(ws, r, 2)
                hsn_sac = _cell(ws, r, 8)
                gst_rate_val = ws.cell(row=r, column=9).value
                gst_rate = f"{int(gst_rate_val)}%" if isinstance(gst_rate_val, (int, float)) else str(gst_rate_val or "0%")
                quantity = ws.cell(row=r, column=10).value or 0
                rate = ws.cell(row=r, column=11).value or 0
                per_uom = _cell(ws, r, 12)  # "Kgs", "NOS", etc.
                amount = ws.cell(row=r, column=13).value or 0

                # Next row(s) may have: weight in kgs (col 8), boxes info (col 2), variant (col 2)
                weight_kgs = 0
                quantity_boxes = 0
                item_remarks = ""
                sub_r = r + 1
                while sub_r <= max_row:
                    sub_sl = ws.cell(row=sub_r, column=1).value
                    sub_desc = _cell(ws, sub_r, 2)
                    sub_col8 = ws.cell(row=sub_r, column=8).value

                    # If next row has a number in Col A or is "Total", it's a new item
                    if isinstance(sub_sl, (int, float)) and sub_sl > 0:
                        break
                    if sub_desc == "Total":
                        break

                    # Weight in kgs — numeric value in col 8 on sub-row
                    if isinstance(sub_col8, (int, float)) and sub_col8 > 0:
                        weight_kgs = float(sub_col8)

                    # Boxes info — e.g., "100 Boxes", "16units X 60 Ctns"
                    if sub_desc:
                        box_match = re.search(r'(\d+)\s*(Box|Boxes|Ctns|Carton)', sub_desc, re.IGNORECASE)
                        if box_match:
                            quantity_boxes = int(box_match.group(1))
                            item_remarks = sub_desc
                        elif sub_desc and not sub_desc.startswith("Total"):
                            # Could be variant/mark info
                            if item_remarks:
                                item_remarks += " | " + sub_desc
                            else:
                                item_remarks = sub_desc

                    sub_r += 1

                # Determine quantity_kgs vs quantity_nos
                quantity_kgs = 0.0
                quantity_nos = 0
                if per_uom.upper() in ("KGS", "KG", "KGS."):
                    quantity_kgs = float(quantity)
                elif per_uom.upper() in ("NOS", "NOS.", "PCS"):
                    quantity_nos = int(quantity) if quantity else 0
                    quantity_kgs = float(weight_kgs) if weight_kgs else 0.0
                else:
                    quantity_kgs = float(quantity) if quantity else 0.0

                # If weight_kgs found and quantity_kgs is the same as quantity, use weight_kgs
                if weight_kgs and not quantity_kgs:
                    quantity_kgs = weight_kgs

                line_items.append({
                    "sl_no": int(sl_val),
                    "description": description,
                    "hsn_sac": hsn_sac,
                    "gst_rate": gst_rate,
                    "quantity_kgs": quantity_kgs,
                    "quantity_nos": quantity_nos,
                    "quantity_boxes": quantity_boxes,
                    "rate_per_unit": float(rate),
                    "amount": float(amount),
                    "remarks": item_remarks,
                })

                r = sub_r
                continue
            r += 1

    # ── Remarks (footer) ──
    for r in range(max_row - 5, max_row + 1):
        v = _cell(ws, r, 1)
        if v.startswith("Remarks:"):
            remarks_text = v.replace("Remarks:", "").strip()
            # Check next row too
            next_v = _cell(ws, r + 1, 1)
            if next_v and "Computer Generated" not in next_v and "PAN" not in next_v:
                remarks_text += " " + next_v
            break

    # ── Total row ──
    total_quantity = ""
    total_amount = 0.0
    amount_in_words = ""
    for r in range(header_row + 2, max_row + 1):
        if _cell(ws, r, 2) == "Total":
            total_quantity = str(ws.cell(row=r, column=10).value or "")
            total_amount = float(ws.cell(row=r, column=13).value or 0)
            break
    for r in range(1, max_row + 1):
        v = _cell(ws, r, 1)
        if v.startswith("INR "):
            amount_in_words = v
            break

    return {
        "challan_no": challan_no,
        "dated": dated,
        "date_time_of_issue": date_time_of_issue,
        "e_way_bill_no": e_way_bill_no,
        "reference_no": reference_no,
        "reference_date": reference_date,
        "dispatch_doc_no": dispatch_doc_no,
        "dispatched_through": dispatched_through,
        "mode_terms_of_payment": mode_terms,
        "other_references": other_references,
        "destination": destination,
        "motor_vehicle_no": motor_vehicle_no,
        "duration_of_process": duration_of_process,
        "nature_of_processing": nature_of_processing,
        "remarks": remarks_text,
        "dispatch_to": {
            "name": dispatch_name,
            "address": dispatch_addr,
            "state": dispatch_state,
            "state_code": dispatch_state_code,
            "gstin": dispatch_gstin,
        },
        "party": {
            "name": party_name,
            "address": party_addr,
            "state": party_state,
            "state_code": party_state_code,
            "gstin": party_gstin,
        },
        "line_items": line_items,
        "totals": {
            "total_quantity": total_quantity,
            "total_amount": total_amount,
            "amount_in_words": amount_in_words,
        },
    }


@router.post("/extract-excel")
async def extract_excel(file: UploadFile = File(...)):
    """
    Parse a Material Out Excel (.xlsx) file and return structured JSON.
    No AI/API key needed — pure openpyxl parsing.
    """
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Uploaded file must be an Excel (.xlsx) file.")

    try:
        file_bytes = await file.read()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read uploaded file: {exc}")

    try:
        extracted = _parse_material_out_excel(file_bytes)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to parse Excel file: {exc}")

    return JSONResponse(content=extracted)


@router.post("/extract-pdf")
async def extract_pdf(file: UploadFile = File(...)):
    """
    Accept a Material Out challan PDF upload, send it to Claude for
    field extraction, and return structured JSON.
    """
    if not file.content_type or "pdf" not in file.content_type.lower():
        raise HTTPException(status_code=400, detail="Uploaded file must be a PDF.")

    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="ANTHROPIC_API_KEY is not configured.")

    try:
        pdf_bytes = await file.read()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to read uploaded file: {exc}")

    pdf_base64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

    extraction_prompt = """You are a precise document-extraction assistant.
I am uploading a Material Out challan PDF. Extract ALL fields and return ONLY valid JSON (no markdown, no code blocks, no explanation).

Return the JSON in exactly this structure:

{
    "challan_no": "",
    "dated": "",
    "date_time_of_issue": "",
    "e_way_bill_no": "",
    "reference_no": "",
    "reference_date": "",
    "dispatch_doc_no": "",
    "dispatched_through": "",
    "mode_terms_of_payment": "",
    "other_references": "",
    "destination": "",
    "motor_vehicle_no": "",
    "duration_of_process": "",
    "nature_of_processing": "",
    "remarks": "",
    "dispatch_to": {
        "name": "",
        "address": "",
        "state": "",
        "state_code": "",
        "gstin": ""
    },
    "party": {
        "name": "",
        "address": "",
        "state": "",
        "state_code": "",
        "gstin": ""
    },
    "line_items": [
        {
            "sl_no": 0,
            "description": "",
            "hsn_sac": "",
            "gst_rate": "",
            "quantity_kgs": 0.0,
            "quantity_nos": 0,
            "quantity_boxes": 0,
            "rate_per_unit": 0.0,
            "amount": 0.0,
            "remarks": ""
        }
    ],
    "totals": {
        "total_quantity": "",
        "total_amount": 0.0,
        "amount_in_words": ""
    }
}

Rules:
- Extract every field visible in the PDF. If a field is blank or not present, use an empty string for text fields and 0 or 0.0 for numeric fields.
- For line_items, include every row in the table. Parse quantities carefully — look for KGS, NOS, and Boxes/Ctns columns.
- For rate_per_unit and amount, return numeric values (not strings).
- For the "remarks" field in line_items, include any packing details like "16units X 60 Ctns".
- The "dispatch_to" section is the consignee/delivery address block.
- The "party" section is the buyer/party details block.
- Return ONLY the JSON object. No other text."""

    try:
        client = anthropic.Anthropic()
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "document",
                            "source": {
                                "type": "base64",
                                "media_type": "application/pdf",
                                "data": pdf_base64,
                            },
                        },
                        {
                            "type": "text",
                            "text": extraction_prompt,
                        },
                    ],
                }
            ],
        )
    except anthropic.BadRequestError as exc:
        msg = str(exc)
        if "credit balance" in msg.lower():
            raise HTTPException(status_code=402, detail="Anthropic API credit balance is too low. Please add credits at console.anthropic.com.")
        raise HTTPException(status_code=400, detail=f"Bad request to Claude API: {msg}")
    except anthropic.AuthenticationError:
        raise HTTPException(status_code=401, detail="Invalid Anthropic API key. Check ANTHROPIC_API_KEY in .env.")
    except anthropic.APIError as exc:
        raise HTTPException(status_code=502, detail=f"Claude API error: {exc}")
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Unexpected error calling Claude API: {exc}")

    raw_text = message.content[0].text.strip()

    # Strip markdown code fences if Claude added them despite instructions
    if raw_text.startswith("```"):
        raw_text = raw_text.split("\n", 1)[1] if "\n" in raw_text else raw_text[3:]
    if raw_text.endswith("```"):
        raw_text = raw_text[:-3].strip()

    try:
        extracted = json.loads(raw_text)
    except json.JSONDecodeError:
        raise HTTPException(
            status_code=500,
            detail="Claude returned invalid JSON. Raw response: " + raw_text[:500],
        )

    return JSONResponse(
        content=extracted,
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "POST, OPTIONS",
            "Access-Control-Allow-Headers": "*",
        },
    )


# ════════════════════════════════════════════════════════════
#  GET /job-work/reports/dashboard  — Reporting & Analytics
# ════════════════════════════════════════════════════════════

@router.get("/reports/dashboard")
def job_work_dashboard(
    period: str = Query("all"),          # all, monthly, quarterly, yearly, custom
    from_date: str = Query(""),
    to_date: str = Query(""),
    sub_category: str = Query(""),       # process type filter
    item: str = Query(""),               # item description filter
    vendor: str = Query(""),             # vendor/party filter
    db: Session = Depends(get_db),
):
    _ensure_tables(db)

    # Build date filter
    date_conditions = []
    params: dict = {}
    if from_date:
        date_conditions.append("h.job_work_date >= :from_date")
        params["from_date"] = from_date
    if to_date:
        date_conditions.append("h.job_work_date <= :to_date")
        params["to_date"] = to_date
    if sub_category:
        date_conditions.append("LOWER(h.sub_category) = LOWER(:sub_cat)")
        params["sub_cat"] = sub_category
    if vendor:
        date_conditions.append("LOWER(h.to_party) LIKE LOWER(:vendor)")
        params["vendor"] = f"%{vendor}%"

    where_h = "WHERE 1=1" + ("".join(f" AND {c}" for c in date_conditions))

    item_join = ""
    if item:
        item_join = " JOIN jb_materialout_lines l_filter ON l_filter.header_id = h.id AND LOWER(l_filter.item_description) LIKE LOWER(:item_filter)"
        params["item_filter"] = f"%{item}%"

    # ── 1. Summary KPIs ──
    summary = db.execute(text(f"""
        SELECT
            COUNT(DISTINCT h.id) as total_jwo,
            COALESCE(SUM(CAST(l.net_weight AS NUMERIC)), 0) as total_dispatched_kgs,
            COUNT(DISTINCT h.to_party) as unique_vendors
        FROM jb_materialout_header h
        {item_join}
        LEFT JOIN jb_materialout_lines l ON l.header_id = h.id
        {where_h}
    """), params).fetchone()

    total_jwo = summary[0] or 0
    total_dispatched = float(summary[1] or 0)
    unique_vendors = summary[2] or 0

    # ── 2. Status Counts ──
    status_rows = db.execute(text(f"""
        SELECT h.status, COUNT(*) FROM jb_materialout_header h {item_join} {where_h} GROUP BY h.status ORDER BY COUNT(*) DESC
    """), params).fetchall()
    status_counts = {r[0]: r[1] for r in status_rows}

    # ── 3. Inward totals ──
    inward_summary = db.execute(text(f"""
        SELECT
            COUNT(DISTINCT ir.id) as total_irs,
            COALESCE(SUM(il.finished_goods_kgs), 0) as total_fg_kgs,
            COALESCE(SUM(il.waste_kgs), 0) as total_waste_kgs,
            COALESCE(SUM(il.rejection_kgs), 0) as total_rejection_kgs
        FROM jb_work_inward_receipt ir
        JOIN jb_work_inward_lines il ON il.inward_receipt_id = ir.id
        JOIN jb_materialout_header h ON h.id = ir.header_id
        {item_join}
        {where_h}
    """), params).fetchone()

    total_irs = inward_summary[0] or 0
    total_fg = float(inward_summary[1] or 0)
    total_waste = float(inward_summary[2] or 0)
    total_rejection = float(inward_summary[3] or 0)
    total_accounted = total_fg + total_waste + total_rejection
    unaccounted = total_dispatched - total_accounted
    overall_loss_pct = (unaccounted / total_dispatched * 100) if total_dispatched > 0 else 0

    # ── 4. By Process Type (Sub Category) ──
    process_rows = db.execute(text(f"""
        SELECT
            COALESCE(h.sub_category, 'Unknown') as process,
            COUNT(DISTINCT h.id) as jwo_count,
            COALESCE(SUM(CAST(l.net_weight AS NUMERIC)), 0) as dispatched_kgs,
            COALESCE((
                SELECT SUM(il.finished_goods_kgs) FROM jb_work_inward_lines il
                JOIN jb_work_inward_receipt ir ON ir.id = il.inward_receipt_id
                WHERE ir.header_id = h.id
            ), 0) as fg_kgs
        FROM jb_materialout_header h
        {item_join}
        LEFT JOIN jb_materialout_lines l ON l.header_id = h.id
        {where_h}
        GROUP BY h.sub_category, h.id
    """), params).fetchall()

    # Aggregate by process
    process_agg: dict = {}
    for r in process_rows:
        p = r[0] or "Unknown"
        if p not in process_agg:
            process_agg[p] = {"jwo_count": 0, "dispatched_kgs": 0, "fg_kgs": 0}
        process_agg[p]["jwo_count"] += 1
        process_agg[p]["dispatched_kgs"] += float(r[2] or 0)
        process_agg[p]["fg_kgs"] += float(r[3] or 0)

    by_process = []
    for p, v in sorted(process_agg.items(), key=lambda x: x[1]["dispatched_kgs"], reverse=True):
        by_process.append({
            "process": p,
            "jwo_count": v["jwo_count"],
            "dispatched_kgs": round(v["dispatched_kgs"], 2),
            "fg_kgs": round(v["fg_kgs"], 2),
        })

    # ── 5. By Vendor ──
    vendor_rows = db.execute(text(f"""
        SELECT
            h.to_party,
            COUNT(DISTINCT h.id) as jwo_count,
            COALESCE(SUM(CAST(l.net_weight AS NUMERIC)), 0) as dispatched_kgs
        FROM jb_materialout_header h
        {item_join}
        LEFT JOIN jb_materialout_lines l ON l.header_id = h.id
        {where_h}
        GROUP BY h.to_party
        ORDER BY dispatched_kgs DESC
    """), params).fetchall()

    by_vendor = [
        {"vendor": r[0] or "Unknown", "jwo_count": r[1], "dispatched_kgs": round(float(r[2] or 0), 2)}
        for r in vendor_rows
    ]

    # ── 6. By Item ──
    item_rows = db.execute(text(f"""
        SELECT
            l.item_description,
            COUNT(DISTINCT h.id) as jwo_count,
            COALESCE(SUM(CAST(l.net_weight AS NUMERIC)), 0) as dispatched_kgs,
            SUM(l.quantity_boxes) as total_boxes
        FROM jb_materialout_header h
        JOIN jb_materialout_lines l ON l.header_id = h.id
        {where_h}
        {"AND LOWER(l.item_description) LIKE LOWER(:item_filter)" if item else ""}
        GROUP BY l.item_description
        ORDER BY dispatched_kgs DESC
    """), params).fetchall()

    by_item = [
        {"item": r[0] or "Unknown", "jwo_count": r[1], "dispatched_kgs": round(float(r[2] or 0), 2), "total_boxes": int(r[3] or 0)}
        for r in item_rows
    ]

    # ── 7. Monthly Trend ──
    monthly_rows = db.execute(text(f"""
        SELECT
            SUBSTRING(h.job_work_date, 4, 7) as month_year,
            COUNT(DISTINCT h.id) as jwo_count,
            COALESCE(SUM(CAST(l.net_weight AS NUMERIC)), 0) as dispatched_kgs
        FROM jb_materialout_header h
        {item_join}
        LEFT JOIN jb_materialout_lines l ON l.header_id = h.id
        {where_h}
        AND h.job_work_date IS NOT NULL AND LENGTH(h.job_work_date) >= 10
        GROUP BY SUBSTRING(h.job_work_date, 4, 7)
        ORDER BY month_year DESC
        LIMIT 12
    """), params).fetchall()

    monthly_trend = [
        {"month": r[0] or "", "jwo_count": r[1], "dispatched_kgs": round(float(r[2] or 0), 2)}
        for r in monthly_rows
    ]

    # ── 8. Vendor × Item Matrix (top combos) ──
    vendor_item_rows = db.execute(text(f"""
        SELECT
            h.to_party, l.item_description,
            COUNT(DISTINCT h.id) as jwo_count,
            COALESCE(SUM(CAST(l.net_weight AS NUMERIC)), 0) as dispatched_kgs
        FROM jb_materialout_header h
        JOIN jb_materialout_lines l ON l.header_id = h.id
        {where_h}
        {"AND LOWER(l.item_description) LIKE LOWER(:item_filter)" if item else ""}
        GROUP BY h.to_party, l.item_description
        ORDER BY dispatched_kgs DESC
        LIMIT 20
    """), params).fetchall()

    vendor_item_matrix = [
        {"vendor": r[0] or "", "item": r[1] or "", "jwo_count": r[2], "dispatched_kgs": round(float(r[3] or 0), 2)}
        for r in vendor_item_rows
    ]

    # ── 9. Filter options (for dropdowns) ──
    sub_cats = db.execute(text("SELECT DISTINCT sub_category FROM jb_materialout_header WHERE sub_category IS NOT NULL AND sub_category != '' ORDER BY sub_category")).fetchall()
    vendors_list = db.execute(text("SELECT DISTINCT to_party FROM jb_materialout_header WHERE to_party IS NOT NULL AND to_party != '' ORDER BY to_party")).fetchall()
    items_list = db.execute(text("SELECT DISTINCT item_description FROM jb_materialout_lines WHERE item_description IS NOT NULL AND item_description != '' ORDER BY item_description")).fetchall()

    return {
        "summary": {
            "total_jwo": total_jwo,
            "total_dispatched_kgs": round(total_dispatched, 2),
            "unique_vendors": unique_vendors,
            "total_irs": total_irs,
            "total_fg_kgs": round(total_fg, 2),
            "total_waste_kgs": round(total_waste, 2),
            "total_rejection_kgs": round(total_rejection, 2),
            "unaccounted_kgs": round(unaccounted, 2),
            "overall_loss_pct": round(overall_loss_pct, 2),
        },
        "status_counts": status_counts,
        "by_process": by_process,
        "by_vendor": by_vendor,
        "by_item": by_item,
        "monthly_trend": monthly_trend,
        "vendor_item_matrix": vendor_item_matrix,
        "filter_options": {
            "sub_categories": [r[0] for r in sub_cats],
            "vendors": [r[0] for r in vendors_list],
            "items": [r[0] for r in items_list],
        },
    }
