from typing import Optional, List
from io import BytesIO

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.orm import Session

from shared.database import get_db
from shared.logger import get_logger

logger = get_logger("cold_storage_server")

router = APIRouter(prefix="/cold-storage", tags=["cold-storage"])

COMPANY_TABLE_MAP = {
    "cfpl": "cfpl_cold_stocks",
    "cdpl": "cdpl_cold_stocks",
}


def _get_cold_table(company: Optional[str]) -> str:
    """Resolve the cold stocks table name from company code."""
    if not company:
        raise HTTPException(status_code=400, detail="company query parameter is required")
    table = COMPANY_TABLE_MAP.get(company.strip().lower())
    if not table:
        raise HTTPException(status_code=400, detail=f"Unknown company: {company}. Use 'cfpl' or 'cdpl'.")
    return table


@router.get("/storage-locations")
def get_storage_locations(
    company: str = Query(..., description="Company code: cfpl or cdpl"),
    db: Session = Depends(get_db),
):
    """Return distinct storage_location values from both cfpl and cdpl cold stocks tables."""
    locations = set()
    for table in ["cfpl_cold_stocks", "cdpl_cold_stocks"]:
        tbl_exists = db.execute(text("SELECT to_regclass(:t)"), {"t": f"public.{table}"}).scalar()
        if not tbl_exists:
            continue
        rows = db.execute(
            text(f"""
                SELECT DISTINCT storage_location
                FROM {table}
                WHERE storage_location IS NOT NULL AND storage_location != ''
            """)
        ).fetchall()
        for r in rows:
            locations.add(r.storage_location)
    return sorted(locations)


@router.get("/stocks/download-summary")
def download_cold_stocks_summary(
    db: Session = Depends(get_db),
):
    """Download consolidated cold storage summary from both cfpl and cdpl tables as Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on the server")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    columns = [
        ("Inward Date", 14), ("Unit", 8), ("Inward No", 22),
        ("Item Description", 35), ("Item Mark", 15), ("Vakkal", 12),
        ("Lot No", 12), ("No of Cartons", 14), ("Weight (Kg)", 14),
        ("Total Inv (Kg)", 15), ("Group Name", 18), ("Storage Location", 20),
        ("Exporter", 20), ("Rate", 12), ("Value", 14),
    ]

    for prefix, label in [("cfpl", "CFPL"), ("cdpl", "CDPL")]:
        table = f"{prefix}_cold_stocks"
        table_exists = db.execute(
            text("SELECT to_regclass(:t)"), {"t": f"public.{table}"}
        ).scalar()

        ws = wb.create_sheet(title=label)

        # Write headers
        for col_idx, (col_name, col_width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col_width

        if not table_exists:
            ws.cell(row=2, column=1, value="Table not found")
            continue

        rows = db.execute(
            text(f"""
                SELECT inward_dt, unit, inward_no, item_description, item_mark,
                       vakkal, lot_no,
                       SUM(no_of_cartons) AS no_of_cartons,
                       MIN(weight_kg) AS weight_kg,
                       SUM(COALESCE(no_of_cartons, 0) * COALESCE(weight_kg, 0)) AS total_inventory_kgs,
                       group_name, storage_location, exporter,
                       MIN(last_purchase_rate) AS last_purchase_rate,
                       SUM(value) AS value
                FROM {table}
                GROUP BY inward_dt, unit, inward_no, item_description, item_mark,
                         vakkal, lot_no, group_name, storage_location, exporter
                ORDER BY inward_dt ASC, item_description ASC
            """)
        ).fetchall()

        for row_idx, r in enumerate(rows, 2):
            data = [
                str(r.inward_dt) if r.inward_dt else "",
                r.unit or "",
                r.inward_no or "",
                r.item_description or "",
                r.item_mark or "",
                r.vakkal or "",
                str(r.lot_no) if r.lot_no else "",
                float(r.no_of_cartons) if r.no_of_cartons else 0,
                float(r.weight_kg) if r.weight_kg else 0,
                float(r.total_inventory_kgs) if r.total_inventory_kgs else 0,
                r.group_name or "",
                r.storage_location or "",
                r.exporter or "",
                float(r.last_purchase_rate) if r.last_purchase_rate else 0,
                float(r.value) if r.value else 0,
            ]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(columns))}1"
        ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=cold_storage_summary.xlsx"},
    )


@router.get("/stocks/download-category-summary")
def download_cold_stocks_category_summary(
    db: Session = Depends(get_db),
):
    """Download category-wise summary split into Dates and Other than Dates, grouped by storage_location, group_name, item_mark."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed on the server")

    from datetime import date as date_type

    # Fetch consolidated data from both tables
    all_rows = []
    for prefix in ("cfpl", "cdpl"):
        table = f"{prefix}_cold_stocks"
        exists = db.execute(text("SELECT to_regclass(:t)"), {"t": f"public.{table}"}).scalar()
        if not exists:
            continue
        rows = db.execute(text(f"""
            SELECT storage_location,
                   group_name,
                   COALESCE(item_mark, '') AS item_mark,
                   SUM(COALESCE(no_of_cartons, 0) * COALESCE(weight_kg, 0)) AS total_kgs,
                   CASE WHEN SUM(COALESCE(no_of_cartons, 0) * COALESCE(weight_kg, 0)) > 0
                        THEN SUM(value) / SUM(COALESCE(no_of_cartons, 0) * COALESCE(weight_kg, 0))
                        ELSE 0 END AS avg_rate,
                   SUM(value) AS total_value
            FROM {table}
            GROUP BY storage_location, group_name, item_mark
            ORDER BY storage_location, group_name, item_mark
        """)).fetchall()
        all_rows.extend(rows)

    # Split into dates vs non-dates based on group_name
    date_keywords = ['date', 'dates', 'ajwa', 'mabroom', 'safawi', 'khalas', 'khidri', 'khanizee', 'barhi', 'fard', 'sayer', 'zahidi', 'mazafati', 'piyarom', 'reziz', 'sufri', 'rabla', 'lulu', 'deri']
    def is_date_group(gname):
        if not gname:
            return False
        gl = gname.lower()
        return any(kw in gl for kw in date_keywords)

    non_dates = []
    dates = []
    for r in all_rows:
        entry = {
            "storage_location": r.storage_location or "",
            "group_name": r.group_name or "",
            "item_mark": r.item_mark or "",
            "total_kgs": float(r.total_kgs or 0),
            "avg_rate": round(float(r.avg_rate or 0), 2),
            "total_value": round(float(r.total_value or 0), 2),
        }
        if is_date_group(r.group_name):
            dates.append(entry)
        else:
            non_dates.append(entry)

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Styles
    title_font = Font(bold=True, size=11, color="000080")
    header_font = Font(bold=True, size=10, color="FFFFFF")
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    num_fmt = '#,##0'
    rate_fmt = '#,##0.00'

    today_str = date_type.today().strftime("%d-%b-%y")

    def write_section(ws, start_col, category_label, data_rows):
        """Write one section (Dates or Other than Dates) starting at start_col."""
        c = start_col
        col_names = ["Storage Location", "Group Name*", "Sub Group Name*", "Total in Kgs", "Avg Rate/Kg", "Approximate Value"]
        col_widths = [18, 20, 22, 14, 14, 18]

        # Row 1: Title headers
        ws.cell(row=1, column=c, value="All Cold Warehouses").font = title_font
        ws.cell(row=1, column=c+1, value="Category --->").font = title_font
        ws.cell(row=1, column=c+2, value=category_label).font = Font(bold=True, size=11, color="FF0000")

        grand_total_kgs = sum(r["total_kgs"] for r in data_rows)
        grand_total_value = sum(r["total_value"] for r in data_rows)
        grand_avg_rate = round(grand_total_value / grand_total_kgs, 2) if grand_total_kgs > 0 else 0

        ws.cell(row=1, column=c+3, value="In Kgs").font = title_font
        ws.cell(row=1, column=c+4, value="Value/KG").font = title_font

        # Row 2: Date + CDPL+CFPL + Summary + totals
        ws.cell(row=2, column=c, value=today_str).font = Font(bold=True, size=10)
        ws.cell(row=2, column=c+1, value="CDPL+CFPL").font = Font(bold=True, size=10)
        ws.cell(row=2, column=c+2, value="Summary").font = Font(bold=True, size=10)
        cell_kgs = ws.cell(row=2, column=c+3, value=grand_total_kgs)
        cell_kgs.font = Font(bold=True, size=10)
        cell_kgs.number_format = num_fmt
        cell_rate = ws.cell(row=2, column=c+4, value=grand_avg_rate)
        cell_rate.font = Font(bold=True, size=10)
        cell_rate.number_format = rate_fmt
        cell_val = ws.cell(row=2, column=c+5, value=grand_total_value)
        cell_val.font = Font(bold=True, size=10)
        cell_val.number_format = num_fmt

        # Row 3: Column headers
        for i, (name, width) in enumerate(zip(col_names, col_widths)):
            cell = ws.cell(row=3, column=c+i, value=name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(c+i)].width = width

        # Data rows grouped by storage_location
        # Sort: by storage_location, group_name, item_mark
        sorted_data = sorted(data_rows, key=lambda x: (x["storage_location"], x["group_name"], x["item_mark"]))

        row = 4
        current_location = None
        location_kgs = 0
        location_value = 0

        for entry in sorted_data:
            # New storage location — write total for previous one
            if entry["storage_location"] != current_location:
                if current_location is not None:
                    # Write location total row
                    ws.cell(row=row, column=c, value=f"{current_location} Total").font = total_font
                    cell = ws.cell(row=row, column=c+3, value=location_kgs)
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.number_format = num_fmt
                    cell.border = thin_border
                    loc_avg = round(location_value / location_kgs, 2) if location_kgs > 0 else 0
                    cell2 = ws.cell(row=row, column=c+4, value=loc_avg)
                    cell2.font = total_font
                    cell2.fill = total_fill
                    cell2.number_format = rate_fmt
                    cell2.border = thin_border
                    row += 1

                current_location = entry["storage_location"]
                location_kgs = 0
                location_value = 0

            # Write data row — only show storage_location and group_name when they change
            ws.cell(row=row, column=c, value=entry["storage_location"]).border = thin_border
            ws.cell(row=row, column=c+1, value=entry["group_name"]).border = thin_border
            ws.cell(row=row, column=c+2, value=entry["item_mark"]).border = thin_border

            cell_k = ws.cell(row=row, column=c+3, value=entry["total_kgs"])
            cell_k.number_format = num_fmt
            cell_k.border = thin_border

            cell_r = ws.cell(row=row, column=c+4, value=entry["avg_rate"])
            cell_r.number_format = rate_fmt
            cell_r.border = thin_border

            cell_v = ws.cell(row=row, column=c+5, value=entry["total_value"])
            cell_v.number_format = num_fmt
            cell_v.border = thin_border

            location_kgs += entry["total_kgs"]
            location_value += entry["total_value"]
            row += 1

        # Write last location total
        if current_location is not None:
            ws.cell(row=row, column=c, value=f"{current_location} Total").font = total_font
            cell = ws.cell(row=row, column=c+3, value=location_kgs)
            cell.font = total_font
            cell.fill = total_fill
            cell.number_format = num_fmt
            cell.border = thin_border
            loc_avg = round(location_value / location_kgs, 2) if location_kgs > 0 else 0
            cell2 = ws.cell(row=row, column=c+4, value=loc_avg)
            cell2.font = total_font
            cell2.fill = total_fill
            cell2.number_format = rate_fmt
            cell2.border = thin_border

        return row

    # Write "Other than Dates" section starting at column A (1)
    write_section(ws, 1, "Other than Dates", non_dates)

    # Write "Dates" section starting at column H (8) — gap of 1 column
    write_section(ws, 8, "Dates", dates)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"cold_storage_summary_{date_type.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/stocks/search")
def search_cold_storage_stocks(
    company: str = Query(..., description="Company code: cfpl or cdpl"),
    lot_no: Optional[str] = Query(None),
    item_description: Optional[str] = Query(None),
    group_name: Optional[str] = Query(None),
    inward_dt: Optional[str] = Query(None),
    unit: Optional[str] = Query(None),
    storage_location: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
):
    where_clauses = []
    params: dict = {}

    if lot_no:
        where_clauses.append("CAST(lot_no AS TEXT) ILIKE :lot_no")
        params["lot_no"] = f"%{lot_no}%"

    if item_description:
        where_clauses.append("item_description ILIKE :item_description")
        params["item_description"] = f"%{item_description}%"

    if group_name:
        where_clauses.append("group_name ILIKE :group_name")
        params["group_name"] = f"%{group_name}%"

    if inward_dt:
        where_clauses.append("CAST(inward_dt AS TEXT) ILIKE :inward_dt")
        params["inward_dt"] = f"%{inward_dt}%"

    if unit:
        where_clauses.append("unit ILIKE :unit")
        params["unit"] = f"%{unit}%"

    if storage_location:
        where_clauses.append("storage_location = :storage_location")
        params["storage_location"] = storage_location

    if q:
        where_clauses.append(
            "(item_description ILIKE :q OR group_name ILIKE :q OR CAST(lot_no AS TEXT) ILIKE :q)"
        )
        params["q"] = f"%{q}%"

    where_sql = " AND ".join(where_clauses) if where_clauses else "TRUE"
    params["limit"] = limit
    table = _get_cold_table(company)

    rows = db.execute(
        text(f"""
            SELECT MIN(id) AS id,
                   MIN(inward_dt) AS inward_dt,
                   MIN(unit) AS unit,
                   COALESCE(inward_no, '') AS inward_no,
                   item_description,
                   COALESCE(item_mark, '') AS item_mark,
                   MAX(vakkal) AS vakkal,
                   COALESCE(lot_no, '') AS lot_no,
                   SUM(no_of_cartons) AS no_of_cartons,
                   MIN(weight_kg) AS weight_kg,
                   SUM(COALESCE(no_of_cartons, 0) * COALESCE(weight_kg, 0)) AS total_inventory_kgs,
                   MAX(group_name) AS group_name,
                   COALESCE(storage_location, '') AS storage_location,
                   MAX(exporter) AS exporter,
                   MIN(last_purchase_rate) AS last_purchase_rate,
                   SUM(value) AS value,
                   MIN(box_id) AS box_id,
                   MIN(transaction_no) AS transaction_no
            FROM {table}
            WHERE {where_sql}
            GROUP BY item_description, COALESCE(lot_no, ''), COALESCE(inward_no, ''), COALESCE(item_mark, ''), COALESCE(storage_location, '')
            ORDER BY MIN(inward_dt) ASC, MIN(id) ASC
            LIMIT :limit
        """),
        params,
    ).fetchall()

    results = [
        {
            "id": r.id,
            "inward_dt": str(r.inward_dt) if r.inward_dt else None,
            "unit": r.unit,
            "inward_no": r.inward_no,
            "item_description": r.item_description,
            "item_mark": r.item_mark,
            "vakkal": r.vakkal,
            "lot_no": r.lot_no,
            "net_qty_on_cartons": float(r.no_of_cartons) if r.no_of_cartons else None,
            "weight_kg": float(r.weight_kg) if r.weight_kg else None,
            "total_inventory_kgs": float(r.total_inventory_kgs) if r.total_inventory_kgs else None,
            "group_name": r.group_name,
            "storage_location": r.storage_location,
            "stock": None,
            "exporter": r.exporter,
            "last_purchase_rate": float(r.last_purchase_rate) if r.last_purchase_rate else None,
            "value": float(r.value) if r.value else None,
            "box_id": r.box_id,
            "transaction_no": r.transaction_no,
        }
        for r in rows
    ]

    return {"results": results, "total": len(results)}


@router.get("/stocks/pick-boxes")
def pick_boxes(
    company: str = Query(..., description="Company code: cfpl or cdpl"),
    item_description: str = Query(...),
    lot_no: str = Query(...),
    inward_no: str = Query(...),
    qty: int = Query(..., ge=1, description="Number of boxes to pick"),
    db: Session = Depends(get_db),
):
    """Return individual box rows in FIFO order (by id ASC) for a given item+lot+inward_no."""
    table = _get_cold_table(company)

    rows = db.execute(
        text(f"""
            SELECT id, box_id, transaction_no, weight_kg, item_mark,
                   inward_dt, unit, inward_no, item_description,
                   vakkal, lot_no, no_of_cartons, total_inventory_kgs,
                   group_name, storage_location, exporter, last_purchase_rate, value
            FROM {table}
            WHERE item_description = :item_description
              AND CAST(lot_no AS TEXT) = :lot_no
              AND inward_no = :inward_no
            ORDER BY id ASC
            LIMIT :qty
        """),
        {
            "item_description": item_description,
            "lot_no": lot_no,
            "inward_no": inward_no,
            "qty": qty,
        },
    ).fetchall()

    return {
        "boxes": [
            {
                "id": r.id,
                "box_id": r.box_id,
                "transaction_no": r.transaction_no,
                "weight_kg": float(r.weight_kg) if r.weight_kg else 0,
                "item_mark": r.item_mark or "",
                "inward_dt": str(r.inward_dt) if r.inward_dt else "",
                "unit": r.unit or "",
                "inward_no": r.inward_no or "",
                "item_description": r.item_description or "",
                "vakkal": r.vakkal or "",
                "lot_no": str(r.lot_no) if r.lot_no else "",
                "no_of_cartons": int(r.no_of_cartons) if r.no_of_cartons else 0,
                "total_inventory_kgs": float(r.total_inventory_kgs) if r.total_inventory_kgs else 0,
                "group_name": r.group_name or "",
                "storage_location": r.storage_location or "",
                "exporter": r.exporter or "",
                "last_purchase_rate": float(r.last_purchase_rate) if r.last_purchase_rate else 0,
                "value": float(r.value) if r.value else 0,
            }
            for r in rows
        ]
    }


# ── Inner Cold Transfer: Update lot numbers in cold_storage_stocks ──

class InnerTransferLine(BaseModel):
    stock_record_id: Optional[int] = None
    item_category: Optional[str] = None
    item_description: Optional[str] = None
    net_weight: Optional[float] = None
    quantity: int
    old_lot_number: str
    new_lot_number: str
    new_storage_location: Optional[str] = None


class InnerTransferHeader(BaseModel):
    challan_no: Optional[str] = None
    transfer_name: Optional[str] = None
    from_warehouse: Optional[str] = None
    remark: Optional[str] = None
    reason_code: Optional[str] = None
    transfer_type: Optional[str] = None


class InnerTransferPayload(BaseModel):
    header: InnerTransferHeader
    lines: List[InnerTransferLine]
    company: Optional[str] = None


def _ensure_inner_cold_transfer_table(db: Session):
    """Create inner_cold_transfer table if it does not exist."""
    db.execute(text("""
        CREATE TABLE IF NOT EXISTS inner_cold_transfer (
            id SERIAL PRIMARY KEY,
            challan_no VARCHAR(50),
            transfer_date VARCHAR(20),
            from_warehouse VARCHAR(100),
            reason_code VARCHAR(100),
            remark TEXT,
            stock_record_id INTEGER,
            item_category VARCHAR(255),
            item_description VARCHAR(255),
            net_weight_kg NUMERIC(12, 3),
            quantity INTEGER,
            old_lot_number VARCHAR(100),
            new_lot_number VARCHAR(100),
            new_storage_location VARCHAR(255),
            transfer_type VARCHAR(50) DEFAULT 'INNER_COLD',
            status VARCHAR(20) DEFAULT 'COMPLETED',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """))
    # Add column if table already existed without it
    db.execute(text("""
        DO $$
        BEGIN
            IF NOT EXISTS (
                SELECT 1 FROM information_schema.columns
                WHERE table_name = 'inner_cold_transfer' AND column_name = 'new_storage_location'
            ) THEN
                ALTER TABLE inner_cold_transfer ADD COLUMN new_storage_location VARCHAR(255);
            END IF;
        END $$;
    """))
    db.commit()


@router.post("/inner-transfer")
def inner_cold_transfer(payload: InnerTransferPayload, db: Session = Depends(get_db)):
    """
    For each line, transfer `quantity` boxes from old lot to new lot in company-specific cold stocks table.
    - If transferring ALL boxes: just update the lot_no on the existing record.
    - If transferring PARTIAL boxes: reduce the original record's cartons/weight,
      and INSERT a new record with the new lot_no and the transferred cartons/weight.
    - Save every transfer line into inner_cold_transfer table for record-keeping.
    """
    cold_table = _get_cold_table(payload.company)
    logger.info(f"Inner cold transfer request: challan={payload.header.challan_no}, company={payload.company}, table={cold_table}, lines={len(payload.lines)}")

    _ensure_inner_cold_transfer_table(db)

    updated_count = 0
    errors = []

    for idx, line in enumerate(payload.lines):
        try:
            logger.info(f"Processing line {idx + 1}: stock_id={line.stock_record_id}, qty={line.quantity}, old_lot={line.old_lot_number}, new_lot={line.new_lot_number}")

            if not line.stock_record_id:
                errors.append(f"Line {idx + 1}: stock_record_id is required")
                continue

            # Fetch the reference record to get item details
            ref_row = db.execute(
                text(f"""
                    SELECT id, inward_dt, unit, inward_no, item_description, item_mark,
                           vakkal, lot_no, no_of_cartons, weight_kg,
                           total_inventory_kgs, group_name, storage_location,
                           exporter, last_purchase_rate, value,
                           box_id, transaction_no
                    FROM {cold_table}
                    WHERE id = :record_id
                """),
                {"record_id": line.stock_record_id},
            ).fetchone()

            if not ref_row:
                errors.append(f"Line {idx + 1}: Record ID {line.stock_record_id} not found")
                continue

            # Fetch ALL rows matching this item/lot group (since stock search groups by item+lot)
            all_rows = db.execute(
                text(f"""
                    SELECT id, no_of_cartons, weight_kg, value
                    FROM {cold_table}
                    WHERE item_description = :desc
                      AND COALESCE(lot_no, '') = COALESCE(:lot_no, '')
                      AND COALESCE(inward_no, '') = COALESCE(:inward_no, '')
                    ORDER BY id ASC
                """),
                {
                    "desc": ref_row.item_description,
                    "lot_no": ref_row.lot_no,
                    "inward_no": ref_row.inward_no,
                },
            ).fetchall()

            current_cartons = sum(float(r.no_of_cartons or 0) for r in all_rows)
            transfer_qty = line.quantity

            logger.info(
                "Inner transfer line %d: found %d matching rows, total_cartons=%.1f, transfer_qty=%d, row_cartons=%s",
                idx + 1, len(all_rows), current_cartons, transfer_qty,
                [float(r.no_of_cartons or 0) for r in all_rows[:5]],
            )

            if transfer_qty <= 0:
                errors.append(f"Line {idx + 1}: Quantity must be greater than 0")
                continue

            if transfer_qty > current_cartons:
                errors.append(
                    f"Line {idx + 1}: Transfer qty ({transfer_qty}) exceeds available ({current_cartons})"
                )
                continue

            # Use reference row for weight/value per carton
            ref_cartons = float(ref_row.no_of_cartons) if ref_row.no_of_cartons else 1
            weight_per_carton = (float(ref_row.weight_kg) / ref_cartons) if (ref_row.weight_kg and ref_cartons > 0) else 0
            total_inv_per_carton = weight_per_carton
            value_per_carton = (float(ref_row.value) / ref_cartons) if (ref_row.value and ref_cartons > 0) else 0

            transferred_weight = round(weight_per_carton * transfer_qty, 3)
            row = ref_row

            new_location = line.new_storage_location if line.new_storage_location else None

            if transfer_qty == current_cartons:
                # Transferring ALL boxes — update lot_no (and optionally location) on all matching rows
                row_ids = [r.id for r in all_rows]
                for rid in row_ids:
                    if new_location:
                        db.execute(
                            text(f"UPDATE {cold_table} SET lot_no = :new_lot_no, storage_location = :new_location WHERE id = :rid"),
                            {"new_lot_no": line.new_lot_number, "new_location": new_location, "rid": rid},
                        )
                    else:
                        db.execute(
                            text(f"UPDATE {cold_table} SET lot_no = :new_lot_no WHERE id = :rid"),
                            {"new_lot_no": line.new_lot_number, "rid": rid},
                        )
            else:
                # PARTIAL transfer across multiple rows — update first N rows to new lot
                # Each row typically has no_of_cartons=1, so we update transfer_qty rows
                rows_to_transfer = []
                remaining_qty = transfer_qty
                for r in all_rows:
                    if remaining_qty <= 0:
                        break
                    r_cartons = float(r.no_of_cartons or 0)
                    if r_cartons <= remaining_qty:
                        # Transfer this entire row
                        rows_to_transfer.append((r.id, r_cartons))
                        remaining_qty -= r_cartons
                    else:
                        # Partial row — need to split
                        rows_to_transfer.append((r.id, remaining_qty))
                        remaining_qty = 0

                for rid, qty in rows_to_transfer:
                    r_row = next((r for r in all_rows if r.id == rid), None)
                    if not r_row:
                        continue
                    r_cartons = float(r_row.no_of_cartons or 0)

                    if qty >= r_cartons:
                        # Transfer entire row — just update lot_no
                        if new_location:
                            db.execute(
                                text(f"UPDATE {cold_table} SET lot_no = :new_lot, storage_location = :new_loc WHERE id = :rid"),
                                {"new_lot": line.new_lot_number, "new_loc": new_location, "rid": rid},
                            )
                        else:
                            db.execute(
                                text(f"UPDATE {cold_table} SET lot_no = :new_lot WHERE id = :rid"),
                                {"new_lot": line.new_lot_number, "rid": rid},
                            )
                    else:
                        # Split this row: reduce original, insert new
                        r_weight = float(r_row.weight_kg or 0)
                        r_value = float(r_row.value or 0)
                        wt_per = r_weight / r_cartons if r_cartons > 0 else 0
                        val_per = r_value / r_cartons if r_cartons > 0 else 0

                        rem = r_cartons - qty
                        db.execute(
                            text(f"""
                                UPDATE {cold_table}
                                SET no_of_cartons = :rem, weight_kg = :rem_wt,
                                    total_inventory_kgs = :rem_wt, value = :rem_val
                                WHERE id = :rid
                            """),
                            {"rem": rem, "rem_wt": round(wt_per * rem, 3), "rem_val": round(val_per * rem, 2), "rid": rid},
                        )
                        db.execute(
                            text(f"""
                                INSERT INTO {cold_table}
                                    (inward_dt, unit, inward_no, item_description, item_mark,
                                     vakkal, lot_no, no_of_cartons, weight_kg,
                                     total_inventory_kgs, group_name, storage_location,
                                     exporter, last_purchase_rate, value,
                                     box_id, transaction_no)
                                VALUES
                                    (:inward_dt, :unit, :inward_no, :item_description, :item_mark,
                                     :vakkal, :new_lot, :qty, :wt,
                                     :wt, :group_name, :storage_location,
                                     :exporter, :rate, :val,
                                     :box_id, :txn)
                            """),
                            {
                                "inward_dt": row.inward_dt, "unit": row.unit, "inward_no": row.inward_no,
                                "item_description": row.item_description, "item_mark": row.item_mark,
                                "vakkal": row.vakkal, "new_lot": line.new_lot_number,
                                "qty": qty, "wt": round(wt_per * qty, 3),
                                "group_name": row.group_name,
                                "storage_location": new_location or row.storage_location,
                                "exporter": row.exporter,
                                "rate": float(row.last_purchase_rate) if row.last_purchase_rate else None,
                                "val": round(val_per * qty, 2),
                                "box_id": None, "txn": None,
                            },
                        )

            # Save transfer record in inner_cold_transfer table
            db.execute(
                text("""
                    INSERT INTO inner_cold_transfer
                        (challan_no, transfer_date, from_warehouse, reason_code, remark,
                         stock_record_id, item_category, item_description, net_weight_kg,
                         quantity, old_lot_number, new_lot_number, new_storage_location, transfer_type)
                    VALUES
                        (:challan_no, :transfer_date, :from_warehouse, :reason_code, :remark,
                         :stock_record_id, :item_category, :item_description, :net_weight_kg,
                         :quantity, :old_lot_number, :new_lot_number, :new_storage_location, :transfer_type)
                """),
                {
                    "challan_no": payload.header.challan_no,
                    "transfer_date": payload.header.transfer_name,
                    "from_warehouse": payload.header.from_warehouse,
                    "reason_code": payload.header.reason_code,
                    "remark": payload.header.remark,
                    "stock_record_id": line.stock_record_id,
                    "item_category": line.item_category,
                    "item_description": line.item_description,
                    "net_weight_kg": transferred_weight,
                    "quantity": transfer_qty,
                    "old_lot_number": line.old_lot_number,
                    "new_lot_number": line.new_lot_number,
                    "new_storage_location": new_location,
                    "transfer_type": payload.header.transfer_type or "INNER_COLD",
                },
            )

            updated_count += 1

        except Exception as e:
            logger.error(f"Inner transfer line {idx + 1} failed: {e}")
            errors.append(f"Line {idx + 1}: {str(e)}")

    try:
        db.commit()
        logger.info(f"Inner cold transfer committed: updated={updated_count}, errors={len(errors)}")
    except Exception as commit_err:
        db.rollback()
        logger.error(f"Inner cold transfer commit failed: {commit_err}")
        return {
            "status": "error",
            "updated_records": 0,
            "errors": [f"Database commit failed: {str(commit_err)}"],
            "challan_no": payload.header.challan_no,
        }

    if errors:
        logger.warning(f"Inner cold transfer completed with errors: {errors}")

    return {
        "status": "success" if not errors else "partial",
        "updated_records": updated_count,
        "errors": errors,
        "challan_no": payload.header.challan_no,
    }


@router.get("/inner-transfer/list")
def list_inner_cold_transfers(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """List inner cold transfer records grouped by challan_no."""
    # Check if table exists
    table_check = db.execute(
        text("SELECT to_regclass('public.inner_cold_transfer')")
    ).scalar()
    if not table_check:
        return {"records": [], "total": 0, "page": page, "per_page": per_page, "total_pages": 0}

    # Get total count of unique challans
    total = db.execute(
        text("SELECT COUNT(DISTINCT challan_no) FROM inner_cold_transfer")
    ).scalar() or 0

    total_pages = max(1, -(-total // per_page))  # ceil division
    offset = (page - 1) * per_page

    # Get distinct challans with aggregated info
    rows = db.execute(
        text("""
            SELECT
                challan_no,
                MIN(transfer_date) AS transfer_date,
                MIN(from_warehouse) AS from_warehouse,
                MIN(reason_code) AS reason_code,
                MIN(remark) AS remark,
                MIN(status) AS status,
                COUNT(*) AS line_count,
                SUM(quantity) AS total_boxes,
                MIN(created_at) AS created_at
            FROM inner_cold_transfer
            GROUP BY challan_no
            ORDER BY MIN(created_at) DESC
            LIMIT :limit OFFSET :offset
        """),
        {"limit": per_page, "offset": offset},
    ).fetchall()

    records = []
    for r in rows:
        # Get the line details for this challan
        lines = db.execute(
            text("""
                SELECT item_description, item_category, quantity,
                       old_lot_number, new_lot_number, net_weight_kg,
                       new_storage_location
                FROM inner_cold_transfer
                WHERE challan_no = :challan_no
                ORDER BY id
            """),
            {"challan_no": r.challan_no},
        ).fetchall()

        records.append({
            "challan_no": r.challan_no,
            "transfer_date": r.transfer_date,
            "from_warehouse": r.from_warehouse,
            "reason_code": r.reason_code,
            "remark": r.remark,
            "status": r.status or "COMPLETED",
            "line_count": r.line_count,
            "total_boxes": r.total_boxes,
            "created_at": str(r.created_at) if r.created_at else None,
            "lines": [
                {
                    "item_description": l.item_description,
                    "item_category": l.item_category,
                    "quantity": l.quantity,
                    "old_lot_number": l.old_lot_number,
                    "new_lot_number": l.new_lot_number,
                    "net_weight_kg": float(l.net_weight_kg) if l.net_weight_kg else 0,
                    "new_storage_location": l.new_storage_location,
                }
                for l in lines
            ],
        })

    return {
        "records": records,
        "total": total,
        "page": page,
        "per_page": per_page,
        "total_pages": total_pages,
    }


INNER_COLD_DELETE_ALLOWED_EMAILS = {"hrithik@candorfoods.in", "yash@candorfoods.in"}


@router.get("/inner-transfer/{challan_no}")
def get_inner_cold_transfer(challan_no: str, db: Session = Depends(get_db)):
    """Get a single inner cold transfer by challan_no."""
    table_check = db.execute(
        text("SELECT to_regclass('public.inner_cold_transfer')")
    ).scalar()
    if not table_check:
        raise HTTPException(status_code=404, detail="Inner cold transfer table not found")

    header = db.execute(
        text("""
            SELECT challan_no, MIN(transfer_date) AS transfer_date,
                   MIN(from_warehouse) AS from_warehouse,
                   MIN(reason_code) AS reason_code,
                   MIN(remark) AS remark,
                   MIN(status) AS status,
                   MIN(created_at) AS created_at
            FROM inner_cold_transfer
            WHERE challan_no = :challan_no
            GROUP BY challan_no
        """),
        {"challan_no": challan_no},
    ).fetchone()

    if not header:
        raise HTTPException(status_code=404, detail="Inner cold transfer not found")

    lines = db.execute(
        text("""
            SELECT id, stock_record_id, item_category, item_description,
                   net_weight_kg, quantity, old_lot_number, new_lot_number,
                   new_storage_location
            FROM inner_cold_transfer
            WHERE challan_no = :challan_no
            ORDER BY id
        """),
        {"challan_no": challan_no},
    ).fetchall()

    return {
        "challan_no": header.challan_no,
        "transfer_date": header.transfer_date,
        "from_warehouse": header.from_warehouse,
        "reason_code": header.reason_code,
        "remark": header.remark,
        "status": header.status or "COMPLETED",
        "created_at": str(header.created_at) if header.created_at else None,
        "lines": [
            {
                "id": l.id,
                "stock_record_id": l.stock_record_id,
                "item_category": l.item_category,
                "item_description": l.item_description,
                "net_weight_kg": float(l.net_weight_kg) if l.net_weight_kg else 0,
                "quantity": l.quantity,
                "old_lot_number": l.old_lot_number,
                "new_lot_number": l.new_lot_number,
                "new_storage_location": l.new_storage_location,
            }
            for l in lines
        ],
    }


@router.delete("/inner-transfer/{challan_no}")
def delete_inner_cold_transfer(
    challan_no: str,
    user_email: str = Query(..., description="Email of the user performing the delete"),
    db: Session = Depends(get_db),
):
    """Delete an inner cold transfer by challan_no. Restricted to authorized users."""
    if user_email not in INNER_COLD_DELETE_ALLOWED_EMAILS:
        raise HTTPException(
            status_code=403,
            detail="You are not authorized to delete inner cold transfers.",
        )

    table_check = db.execute(
        text("SELECT to_regclass('public.inner_cold_transfer')")
    ).scalar()
    if not table_check:
        raise HTTPException(status_code=404, detail="Inner cold transfer table not found")

    result = db.execute(
        text("DELETE FROM inner_cold_transfer WHERE challan_no = :challan_no"),
        {"challan_no": challan_no},
    )
    db.commit()

    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Inner cold transfer not found")

    logger.info(f"Inner cold transfer {challan_no} deleted by {user_email}")
    return {"status": "success", "message": f"Inner cold transfer {challan_no} deleted successfully."}
