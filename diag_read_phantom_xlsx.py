import io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import openpyxl

wb = openpyxl.load_workbook("../Corrections/Phantom_Transactions_List.xlsx", read_only=True, data_only=True)

def dump(name, maxrows=200):
    ws = wb[name]
    print(f"\n{'='*90}\nSHEET: {name}  ({ws.max_row} rows)\n{'='*90}")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > maxrows:
            print("  ... (truncated)"); break
        cells = [("" if c is None else str(c)) for c in row]
        # skip fully-empty rows
        if any(c.strip() for c in cells):
            print("  " + " | ".join(cells))

for s in ["Summary", "Relabel (FIX 1)", "Phantom In-Transit (FIX 2)",
          "Present, no system record", "System stock, not physical"]:
    dump(s)
