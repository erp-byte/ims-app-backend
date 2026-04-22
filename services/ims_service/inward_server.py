from io import BytesIO
from datetime import date
from typing import Optional

from fastapi import APIRouter, Depends, Query, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from shared.logger import get_logger

logger = get_logger("inward_server")

from shared.database import get_db
from services.ims_service.inward_models import (
    Company,
    InwardPayloadFlexible,
    InwardUpdatePayload,
    InwardListResponse,
    POExtractResponse,
    MultiPOExtractResponse,
    POUploadResponse,
    PageExtractResponse,
    SKULookupRequest,
    SKULookupResponse,
    SKUDropdownResponse,
    SKUGlobalSearchResponse,
    SKUIdResponse,
    ApprovalRequest,
    BoxUpsertRequest,
    BoxUpsertResponse,
    BoxEditLogRequest,
    BulkStickerPayload,
    BulkStickerResponse,
)
from services.ims_service.inward_tools import (
    list_inward_records,
    list_distinct_warehouses,
    export_inward_records,
    create_inward,
    create_inward_bulk_sticker,
    get_inward,
    update_inward,
    delete_inward,
    approve_inward,
    extract_po_from_pdf,
    upload_po_pdf,
    extract_single_page,
    lookup_sku,
    sku_dropdown,
    sku_global_search,
    sku_id_lookup,
    upsert_box,
    log_box_edits,
    get_vendors,
)

router = APIRouter(prefix="/inward", tags=["inward"])


# ── Per-page extraction endpoints (avoids App Runner 120s timeout) ──────


@router.post("/extract-po/upload", response_model=POUploadResponse)
def upload_po_endpoint(file: UploadFile = File(...)):
    """Upload a PO PDF, split into pages, return job_id + total_pages."""
    try:
        contents = file.file.read()
        logger.info(f"extract-po/upload: received '{file.filename}', size={len(contents)} bytes")
        result = upload_po_pdf(contents)
        return POUploadResponse(**result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"extract-po/upload failed: {type(e).__name__}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Upload failed: {type(e).__name__}: {str(e)}")


@router.post("/extract-po/{job_id}/page/{page_num}", response_model=PageExtractResponse)
def extract_page_endpoint(job_id: str, page_num: int):
    """Extract POs from a single page of a previously uploaded PDF."""
    try:
        result = extract_single_page(job_id, page_num)
        return PageExtractResponse(**result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"extract-po/{job_id}/page/{page_num} failed: {type(e).__name__}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Page extraction failed: {type(e).__name__}: {str(e)}")


# ── Legacy single-request extraction (kept for backward compat) ─────────


@router.post("/extract-po", response_model=MultiPOExtractResponse)
def extract_po_endpoint(file: UploadFile = File(...)):
    """Upload a PO PDF and extract one or more POs via Claude Sonnet 4.5."""
    try:
        contents = file.file.read()
        logger.info(f"extract-po: received file '{file.filename}', size={len(contents)} bytes")
        result = extract_po_from_pdf(contents)
        return MultiPOExtractResponse(**result)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"extract-po failed: {type(e).__name__}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"PO extraction failed: {type(e).__name__}: {str(e)}")


@router.post("/sku-lookup/{company}", response_model=SKULookupResponse)
def sku_lookup_endpoint(
    company: Company,
    body: SKULookupRequest,
    db: Session = Depends(get_db),
):
    """Lookup SKU details by item description."""
    result = lookup_sku(body.item_description, company, db)
    if result is None:
        return SKULookupResponse(item_description=body.item_description)
    return SKULookupResponse(**result)


@router.get("", response_model=InwardListResponse)
def list_inward_records_query(
    company: Company = Query(..., description="Company code"),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=1000),
    skip: int = Query(0, ge=0),
    limit: int = Query(1000, ge=1, le=1000),
    status: Optional[str] = Query(None, description="Filter by status (pending, approved)"),
    grn_status: Optional[str] = Query(None, description="Filter by GRN status (completed, pending)"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse name (exact match, case-insensitive)"),
    search: Optional[str] = Query(None, description="Search across all transaction fields"),
    from_date: Optional[str] = Query(None, description="Filter from date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="Filter to date (YYYY-MM-DD)"),
    sort_by: Optional[str] = Query("entry_date", description="Sort field"),
    sort_order: Optional[str] = Query("desc", description="Sort order (asc, desc)"),
    db: Session = Depends(get_db),
):
    """List inward + bulk entry records with company as query parameter (backward compat)."""
    if skip > 0 or limit != 1000:
        page = (skip // limit) + 1 if limit > 0 else 1
        per_page = min(limit, 100)

    return list_inward_records(
        company=company,
        page=page,
        per_page=per_page,
        search=search,
        from_date=from_date,
        to_date=to_date,
        sort_by=sort_by,
        sort_order=sort_order,
        db=db,
        status=status,
        grn_status=grn_status,
        warehouse=warehouse,
    )


@router.get("/warehouses")
def list_warehouses_endpoint(
    company: Company = Query(..., description="Company code"),
    db: Session = Depends(get_db),
):
    """Distinct warehouse values across inward and bulk entry transactions (dropdown source)."""
    return {"warehouses": list_distinct_warehouses(company, db)}


@router.get("/vendors")
def vendors_endpoint(
    search: Optional[str] = Query(None, description="Filter vendor names (ILIKE)"),
    db: Session = Depends(get_db),
):
    """Return all vendors (optionally filtered by name)."""
    return get_vendors(search, db)


@router.get("/sku-dropdown", response_model=SKUDropdownResponse)
def sku_dropdown_endpoint(
    company: Company = Query(...),
    material_type: Optional[str] = Query(None),
    item_category: Optional[str] = Query(None),
    sub_category: Optional[str] = Query(None),
    item_description: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    limit: int = Query(200, ge=1, le=10000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    """Cascading SKU dropdown for manual article entry."""
    return sku_dropdown(
        company, material_type, item_category, sub_category,
        item_description, search, limit, offset, db,
    )


@router.get("/sku-search", response_model=SKUGlobalSearchResponse)
def sku_global_search_endpoint(
    company: Company = Query(...),
    search: Optional[str] = Query(None),
    limit: int = Query(200, ge=1, le=10000),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    """Global item description search — bypasses hierarchy."""
    return sku_global_search(company, search, limit, offset, db)


@router.get("/sku-id", response_model=SKUIdResponse)
def sku_id_endpoint(
    company: Company = Query(...),
    item_description: str = Query(...),
    item_category: Optional[str] = Query(None),
    sub_category: Optional[str] = Query(None),
    material_type: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """Get SKU ID for a specific item description."""
    return sku_id_lookup(company, item_description, item_category, sub_category, material_type, db)


@router.get("/export")
def export_inward_endpoint(
    company: Company = Query(..., description="Company code"),
    status: Optional[str] = Query(None),
    grn_status: Optional[str] = Query(None),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse name (exact match, case-insensitive)"),
    search: Optional[str] = Query(None),
    from_date: Optional[str] = Query(None),
    to_date: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("entry_date"),
    sort_order: Optional[str] = Query("desc"),
    db: Session = Depends(get_db),
):
    """Export all filtered inward + bulk entry records as an Excel (.xlsx) file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlalchemy import text as sa_text

    rows = export_inward_records(
        company=company,
        search=search,
        from_date=from_date,
        to_date=to_date,
        sort_by=sort_by,
        sort_order=sort_order,
        db=db,
        status=status,
        grn_status=grn_status,
        warehouse=warehouse,
    )

    # Build edit-log lookup: { (box_id, field_name) } for highlighting
    tx_nos = list({r["Transaction No"] for r in rows if r.get("Transaction No")})
    edited_cells: set[tuple[str, str]] = set()
    if tx_nos:
        edit_rows = db.execute(
            sa_text("""
                SELECT box_id, field_name FROM box_edit_logs
                WHERE transaction_no = ANY(:txnos)
            """),
            {"txnos": tx_nos},
        ).fetchall()
        for er in edit_rows:
            edited_cells.add((er.box_id, er.field_name))

    # Map field_name in edit_logs to Excel header
    field_to_header = {
        "net_weight": "Box Net Weight",
        "gross_weight": "Box Gross Weight",
        "lot_number": "Box Lot Number",
        "count": "Box Count",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Inward Records"

    # Headers from all 3 tables
    headers = [
        # Origin + Transaction
        "Source", "Warehouse",
        "Transaction No", "Entry Date", "Status", "Vehicle Number", "Transporter",
        "LR Number", "Vendor / Supplier", "Customer / Party", "Source Location",
        "Destination", "Challan Number", "Invoice Number", "PO Number",
        "GRN Number", "GRN Quantity", "System GRN Date", "Purchased By",
        "Service Invoice", "DN Number", "Approval Authority",
        "Total Amount", "Tax Amount", "Discount Amount", "PO Quantity",
        "Remark", "Currency",
        # Article
        "SKU ID", "Item Description", "Item Category", "Sub Category",
        "Material Type", "Quality Grade", "UOM", "Art PO Qty", "Units",
        "Quantity Units", "Art Net Weight", "Total Weight", "PO Weight",
        "Art Lot Number", "Mfg Date", "Expiry Date", "Unit Rate",
        "Art Total Amount", "Carton Weight",
        # Box
        "Box ID", "Box Article", "Box Number", "Box Net Weight",
        "Box Gross Weight", "Box Lot Number", "Box Count",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="29417A", end_color="29417A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    edited_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Header-index lookup for box columns that can be edited
    box_header_cols = {h: i + 1 for i, h in enumerate(headers) if h in field_to_header.values()}

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(rows, 2):
        box_id = row_data.get("Box ID", "")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.border = thin_border

            # Highlight edited box fields with light red
            if box_id and header in box_header_cols:
                for field_name, mapped_header in field_to_header.items():
                    if mapped_header == header and (box_id, field_name) in edited_cells:
                        cell.fill = edited_fill
                        break

    # Auto-fit: use header length as minimum width
    for col_idx, header in enumerate(headers, 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(len(header) + 4, 14)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"inward_{company}_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{company}", response_model=InwardListResponse)
def list_inward_records_path(
    company: Company,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=1000),
    status: Optional[str] = Query(None, description="Filter by status (pending, approved)"),
    grn_status: Optional[str] = Query(None, description="Filter by GRN status (completed, pending)"),
    warehouse: Optional[str] = Query(None, description="Filter by warehouse name (exact match, case-insensitive)"),
    search: Optional[str] = Query(None, description="Search across all transaction fields"),
    from_date: Optional[str] = Query(None, description="Filter from date (YYYY-MM-DD)"),
    to_date: Optional[str] = Query(None, description="Filter to date (YYYY-MM-DD)"),
    sort_by: Optional[str] = Query("entry_date", description="Sort field (entry_date, transaction_no, invoice_number)"),
    sort_order: Optional[str] = Query("desc", description="Sort order (asc, desc)"),
    db: Session = Depends(get_db),
):
    """List inward + bulk entry records with comprehensive search and date filtering."""
    return list_inward_records(
        company=company,
        page=page,
        per_page=per_page,
        search=search,
        from_date=from_date,
        to_date=to_date,
        sort_by=sort_by,
        sort_order=sort_order,
        db=db,
        status=status,
        grn_status=grn_status,
        warehouse=warehouse,
    )


@router.post("", status_code=201)
def create_inward_endpoint(payload: InwardPayloadFlexible, db: Session = Depends(get_db)):
    return create_inward(payload, db)


@router.post("/bulk-sticker", status_code=201, response_model=BulkStickerResponse)
def create_inward_bulk_sticker_endpoint(
    payload: BulkStickerPayload,
    db: Session = Depends(get_db),
):
    """Create inward entry with immediate box_id generation for bulk sticker printing."""
    return create_inward_bulk_sticker(payload, db)


@router.put("/{company}/{transaction_no}/box", response_model=BoxUpsertResponse)
def upsert_box_endpoint(
    company: Company,
    transaction_no: str,
    payload: BoxUpsertRequest,
    db: Session = Depends(get_db),
):
    """Upsert a single box row. Returns box_id for QR label printing."""
    return upsert_box(company, transaction_no, payload, db)


@router.post("/box-edit-log")
def log_box_edit_endpoint(
    payload: BoxEditLogRequest,
    db: Session = Depends(get_db),
):
    """Log audit entries for edits to a previously-printed box."""
    return log_box_edits(payload, db)


@router.put("/{company}/{transaction_no}/approve")
def approve_inward_endpoint(
    company: Company,
    transaction_no: str,
    payload: ApprovalRequest,
    db: Session = Depends(get_db),
):
    """Approve or reject a pending inward entry."""
    return approve_inward(company, transaction_no, payload, db)


@router.get("/{company}/{transaction_no}")
def get_inward_endpoint(company: Company, transaction_no: str, db: Session = Depends(get_db)):
    return get_inward(company, transaction_no, db)


@router.put("/{company}/{transaction_no}")
def update_inward_endpoint(
    company: Company,
    transaction_no: str,
    payload: InwardUpdatePayload,
    db: Session = Depends(get_db),
):
    return update_inward(company, transaction_no, payload, db)


@router.delete("/{company}/{transaction_no}")
def delete_inward_endpoint(company: Company, transaction_no: str, db: Session = Depends(get_db)):
    return delete_inward(company, transaction_no, db)
