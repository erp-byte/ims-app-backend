"""IMS vs System box-count reconciliation (Cold_and_IMS_Comparison_14Jun2026.xlsx, 2nd sheet).

For every 'Qty Mismatch' lot, make the cold table's box count equal IMS Cartons:
  • live > IMS  -> DELETE (live-IMS) boxes FROM LAST (highest numeric suffix).
                   Each deleted box is first snapshotted into cold_stock_disposition
                   (disposition_type='manual_correction', snapshot_data=full row) so it
                   is fully restorable.
  • live < IMS  -> ADD (IMS-live) boxes FROM LAST: copy the max-suffix box of the lot,
                   assign box_id = <prefix>-<maxsuffix+i>.
  • live == IMS -> skip (already correct / already fixed).

Per-lot guard: the lot's final count MUST equal IMS, else that lot is rolled back.
Commits per lot. Dry-run by default; --apply to write.

Usage:
  python ims_recon_14jun.py            # dry-run, no writes
  python ims_recon_14jun.py --apply     # execute
"""
import sys
import json
import openpyxl
import psycopg2
import psycopg2.extras

DB = dict(host="wms-postgres-db.cpis084golp7.ap-south-1.rds.amazonaws.com",
          port=5432, dbname="warehouse_db", user="wmsadmin", password="Candorfoods")
XLSX = "../Corrections/Cold_and_IMS_Comparison_14Jun2026.xlsx"
SHEET = "IMS_vs_System_Comparison"
REF_NO = "IMS-RECON-14JUN2026"
APPLY = "--apply" in sys.argv


def suffix(bid):
    if bid and "-" in bid:
        t = bid.rsplit("-", 1)[-1]
        if t.isdigit():
            return int(t)
    return None


def prefix_of(bid):
    return bid.rsplit("-", 1)[0] if bid and "-" in bid else bid


def read_targets():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb[SHEET]
    out = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 1).value == "Qty Mismatch":
            ent = str(ws.cell(r, 2).value).strip().upper()
            lot = str(ws.cell(r, 3).value).strip()
            try:
                ims = int(ws.cell(r, 18).value)
            except Exception:
                ims = None
            out.append((ent, lot, ims))
    return out


def main():
    targets = read_targets()
    conn = psycopg2.connect(**DB)
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    tot_del = tot_add = lots_del = lots_add = lots_skip = lots_err = 0
    print(f"{'MODE':5} = {'APPLY' if APPLY else 'DRY-RUN'}   lots={len(targets)}\n")

    for ent, lot, ims in targets:
        tbl = "cdpl_cold_stocks" if ent == "CDPL" else "cfpl_cold_stocks"
        comp = "cdpl" if ent == "CDPL" else "cfpl"
        if ims is None:
            print(f"  [SKIP] {ent} {lot}: no IMS cartons"); lots_err += 1; continue

        cur.execute(f"SELECT * FROM {tbl} WHERE lot_no = %s", (lot,))
        rows = cur.fetchall()
        live = len(rows)
        if live == 0:
            print(f"  [SKIP] {ent} {lot}: not in {tbl}"); lots_err += 1; continue

        delta = ims - live
        if delta == 0:
            lots_skip += 1
            continue

        # order by numeric suffix DESC (non-numeric suffixes treated as lowest)
        rows_sorted = sorted(rows, key=lambda r: (suffix(r["box_id"]) if suffix(r["box_id"]) is not None else -1,
                                                  r["id"]), reverse=True)

        try:
            if delta < 0:
                n = -delta
                victims = rows_sorted[:n]
                ids = [r["id"] for r in victims]
                # 1) snapshot each into cold_stock_disposition
                for r in victims:
                    if APPLY:
                        cur.execute("""
                            INSERT INTO cold_stock_disposition
                                (box_id, transaction_no, lot_no, item_description,
                                 from_company, unit, from_site, source_table,
                                 disposition_type, disposition_ref_no, disposed_by,
                                 snapshot_data, notes)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,'manual_correction',%s,'ims_recon',
                                    CAST(%s AS JSONB), %s)
                        """, (r["box_id"], (r.get("transaction_no") or ""), lot, r.get("item_description"),
                              comp, r.get("unit"), r.get("storage_location"), tbl,
                              REF_NO, json.dumps(r, default=str),
                              f"IMS recon 14Jun: excess over IMS qty ({live}->{ims}), deleted from last"))
                # 2) delete by exact id list
                if APPLY:
                    cur.execute(f"DELETE FROM {tbl} WHERE id = ANY(%s) AND lot_no = %s", (ids, lot))
                    deleted = cur.rowcount
                    cur.execute(f"SELECT COUNT(*) AS c FROM {tbl} WHERE lot_no = %s", (lot,))
                    final = cur.fetchone()["c"]
                    if deleted != n or final != ims:
                        conn.rollback()
                        print(f"  [ERR ] {ent} {lot}: deleted {deleted}/{n}, final {final}!={ims} -> ROLLBACK")
                        lots_err += 1; continue
                    conn.commit()
                lots_del += 1; tot_del += n
                lo, hi = victims[-1]["box_id"], victims[0]["box_id"]
                print(f"  DELETE {ent} {lot}: live={live} ims={ims}  -{n}  (last {n}: {hi} .. {lo})"
                      + ("  [committed]" if APPLY else ""))

            else:  # delta > 0  -> ADD
                n = delta
                numeric = [r for r in rows_sorted if suffix(r["box_id"]) is not None]
                if not numeric:
                    print(f"  [ERR ] {ent} {lot}: no numeric box suffix to extend"); lots_err += 1; continue
                tmpl = numeric[0]
                base = prefix_of(tmpl["box_id"])
                smax = suffix(tmpl["box_id"])
                new_ids = [f"{base}-{smax + i}" for i in range(1, n + 1)]
                # guard: none of the new ids already exist for this txn
                cur.execute(f"SELECT box_id FROM {tbl} WHERE transaction_no=%s AND box_id = ANY(%s)",
                            (tmpl.get("transaction_no"), new_ids))
                clash = [x["box_id"] for x in cur.fetchall()]
                if clash:
                    print(f"  [ERR ] {ent} {lot}: new box_ids already exist {clash[:3]} -> skip"); lots_err += 1; continue
                if APPLY:
                    cols = [c for c in tmpl.keys() if c not in ("id",)]
                    collist = ",".join(cols)
                    ph = ",".join(["%s"] * len(cols))
                    for nb in new_ids:
                        vals = []
                        for c in cols:
                            vals.append(nb if c == "box_id" else tmpl[c])
                        cur.execute(f"INSERT INTO {tbl} ({collist}) VALUES ({ph})", vals)
                    cur.execute(f"SELECT COUNT(*) AS c FROM {tbl} WHERE lot_no = %s", (lot,))
                    final = cur.fetchone()["c"]
                    if final != ims:
                        conn.rollback()
                        print(f"  [ERR ] {ent} {lot}: final {final}!={ims} after add -> ROLLBACK"); lots_err += 1; continue
                    conn.commit()
                lots_add += 1; tot_add += n
                print(f"  ADD    {ent} {lot}: live={live} ims={ims}  +{n}  ({new_ids[0]} .. {new_ids[-1]})"
                      + ("  [committed]" if APPLY else ""))
        except Exception as e:
            conn.rollback()
            print(f"  [ERR ] {ent} {lot}: {e!r} -> ROLLBACK"); lots_err += 1

    print(f"\n=== SUMMARY ({'APPLIED' if APPLY else 'DRY-RUN'}) ===")
    print(f"  delete lots={lots_del}  boxes={tot_del}")
    print(f"  add    lots={lots_add}  boxes={tot_add}")
    print(f"  already-match skipped={lots_skip}   errors/skipped={lots_err}")
    if not APPLY:
        print("\n  DRY-RUN — nothing written. Re-run with --apply.")
    conn.close()


if __name__ == "__main__":
    main()
