# -*- coding: utf-8 -*-
"""Lots with two or more entries in the inter-unit transfer report.

Reads the report built by build_interunit_transfers_report.py (no DB access) and
writes a Summary sheet (one row per lot) + Detail sheet (every entry of those lots).

    python tools/multi_entry_lots.py [report.xlsx] [out.xlsx]
"""
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "data" / "interunit_transfers_report_till_26Jul2026.xlsx"
OUT = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "data" / "transfer_lots_multi_entry.xlsx"

SHEETS = ["Transfer Out", "Cold Transfer Out", "Transfer In", "Cold Transfer In"]
SUM_COLS = ["Lot", "Total Entries", "Out Entries", "Out Challans", "In Entries", "In GRNs",
            "Out Boxes", "In Boxes", "Box Diff", "Items", "From", "To", "First Date", "Last Date"]
DET_COLS = ["Lot", "Sheet", "Date", "Challan", "GRN No", "From", "To", "Item", "Category",
            "Material", "Qty", "Net Weight", "Boxes", "Status", "Received"]


def read_entries():
    wb = load_workbook(SRC, read_only=True, data_only=True)
    out = []
    for name in SHEETS:
        rows = list(wb[name].values)
        hdr = list(rows[0])
        idx = {c: hdr.index(c) for c in hdr if c}
        for r in rows[1:]:
            lot = str(r[idx["Lot"]] or "").strip()
            if not lot:
                continue
            out.append({"sheet": name, "lot": lot,
                        **{c.lower().replace(" ", "_"): r[i] for c, i in idx.items()}})
    wb.close()
    return out


def uniq(vals):
    return sorted({str(v).strip() for v in vals if v not in (None, "")})


def main():
    by_lot = defaultdict(list)
    for e in read_entries():
        by_lot[e["lot"]].append(e)

    summary, detail = [], []
    for lot, es in sorted(by_lot.items()):
        outs = [e for e in es if e["sheet"].endswith("Out")]
        ins = [e for e in es if e["sheet"].endswith("In")]
        # "two or more entries" on the same side: dispatched twice, or received twice.
        # A single out + single in is the normal pair, not a multi-entry lot.
        if len(outs) < 2 and len(ins) < 2:
            continue
        dates = sorted(str(e["date"]) for e in es if e["date"])
        ob = sum(e["boxes"] or 0 for e in outs)
        ib = sum(e["boxes"] or 0 for e in ins)
        summary.append([lot, len(es), len(outs), len(uniq(e["challan"] for e in outs)),
                        len(ins), len(uniq(e.get("grn_no") for e in ins)), ob, ib, ib - ob,
                        ", ".join(uniq(e["item"] for e in es))[:200],
                        ", ".join(uniq(e["from"] for e in es)),
                        ", ".join(uniq(e["to"] for e in es)),
                        dates[0] if dates else "", dates[-1] if dates else ""])
        for e in sorted(es, key=lambda x: (SHEETS.index(x["sheet"]), str(x["date"] or ""))):
            detail.append([lot, e["sheet"], str(e["date"] or ""), e["challan"], e.get("grn_no", ""),
                           e["from"], e["to"], e["item"], e["category"], e["material"],
                           e["qty"], e["net_weight"], e["boxes"], e["status"], e["received"]])

    summary.sort(key=lambda r: (-r[1], r[0]))

    wb = Workbook()
    wb.remove(wb.active)
    for title, cols, rows in [("Summary", SUM_COLS, summary), ("Detail", DET_COLS, detail)]:
        ws = wb.create_sheet(title)
        ws.append(cols)
        for r in rows:
            ws.append(r)
        for c in range(1, len(cols) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
            ws.cell(row=1, column=c).alignment = Alignment(horizontal="center")
            width = max([len(str(cols[c - 1]))] +
                        [len(str(r[c - 1])) for r in rows[:400] if r[c - 1] is not None]) + 2
            ws.column_dimensions[get_column_letter(c)].width = min(max(width, 9), 48)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  lots with 2+ entries on one side: {len(summary)}  (detail rows: {len(detail)})")
    print(f"  2+ OUT entries: {sum(1 for r in summary if r[2] >= 2)}"
          f"   2+ IN entries: {sum(1 for r in summary if r[4] >= 2)}"
          f"   both: {sum(1 for r in summary if r[2] >= 2 and r[4] >= 2)}")
    print("\n  Top 20 by entry count:")
    print(f"  {'Lot':<12}{'Tot':>4}{'Out':>5}{'In':>4}{'OutBx':>7}{'InBx':>6}  Item")
    for r in summary[:20]:
        print(f"  {r[0]:<12}{r[1]:>4}{r[2]:>5}{r[4]:>4}{r[6]:>7}{r[7]:>6}  {r[9][:44]}")


if __name__ == "__main__":
    main()
