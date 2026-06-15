"""READ-ONLY report: trace cold_stocks transaction_no/box_id against boxes_v2 +
bulk_entry_boxes, lot + item-description wise.

For every lot present in cold AND (v2 OR bulk), per company, compare the cold
transaction_no / box_id set with the v2 and bulk sets. Flag lots whose cold txn or
box_id does NOT line up with v2/bulk (cold txn likely wrong -> trace correct one).

Output: Corrections/v2 and bulk vs cold table.xlsx
  Sheet 'Mismatches'      - only the flagged lots, with suggested correct txn.
  Sheet 'All Overlaps'    - every overlapping lot with match status (full audit).
  Sheet 'Mismatch Boxes'  - box-level rows (cold + v2/bulk) for the flagged lots.
Nothing is written to the DB.
"""
import os, io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
from sqlalchemy import create_engine, text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

raw = os.environ["DATABASE_URL"]
DB = raw.replace("postgresql://", "postgresql+psycopg://", 1) if raw.startswith("postgresql://") else raw
engine = create_engine(DB, pool_pre_ping=True)
OUT = os.path.join(os.path.dirname(__file__), "..", "Corrections",
                   ("v2 and bulk vs cold table (after 3585 fix).xlsx" if "--suffix" in sys.argv
                    else "v2 and bulk vs cold table.xlsx"))

def bases(ids):
    return ", ".join(sorted({(b or "").rsplit("-", 1)[0] for b in ids if b})[:8])

def agg(ids):
    return ", ".join(sorted({t for t in ids if t}))

overlaps = []   # summary rows
mismatch_boxes = []  # box-level rows for flagged lots

for comp in ["CFPL", "CDPL"]:
    cl = comp.lower()
    cs, v2, bk = f"{cl}_cold_stocks", f"{cl}_boxes_v2", f"{cl}_bulk_entry_boxes"
    with engine.connect() as c:
        lots = c.execute(text(f"""
          SELECT DISTINCT TRIM(cs.lot_no) lot FROM {cs} cs WHERE cs.box_id IS NOT NULL
            AND (EXISTS(SELECT 1 FROM {v2} v WHERE TRIM(v.lot_number)=TRIM(cs.lot_no))
              OR EXISTS(SELECT 1 FROM {bk} b WHERE TRIM(b.lot_number)=TRIM(cs.lot_no)))
        """)).scalars().all()
        for lot in sorted(lots):
            cold = c.execute(text(f"SELECT box_id, transaction_no, item_description FROM {cs} WHERE TRIM(lot_no)=:l AND box_id IS NOT NULL"), {"l": lot}).all()
            v2r = c.execute(text(f"SELECT box_id, transaction_no, article_description FROM {v2} WHERE TRIM(lot_number)=:l"), {"l": lot}).all()
            bkr = c.execute(text(f"SELECT box_id, transaction_no, article_description FROM {bk} WHERE TRIM(lot_number)=:l"), {"l": lot}).all()

            cold_txn = {r[1] for r in cold}; v2_txn = {r[1] for r in v2r}; bk_txn = {r[1] for r in bkr}
            cold_box = {r[0] for r in cold}; v2_box = {r[0] for r in v2r}; bk_box = {r[0] for r in bkr}
            src_txn = v2_txn | bk_txn; src_box = v2_box | bk_box

            txn_status = "MATCH" if cold_txn <= src_txn else ("PARTIAL" if (cold_txn & src_txn) else "MISMATCH")
            box_status = "MATCH" if cold_box <= src_box else ("PARTIAL" if (cold_box & src_box) else "MISMATCH")
            flagged = (txn_status != "MATCH") or (box_status != "MATCH")
            suggested = agg(src_txn) if flagged else ""

            cold_desc = "; ".join(sorted({(r[2] or "").strip() for r in cold}))
            src_desc = "; ".join(sorted({(r[2] or "").strip() for r in (v2r + bkr)}))

            row = {
                "Company": comp, "Lot No": lot, "Cold Item Desc": cold_desc, "v2/Bulk Article Desc": src_desc,
                "Cold Boxes": len(cold), "v2 Boxes": len(v2r), "Bulk Boxes": len(bkr),
                "Cold Txn(s)": agg(cold_txn), "v2 Txn(s)": agg(v2_txn), "Bulk Txn(s)": agg(bk_txn),
                "Cold box bases": bases(cold_box), "v2 box bases": bases(v2_box), "Bulk box bases": bases(bk_box),
                "Cold boxes in v2": len(cold_box & v2_box), "Cold boxes in bulk": len(cold_box & bk_box),
                "Txn Match": txn_status, "BoxId Match": box_status,
                "Status": "MISMATCH" if flagged else "OK",
                "Suggested correct txn (v2/bulk)": suggested,
            }
            overlaps.append(row)

            if flagged:
                for bid, txn, desc in cold:
                    mismatch_boxes.append({"Company": comp, "Lot No": lot, "Source": "COLD", "box_id": bid,
                                           "transaction_no": txn, "desc": (desc or "").strip(),
                                           "Suggested correct txn": suggested})
                for bid, txn, desc in v2r:
                    mismatch_boxes.append({"Company": comp, "Lot No": lot, "Source": "V2", "box_id": bid,
                                           "transaction_no": txn, "desc": (desc or "").strip(), "Suggested correct txn": ""})
                for bid, txn, desc in bkr:
                    mismatch_boxes.append({"Company": comp, "Lot No": lot, "Source": "BULK", "box_id": bid,
                                           "transaction_no": txn, "desc": (desc or "").strip(), "Suggested correct txn": ""})

# ---------- write workbook ----------
HDR_FILL = PatternFill("solid", fgColor="1F4E78"); HDR_FONT = Font(bold=True, color="FFFFFF")
BAD_FILL = PatternFill("solid", fgColor="FCE4D6")

def write_sheet(ws, rows, cols):
    ws.append(cols)
    for ci in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=ci); cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    # highlight MISMATCH rows if Status col present
    if "Status" in cols:
        sidx = cols.index("Status") + 1
        for ri in range(2, ws.max_row + 1):
            if ws.cell(row=ri, column=sidx).value == "MISMATCH":
                for ci in range(1, len(cols) + 1):
                    ws.cell(row=ri, column=ci).fill = BAD_FILL
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    for ci, name in enumerate(cols, 1):
        w = max(len(str(name)), *(len(str(r.get(name, ""))) for r in rows)) if rows else len(str(name))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(w + 2, 10), 60)

wb = openpyxl.Workbook()
sum_cols = ["Company","Lot No","Cold Item Desc","v2/Bulk Article Desc","Cold Boxes","v2 Boxes","Bulk Boxes",
            "Cold Txn(s)","v2 Txn(s)","Bulk Txn(s)","Cold box bases","v2 box bases","Bulk box bases",
            "Cold boxes in v2","Cold boxes in bulk","Txn Match","BoxId Match","Status","Suggested correct txn (v2/bulk)"]
mismatches = [r for r in overlaps if r["Status"] == "MISMATCH"]
mismatches.sort(key=lambda r: (-r["Cold Boxes"]))
overlaps.sort(key=lambda r: (r["Status"] != "MISMATCH", r["Company"], -r["Cold Boxes"]))

ws1 = wb.active; ws1.title = "Mismatches"
write_sheet(ws1, mismatches, sum_cols)
write_sheet(wb.create_sheet("All Overlaps"), overlaps, sum_cols)
box_cols = ["Company","Lot No","Source","box_id","transaction_no","desc","Suggested correct txn"]
write_sheet(wb.create_sheet("Mismatch Boxes"), mismatch_boxes, box_cols)

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"Overlapping lots: {len(overlaps)}  | Mismatched lots: {len(mismatches)}  | mismatch box rows: {len(mismatch_boxes)}")
for m in mismatches:
    print(f"  {m['Company']} lot {m['Lot No']:>8} | cold {m['Cold Boxes']}box txn[{m['Cold Txn(s)']}] -> suggested [{m['Suggested correct txn (v2/bulk)']}]")
print("Saved:", os.path.abspath(OUT))
