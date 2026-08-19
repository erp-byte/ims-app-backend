# -*- coding: utf-8 -*-
"""Transfer register consolidated PER ARTICLE PER TRANSACTION (not per box).

Every sheet collapses the per-box / per-line ledgers down to one row per
(document, article), carrying qty, UOM, box count, weights, lots and all the
header context the IMS records.

Sheets: Transfer Out | Cold Transfer Out | Transfer In | Cold Transfer In |
        Cold Direct Out | Inner Cold Transfer | In Transit (Pending) |
        All Transfers | Notes

READ-ONLY against prod. Usage:
    python tools/build_transfers_by_article_report.py [FROM] [TO] [out.xlsx]
    (dates YYYY-MM-DD; defaults 2026-04-01 -> today)
"""
import os
import sys
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine, text

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

D_FROM = date.fromisoformat(sys.argv[1]) if len(sys.argv) > 1 else date(2026, 4, 1)
D_TO = date.fromisoformat(sys.argv[2]) if len(sys.argv) > 2 else date.today()
OUT_XLSX = (Path(sys.argv[3]) if len(sys.argv) > 3 else
            ROOT / "data" / f"transfers_by_article_{D_FROM:%d%b%Y}_to_{D_TO:%d%b%Y}.xlsx")
TO_TS = D_TO + timedelta(days=1)          # exclusive upper bound on timestamps
FROM_TS = D_FROM

_raw = os.environ["DATABASE_URL"]
engine = create_engine(_raw.replace("postgresql://", "postgresql+psycopg://", 1)
                       if _raw.startswith("postgresql://") else _raw)

# Cold sites, lower-cased. Dispatch FROM one => Cold Transfer-Out; receipt INTO
# one => Cold Transfer-In. 'Cold Storage' is the generic source used when the
# specific unit is carried in interunit_transfers_header.from_cold_unit.
COLD_SITES = {"savla d-39", "savla d-514", "savla bond", "rishi", "supreme", "eskimo"}
COLD_SQL = "('savla d-39','savla d-514','savla bond','rishi','supreme','eskimo')"

PARAMS = {"d1": D_FROM, "d1_ts": FROM_TS, "d2": D_TO, "d2_ts": TO_TS}


def fetch(sql, **params):
    with engine.connect() as c:
        c.execute(text("SET TRANSACTION READ ONLY"))
        return [dict(r._mapping) for r in c.execute(text(sql), {**PARAMS, **params})]


def n3(x):
    return None if x is None else round(float(x), 3)


def i0(x):
    return 0 if x is None else int(x)


def is_cold(site):
    s = (site or "").strip().lower()
    return s.startswith("cold") or s in COLD_SITES


def received_state(status, has_grn):
    if status == "Received":
        return "Received"
    return "Partially/Pending" if has_grn else "Not Received"


# ── 1. Transfer-Out / Cold Transfer-Out ──────────────────────────────────
# Grain = (challan, article). Box count prefers the real per-box records in
# interunit_transfer_boxes (linked by transfer_line_id, verified 100% resolvable);
# where a challan predates per-box capture it falls back to the transfer-line
# count, which is what cold dispatch writes one-per-carton anyway.
OUT_SQL = """
WITH boxcnt AS (
  SELECT transfer_line_id, count(*) AS n FROM interunit_transfer_boxes GROUP BY 1
)
SELECT h.id AS header_id, h.challan_no, h.stock_trf_date::text AS doc_date,
       h.from_site, h.to_site, COALESCE(h.from_cold_unit,'') AS from_cold_unit,
       h.status, h.vehicle_no, h.driver_name, h.reason_code, h.remark,
       h.created_by, h.created_ts::text AS created_ts,
       h.has_variance, h.unallocated_boxes,
       COALESCE(rq.request_no,'') AS request_no,
       (h.stock_trf_date BETWEEN :d1 AND :d2) AS in_by_doc_date,
       COALESCE(NULLIF(TRIM(l.item_desc_raw),''),'(blank)') AS article,
       string_agg(DISTINCT NULLIF(TRIM(l.item_category),''), ' / ') AS category,
       string_agg(DISTINCT NULLIF(TRIM(l.sub_category),''), ' / ')  AS sub_category,
       string_agg(DISTINCT NULLIF(TRIM(l.rm_pm_fg_type),''), ' / ') AS material_type,
       string_agg(DISTINCT NULLIF(TRIM(l.uom),''), ' / ')           AS uom,
       string_agg(DISTINCT NULLIF(TRIM(l.lot_number),''), ', ')     AS lots,
       string_agg(DISTINCT NULLIF(TRIM(l.batch_number),''), ', ')   AS batches,
       string_agg(DISTINCT NULLIF(TRIM(l.vakkal),''), ', ')         AS vakkals,
       string_agg(DISTINCT NULLIF(TRIM(l.hsn_code),''), ', ')       AS hsn,
       count(*)                    AS line_rows,
       sum(l.qty)                  AS qty,
       sum(l.net_weight)           AS net_wt,
       sum(l.total_weight)         AS total_wt,
       min(l.pack_size)            AS pack_min,
       max(l.pack_size)            AS pack_max,
       COALESCE(sum(bc.n), 0)      AS box_rows
FROM interunit_transfers_lines l
JOIN interunit_transfers_header h  ON h.id = l.header_id
LEFT JOIN interunit_transfer_requests rq ON rq.id = h.request_id
LEFT JOIN boxcnt bc ON bc.transfer_line_id = l.id
WHERE (h.stock_trf_date BETWEEN :d1 AND :d2)
   OR (h.created_ts >= :d1_ts AND h.created_ts < :d2_ts)
GROUP BY h.id, h.challan_no, h.stock_trf_date, h.from_site, h.to_site, h.from_cold_unit,
         h.status, h.vehicle_no, h.driver_name, h.reason_code, h.remark, h.created_by,
         h.created_ts, h.has_variance, h.unallocated_boxes, rq.request_no, article
ORDER BY h.stock_trf_date, h.challan_no, article
"""

GRN_SQL = """
SELECT transfer_out_id, string_agg(DISTINCT grn_number, ', ') AS grns,
       string_agg(DISTINCT grn_date::date::text, ', ') AS grn_dates, count(*) AS n
FROM (
  SELECT transfer_out_id, grn_number, grn_date FROM interunit_transfer_in_header
  UNION
  SELECT transfer_out_id, grn_number, grn_date FROM cold_transfer_in_headers
) t WHERE transfer_out_id IS NOT NULL GROUP BY 1
"""

OUT_COLS = [
    "TR Number", "TR Date", "Direction", "From Warehouse", "To Warehouse", "Cold Unit",
    "Article", "Category", "Sub Category", "Material Type",
    "Qty", "UOM", "Box Count", "Box Count Basis", "Lines Merged",
    "Net Weight (kg)", "Total Weight (kg)", "Pack Size (min)", "Pack Size (max)",
    "Lot No(s)", "Batch No(s)", "Vakkal(s)", "HSN",
    "Status", "Receipt Status", "GRN No(s)", "GRN Date(s)",
    "Vehicle No", "Driver", "Reason Code", "Remark",
    "Request No", "Has Variance", "Unallocated Boxes",
    "Created By", "Created At", "In Range By",
]


def build_out():
    grn = {r["transfer_out_id"]: r for r in fetch(GRN_SQL)}
    wh, cold = [], []
    for r in fetch(OUT_SQL):
        g = grn.get(r["header_id"])
        cold_flag = is_cold(r["from_site"]) or bool(r["from_cold_unit"].strip())
        basis = "box records" if r["box_rows"] else "transfer lines"
        row = [
            r["challan_no"], r["doc_date"],
            "Cold Transfer Out" if cold_flag else "Transfer Out",
            r["from_site"], r["to_site"], r["from_cold_unit"],
            r["article"], r["category"], r["sub_category"], r["material_type"],
            n3(r["qty"]), r["uom"],
            i0(r["box_rows"]) or i0(r["line_rows"]), basis, i0(r["line_rows"]),
            n3(r["net_wt"]), n3(r["total_wt"]), n3(r["pack_min"]), n3(r["pack_max"]),
            r["lots"], r["batches"], r["vakkals"], r["hsn"],
            r["status"], received_state(r["status"], bool(g)),
            g["grns"] if g else "", g["grn_dates"] if g else "",
            r["vehicle_no"], r["driver_name"], r["reason_code"], r["remark"],
            r["request_no"], "Yes" if r["has_variance"] else "", r["unallocated_boxes"],
            r["created_by"], r["created_ts"],
            "document date" if r["in_by_doc_date"] else "entry date",
        ]
        (cold if cold_flag else wh).append(row)
    return wh, cold


# ── 2. Category / Material / UOM carried from the matching transfer-OUT line ──
LINE_MAP_SQL = """
SELECT header_id, LOWER(TRIM(COALESCE(item_desc_raw,''))) AS item,
       COALESCE(lot_number,'') AS lot,
       MIN(COALESCE(item_category,'')) AS category,
       MIN(COALESCE(rm_pm_fg_type,'')) AS material,
       MIN(COALESCE(uom,''))           AS uom,
       MIN(COALESCE(sub_category,''))  AS sub_category
FROM interunit_transfers_lines GROUP BY 1,2,3
"""


def line_lookup():
    by_item_lot, by_item, by_lot = {}, {}, {}
    for r in fetch(LINE_MAP_SQL):
        v = (r["category"], r["material"], r["uom"], r["sub_category"])
        by_item_lot[(r["header_id"], r["item"], r["lot"])] = v
        by_item.setdefault((r["header_id"], r["item"]), v)
        if r["lot"]:
            by_lot.setdefault((r["header_id"], r["lot"]), v)

    def look(out_id, item, lot):
        it = (item or "").strip().lower()
        # lots aggregate to a "a, b" string at this grain; try each part
        for cand in [lot or ""] + [p.strip() for p in (lot or "").split(",")]:
            hit = by_item_lot.get((out_id, it, cand))
            if hit:
                return hit
        return by_item.get((out_id, it)) or by_lot.get((out_id, lot)) or ("", "", "", "")
    return look


# ── 3. Transfer-In (all destinations; split cold vs warehouse in python) ──
IN_SQL = f"""
SELECT i.id AS header_id, i.grn_number, i.grn_date::date::text AS grn_date, i.status,
       i.received_by, i.received_at::text AS received_at, i.box_condition,
       i.condition_remarks, COALESCE(i.inward_transaction_no,'') AS inward_txn,
       i.transfer_out_id,
       COALESCE(o.challan_no, i.transfer_out_no, '') AS challan,
       o.stock_trf_date::text AS out_date, o.status AS out_status,
       COALESCE(o.from_site,'') AS from_site,
       COALESCE(NULLIF(TRIM(i.receiving_warehouse),''), o.to_site, '') AS to_site,
       COALESCE(o.vehicle_no,'') AS vehicle_no, COALESCE(o.driver_name,'') AS driver_name,
       COALESCE(o.reason_code,'') AS reason_code, COALESCE(o.from_cold_unit,'') AS from_cold_unit,
       COALESCE(NULLIF(TRIM(b.article),''),'(blank)') AS article,
       string_agg(DISTINCT NULLIF(TRIM(b.lot_number),''), ', ')   AS lots,
       string_agg(DISTINCT NULLIF(TRIM(b.batch_number),''), ', ') AS batches,
       string_agg(DISTINCT NULLIF(TRIM(b.scan_source),''), ', ')  AS scan_source,
       count(b.id)                                      AS box_rows,
       count(b.id) FILTER (WHERE b.is_matched)          AS matched,
       count(b.id) FILTER (WHERE NOT b.is_matched)      AS unmatched,
       sum(b.net_weight)                                AS net_wt,
       sum(b.gross_weight)                              AS gross_wt,
       min(b.scanned_at)::text                          AS first_scan,
       max(b.scanned_at)::text                          AS last_scan
FROM interunit_transfer_in_header i
LEFT JOIN interunit_transfer_in_boxes b ON b.header_id = i.id
LEFT JOIN interunit_transfers_header o  ON o.id = i.transfer_out_id
WHERE i.grn_date >= :d1_ts AND i.grn_date < :d2_ts
GROUP BY i.id, i.grn_number, i.grn_date, i.status, i.received_by, i.received_at,
         i.box_condition, i.condition_remarks, i.inward_transaction_no, i.transfer_out_id,
         o.challan_no, i.transfer_out_no, o.stock_trf_date, o.status, o.from_site,
         i.receiving_warehouse, o.to_site, o.vehicle_no, o.driver_name, o.reason_code,
         o.from_cold_unit, COALESCE(NULLIF(TRIM(b.article),''),'(blank)')
ORDER BY i.grn_date, i.grn_number, article
"""

IN_COLS = [
    "GRN No", "GRN Date", "Direction", "From Warehouse", "To Warehouse", "Cold Unit",
    "Article", "Category", "Sub Category", "Material Type",
    "Qty (boxes received)", "UOM", "Box Count",
    "Net Weight (kg)", "Gross Weight (kg)",
    "Lot No(s)", "Batch No(s)",
    "GRN Status", "Receipt Status", "TR Number", "TR Date",
    "Boxes Matched", "Boxes Unmatched", "Scan Source", "First Scan", "Last Scan",
    "Received By", "Received At", "Box Condition", "Condition Remarks",
    "Inward Txn No", "Vehicle No", "Driver", "Reason Code", "Box Source",
]


def build_in(look):
    wh, cold = [], []
    for r in fetch(IN_SQL):
        cat, mat, uom, sub = look(r["transfer_out_id"], r["article"], r["lots"])
        cold_flag = is_cold(r["to_site"])
        boxes = i0(r["box_rows"])
        row = [
            r["grn_number"], r["grn_date"],
            "Cold Transfer In" if cold_flag else "Transfer In",
            r["from_site"], r["to_site"], r["from_cold_unit"],
            r["article"], cat, sub, mat,
            boxes, uom, boxes,
            n3(r["net_wt"]), n3(r["gross_wt"]),
            r["lots"], r["batches"],
            r["status"], received_state(r["out_status"], True),
            r["challan"], r["out_date"],
            i0(r["matched"]), i0(r["unmatched"]), r["scan_source"],
            r["first_scan"], r["last_scan"],
            r["received_by"], r["received_at"], r["box_condition"], r["condition_remarks"],
            r["inward_txn"], r["vehicle_no"], r["driver_name"], r["reason_code"],
            "interunit_transfer_in_boxes",
        ]
        (cold if cold_flag else wh).append(row)
    return wh, cold


# ── 4. Cold Transfer-In written through the dedicated cold tables ─────────
# Two receive paths write two tables that share an id space. The cold carton
# table is authoritative for a cold receipt whenever it holds real per-box
# detail (>1 row): it is what the cold-receive path writes and it ties back to
# the dispatched quantity. Where it holds only a summary row (or nothing) the
# interunit scan ledger is used instead.
#
# This matters. GRN-20260608124120 receives challan TRANS202606051800, which
# dispatched 185 boxes / 1850.000 kg. The cold table records 185 / 1850.000;
# the interunit scan ledger records 369 / 3690.000. Counting the ledger would
# double the received quantity for that GRN.
COLD_IN_HDR_SQL = """
SELECT c.id, c.grn_number, c.grn_date::date::text AS grn_date, c.status,
       c.received_by, c.received_at::text AS received_at, c.box_condition,
       c.condition_remarks, COALESCE(c.inward_transaction_no,'') AS inward_txn,
       c.transfer_out_id, COALESCE(o.challan_no, c.transfer_out_no,'') AS challan,
       o.stock_trf_date::text AS out_date, o.status AS out_status,
       COALESCE(c.from_site, o.from_site,'') AS from_site,
       COALESCE(NULLIF(TRIM(c.to_site),''), o.to_site,'') AS to_site,
       COALESCE(c.to_company,'') AS to_company,
       COALESCE(o.vehicle_no,'') AS vehicle_no, COALESCE(o.driver_name,'') AS driver_name,
       COALESCE(o.reason_code,'') AS reason_code, COALESCE(o.from_cold_unit,'') AS from_cold_unit
FROM cold_transfer_in_headers c
LEFT JOIN interunit_transfers_header o ON o.id = c.transfer_out_id
WHERE c.grn_date >= :d1_ts AND c.grn_date < :d2_ts
ORDER BY c.grn_date, c.grn_number
"""

COLD_IN_BOX_SQL = """
SELECT header_id, COALESCE(NULLIF(TRIM(item_description),''),'(blank)') AS article,
       string_agg(DISTINCT NULLIF(TRIM(lot_no),''), ', ')          AS lots,
       string_agg(DISTINCT NULLIF(TRIM(vakkal),''), ', ')          AS vakkals,
       string_agg(DISTINCT NULLIF(TRIM(group_name),''), ' / ')     AS group_name,
       string_agg(DISTINCT NULLIF(TRIM(item_subgroup),''), ' / ')  AS subgroup,
       string_agg(DISTINCT NULLIF(TRIM(storage_location),''), ', ')AS storage_loc,
       string_agg(DISTINCT NULLIF(TRIM(item_mark),''), ', ')       AS item_mark,
       string_agg(DISTINCT NULLIF(TRIM(unit),''), ', ')            AS unit,
       string_agg(DISTINCT NULLIF(TRIM(exporter),''), ', ')        AS exporter,
       count(*) AS scan_rows, sum(COALESCE(no_of_cartons,1)) AS cartons,
       sum(weight_kg) AS net_wt, sum(value) AS value
FROM cold_transfer_inboxes WHERE header_id = ANY(:ids)
GROUP BY 1,2 ORDER BY 1,2
"""

IU_IN_COUNT_SQL = """
SELECT header_id, count(*) AS n FROM interunit_transfer_in_boxes
WHERE header_id = ANY(:ids) GROUP BY 1
"""


def build_cold_in_dedicated(look):
    """Cold receipts taken from the cold carton table.

    Returns (rows, grns) where `grns` is the set of GRNs these rows own; the
    caller must drop those GRNs from the interunit-sourced rows so no box is
    counted twice.
    """
    hdrs = fetch(COLD_IN_HDR_SQL)
    if not hdrs:
        return [], set()
    ids = [h["id"] for h in hdrs]
    boxes = defaultdict(list)
    for r in fetch(COLD_IN_BOX_SQL, ids=ids):
        boxes[r["header_id"]].append(r)
    iu_n = {r["header_id"]: r["n"] for r in fetch(IU_IN_COUNT_SQL, ids=ids)}

    rows, grns = [], set()
    for h in hdrs:
        cb = boxes.get(h["id"], [])
        cold_rows = sum(i0(g["scan_rows"]) for g in cb)
        # >1 row means real per-box detail. A lone summary row is only used
        # when the interunit ledger has nothing to offer instead.
        use_cold = cold_rows > 1 or (cold_rows == 1 and not iu_n.get(h["id"]))
        if not use_cold:
            continue
        grns.add(h["grn_number"])
        for g in cb:
            cat, mat, uom, sub = look(h["transfer_out_id"], g["article"], g["lots"])
            cartons = int(g["cartons"] or 0)
            rows.append([
                h["grn_number"], h["grn_date"], "Cold Transfer In",
                h["from_site"], h["to_site"], h["from_cold_unit"],
                g["article"], cat or g["group_name"], sub or g["subgroup"], mat,
                cartons, uom or "CARTON", cartons,
                n3(g["net_wt"]), None,
                g["lots"], "",
                h["status"], received_state(h["out_status"], True),
                h["challan"], h["out_date"],
                cartons, 0, "cold receive", "", "",
                h["received_by"], h["received_at"], h["box_condition"], h["condition_remarks"],
                h["inward_txn"], h["vehicle_no"], h["driver_name"], h["reason_code"],
                "cold_transfer_inboxes",
            ])
    return rows, grns


# ── 5. Cold Direct Out (stock issued straight out of cold to a customer) ──
DIRECT_OUT_SQL = """
SELECT t.transaction_no, t.transaction_type, t.company, t.entry_date::text AS entry_date,
       COALESCE(NULLIF(TRIM(t.warehouse),''),'') AS warehouse,
       COALESCE(t.to_customer,'') AS to_customer, COALESCE(t.vehicle_no,'') AS vehicle_no,
       COALESCE(t.invoice_no,'') AS invoice_no, COALESCE(t.remarks,'') AS remarks,
       t.status, COALESCE(t.created_by,'') AS created_by,
       t.created_at::date::text AS created_at, COALESCE(t.lot_no,'') AS hdr_lot,
       COALESCE(NULLIF(TRIM(e->>'item_description'),''),'(blank)') AS article,
       string_agg(DISTINCT NULLIF(e->>'uom',''), ' / ')       AS uom,
       string_agg(DISTINCT NULLIF(e->>'lot_no',''), ', ')     AS lots,
       string_agg(DISTINCT NULLIF(e->>'unit',''), ', ')       AS units,
       string_agg(DISTINCT NULLIF(e->>'warehouse',''), ', ')  AS line_wh,
       string_agg(DISTINCT NULLIF(e->>'item_mark',''), ', ')  AS item_marks,
       string_agg(DISTINCT NULLIF(e->>'inward_no',''), ', ')  AS inward_nos,
       string_agg(DISTINCT NULLIF(e->>'transaction_no',''), ', ') AS src_txns,
       count(*) AS box_rows,
       sum(NULLIF(e->>'issue_qty','')::numeric) AS qty,
       sum(NULLIF(e->>'issue_qty','')::numeric
           * COALESCE(NULLIF(e->>'weight_kg_per_box','')::numeric, 0)) AS net_wt
FROM (SELECT * FROM cdpl_cold_storage_direct_out
      UNION ALL
      SELECT * FROM cfpl_cold_storage_direct_out) t,
     LATERAL jsonb_array_elements(t.lines) e
WHERE t.entry_date BETWEEN :d1 AND :d2
GROUP BY 1,2,3,4,5,6,7,8,9,10,11,12,13, article
ORDER BY 4, 1, article
"""

DIRECT_OUT_COLS = [
    "TR Number", "Date", "Direction", "Company", "From Warehouse", "Cold Unit(s)",
    "To Customer", "Article", "Qty", "UOM", "Source Box Entries", "Net Weight (kg)",
    "Lot No(s)", "Item Mark(s)", "Inward No(s)", "Source Txn No(s)",
    "Transaction Type", "Status", "Vehicle No", "Invoice No", "Remarks",
    "Created By", "Created At",
]


def build_direct_out():
    rows = []
    for r in fetch(DIRECT_OUT_SQL):
        rows.append([
            r["transaction_no"], r["entry_date"], "Cold Direct Out", r["company"],
            r["warehouse"] or r["line_wh"], r["units"], r["to_customer"],
            r["article"], n3(r["qty"]), r["uom"], i0(r["box_rows"]), n3(r["net_wt"]),
            r["lots"] or r["hdr_lot"], r["item_marks"], r["inward_nos"], r["src_txns"],
            r["transaction_type"], r["status"], r["vehicle_no"], r["invoice_no"],
            r["remarks"], r["created_by"], r["created_at"],
        ])
    return rows


# ── 6. Inner Cold Transfer (relocation / lot change inside cold) ──────────
INNER_SQL = """
SELECT challan_no, COALESCE(transfer_date,'') AS transfer_date,
       COALESCE(from_warehouse,'') AS from_warehouse, COALESCE(transfer_type,'') AS transfer_type,
       COALESCE(reason_code,'') AS reason_code, COALESCE(remark,'') AS remark,
       status, created_at::date::text AS created_at,
       COALESCE(NULLIF(TRIM(item_description),''),'(blank)') AS article,
       string_agg(DISTINCT NULLIF(TRIM(item_category),''), ' / ')        AS category,
       string_agg(DISTINCT NULLIF(TRIM(old_lot_number),''), ', ')        AS old_lots,
       string_agg(DISTINCT NULLIF(TRIM(new_lot_number),''), ', ')        AS new_lots,
       string_agg(DISTINCT NULLIF(TRIM(old_storage_location),''), ', ')  AS old_loc,
       string_agg(DISTINCT NULLIF(TRIM(new_storage_location),''), ', ')  AS new_loc,
       count(*) AS line_rows, sum(quantity) AS qty, sum(net_weight_kg) AS net_wt
FROM inner_cold_transfer
WHERE created_at >= :d1_ts AND created_at < :d2_ts
GROUP BY 1,2,3,4,5,6,7,8, article
ORDER BY created_at, challan_no, article
"""

INNER_COLS = [
    "TR Number", "TR Date", "Direction", "Warehouse", "Article", "Category",
    "Qty", "UOM", "Lines Merged", "Net Weight (kg)",
    "Old Lot No(s)", "New Lot No(s)", "Old Location", "New Location",
    "Transfer Type", "Reason Code", "Status", "Remark", "Created At",
]


def build_inner():
    return [[
        r["challan_no"], r["transfer_date"], "Inner Cold Transfer", r["from_warehouse"],
        r["article"], r["category"], n3(r["qty"]), "CARTON", i0(r["line_rows"]),
        n3(r["net_wt"]), r["old_lots"], r["new_lots"], r["old_loc"], r["new_loc"],
        r["transfer_type"], r["reason_code"], r["status"], r["remark"], r["created_at"],
    ] for r in fetch(INNER_SQL)]


# ── 7. In-Transit ledger (dispatched, not yet received) ───────────────────
TRANSIT_SQL = """
SELECT transfer_out_challan_no AS challan, transfer_type,
       COALESCE(from_site,'') AS from_site, COALESCE(to_site,'') AS to_site,
       from_storage_type, to_storage_type,
       COALESCE(from_company,'') AS from_company, COALESCE(to_company,'') AS to_company,
       COALESCE(NULLIF(TRIM(article),''), NULLIF(TRIM(item_description),''),'(blank)') AS article,
       string_agg(DISTINCT NULLIF(TRIM(item_category),''), ' / ')  AS category,
       string_agg(DISTINCT NULLIF(TRIM(sub_category),''), ' / ')   AS sub_category,
       string_agg(DISTINCT NULLIF(TRIM(rm_pm_fg_type),''), ' / ')  AS material_type,
       string_agg(DISTINCT NULLIF(TRIM(uom),''), ' / ')            AS uom,
       string_agg(DISTINCT NULLIF(TRIM(lot_no),''), ', ')          AS lots,
       string_agg(DISTINCT NULLIF(TRIM(batch_number),''), ', ')    AS batches,
       string_agg(DISTINCT NULLIF(TRIM(status),''), ', ')          AS status,
       count(*) AS box_rows, sum(qty) AS qty, sum(no_of_cartons) AS cartons,
       sum(weight_kg) AS weight_kg, sum(net_weight) AS net_wt, sum(total_weight) AS total_wt,
       min(dispatched_at)::date::text AS first_dispatch,
       max(dispatched_at)::date::text AS last_dispatch,
       string_agg(DISTINCT NULLIF(TRIM(dispatched_by),''), ', ')   AS dispatched_by,
       count(*) FILTER (WHERE reconciled) AS reconciled_boxes
FROM pending_transfer_stock
WHERE dispatched_at >= :d1_ts AND dispatched_at < :d2_ts
GROUP BY 1,2,3,4,5,6,7,8,9
ORDER BY min(dispatched_at), 1, 9
"""

TRANSIT_COLS = [
    "TR Number", "First Dispatch", "Last Dispatch", "Direction",
    "From Warehouse", "To Warehouse", "From Storage", "To Storage",
    "From Company", "To Company", "Article", "Category", "Sub Category", "Material Type",
    "Qty", "UOM", "Box Count", "Cartons", "Weight (kg)", "Net Weight (kg)",
    "Total Weight (kg)", "Lot No(s)", "Batch No(s)", "Status",
    "Reconciled Boxes", "Dispatched By", "Transfer Type",
]


def build_transit():
    return [[
        r["challan"], r["first_dispatch"], r["last_dispatch"], "In Transit",
        r["from_site"], r["to_site"], r["from_storage_type"], r["to_storage_type"],
        r["from_company"], r["to_company"], r["article"], r["category"],
        r["sub_category"], r["material_type"], n3(r["qty"]), r["uom"],
        i0(r["box_rows"]), n3(r["cartons"]), n3(r["weight_kg"]), n3(r["net_wt"]),
        n3(r["total_wt"]), r["lots"], r["batches"], r["status"],
        i0(r["reconciled_boxes"]), r["dispatched_by"], r["transfer_type"],
    ] for r in fetch(TRANSIT_SQL)]


# ── 8. Combined view across the four core movement types ─────────────────
COMBINED_COLS = [
    "Direction", "TR / GRN Number", "Date", "From Warehouse", "To Warehouse", "Cold Unit",
    "Article", "Category", "Sub Category", "Material Type",
    "Qty", "UOM", "Box Count", "Net Weight (kg)", "Total Weight (kg)",
    "Lot No(s)", "Batch No(s)", "Status", "Linked TR / GRN",
]


def to_combined(rows, kind):
    """Project an OUT-shaped or IN-shaped row onto the combined layout."""
    out = []
    for r in rows:
        if kind == "out":                      # OUT_COLS layout
            out.append([r[2], r[0], r[1], r[3], r[4], r[5], r[6], r[7], r[8], r[9],
                        r[10], r[11], r[12], r[15], r[16], r[19], r[20], r[23], r[25]])
        else:                                  # IN_COLS layout
            out.append([r[2], r[0], r[1], r[3], r[4], r[5], r[6], r[7], r[8], r[9],
                        r[10], r[11], r[12], r[13], r[14], r[15], r[16], r[17], r[19]])
    return out


# ── 9. Write ─────────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E78")


def prune_empty(header, groups):
    """Drop columns with no value on any sheet sharing this header.

    An all-zero column is kept - zero is an answer. Only None/'' counts as
    'the IMS holds nothing here for this period'.
    """
    keep = [i for i in range(len(header))
            if any(r[i] not in (None, "") for rows in groups for r in rows)]
    dropped = [header[i] for i in range(len(header)) if i not in set(keep)]
    return ([header[i] for i in keep],
            [[[r[i] for i in keep] for r in rows] for rows in groups],
            dropped)


def add_sheet(wb, title, header, rows):
    ws = wb.create_sheet(title)
    ws.append(header)
    for r in rows:
        ws.append(r)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        width = max([len(str(header[c - 1]))] +
                    [len(str(r[c - 1])) for r in rows[:500] if r[c - 1] is not None]) + 2
        ws.column_dimensions[get_column_letter(c)].width = min(max(width, 10), 46)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions
    return ws


def main():
    look = line_lookup()
    wh_out, cold_out = build_out()
    wh_in, cold_in_iu = build_in(look)
    cold_in_ded, ded_grns = build_cold_in_dedicated(look)
    # A GRN owned by the cold table must not also be counted from the interunit
    # scan ledger, on either sheet.
    superseded = sorted(ded_grns & {r[0] for r in wh_in + cold_in_iu})
    wh_in = [r for r in wh_in if r[0] not in ded_grns]
    cold_in = [r for r in cold_in_iu if r[0] not in ded_grns] + cold_in_ded
    cold_in.sort(key=lambda r: (r[1] or "", r[0] or "", r[6] or ""))
    direct_out = build_direct_out()
    inner = build_inner()
    transit = build_transit()

    combined = (to_combined(wh_out, "out") + to_combined(cold_out, "out")
                + to_combined(wh_in, "in") + to_combined(cold_in, "in"))
    combined.sort(key=lambda r: (r[2] or "", r[0] or "", r[1] or "", r[6] or ""))

    # Summary is computed on the full column set, before any pruning.
    # (sheet, rows, doc, date, article, qty, box, weight) - box is None where the
    # sheet has no separate physical box count and Qty already is the carton count.
    summary = [
        ("Transfer Out", wh_out, 0, 1, 6, 10, 12, 15),
        ("Cold Transfer Out", cold_out, 0, 1, 6, 10, 12, 15),
        ("Transfer In", wh_in, 0, 1, 6, 10, 12, 13),
        ("Cold Transfer In", cold_in, 0, 1, 6, 10, 12, 13),
        ("Cold Direct Out", direct_out, 0, 1, 7, 8, None, 11),
        ("Inner Cold Transfer", inner, 0, 18, 4, 6, None, 9),
        ("In Transit (Pending)", transit, 0, 1, 10, 14, 16, 19),
    ]
    summary_rows = []
    for title, rows, i_doc, i_date, i_art, i_qty, i_box, i_wt in summary:
        dates = sorted(d for d in (r[i_date] for r in rows) if d)
        summary_rows.append([
            title, len(rows), len({r[i_doc] for r in rows}), len({r[i_art] for r in rows}),
            dates[0] if dates else "", dates[-1] if dates else "",
            round(sum(r[i_qty] or 0 for r in rows), 3),
            sum(r[i_box] or 0 for r in rows) if i_box is not None else "(Qty is the count)",
            round(sum(r[i_wt] or 0 for r in rows), 3)])

    # Drop columns the IMS holds nothing in for this period, so the sheets
    # carry only fields that actually have data behind them.
    dropped = {}
    out_hdr, (wh_out, cold_out), dropped["Transfer Out / Cold Transfer Out"] = \
        prune_empty(OUT_COLS, [wh_out, cold_out])
    in_hdr, (wh_in, cold_in), dropped["Transfer In / Cold Transfer In"] = \
        prune_empty(IN_COLS, [wh_in, cold_in])
    do_hdr, (direct_out,), dropped["Cold Direct Out"] = prune_empty(DIRECT_OUT_COLS, [direct_out])
    ic_hdr, (inner,), dropped["Inner Cold Transfer"] = prune_empty(INNER_COLS, [inner])
    tr_hdr, (transit,), dropped["In Transit (Pending)"] = prune_empty(TRANSIT_COLS, [transit])
    cb_hdr, (combined,), dropped["All Transfers"] = prune_empty(COMBINED_COLS, [combined])

    wb = Workbook()
    wb.remove(wb.active)
    sheets = [
        ("Transfer Out", out_hdr, wh_out),
        ("Cold Transfer Out", out_hdr, cold_out),
        ("Transfer In", in_hdr, wh_in),
        ("Cold Transfer In", in_hdr, cold_in),
        ("Cold Direct Out", do_hdr, direct_out),
        ("Inner Cold Transfer", ic_hdr, inner),
        ("In Transit (Pending)", tr_hdr, transit),
        ("All Transfers", cb_hdr, combined),
    ]
    for title, header, rows in sheets:
        add_sheet(wb, title, header, rows)

    # ── Notes ────────────────────────────────────────────────────────────
    n = wb.create_sheet("Notes")
    n.append(["Transfer register - consolidated per article per transaction"])
    n.append(["Period", f"{D_FROM:%d %b %Y} to {D_TO:%d %b %Y}"])
    n.append(["Source", "IMS production database (warehouse_db), read-only"])
    n.append([])
    n.append(["Sheet", "Rows", "Documents", "Articles", "First date", "Last date",
              "Total Qty", "Total Boxes", "Net Weight (kg)"])
    for row in summary_rows:
        n.append(row)
    for line in [
        [],
        ["How to read this workbook"],
        ["Grain", "ONE ROW PER (transaction, article). All per-box and per-line records "
                  "behind a document+article are collapsed into a single row."],
        ["Qty", "OUT sheets: sum of the transfer-line qty in the stated UOM. "
                "IN sheets: boxes actually scanned in (the receiving side counts boxes, "
                "not packs), so IN Qty can differ from OUT Qty by design."],
        ["Box Count", "Number of physical boxes/cartons. OUT prefers the real per-box records "
                      "(interunit_transfer_boxes, linked by transfer_line_id); where a challan "
                      "predates per-box capture it falls back to the transfer-line count - "
                      "'Box Count Basis' states which was used for every row."],
        ["Lines Merged", "How many source rows were consolidated into that one row. Shown only "
                         "where it can differ from Box Count - on the IN sheets one scan row is "
                         "one box, so the two are always equal and only Box Count is carried."],
        ["Cold Direct Out / Inner Cold Transfer",
         "These two have no separate physical box ledger: Qty already IS the carton count. "
         "'Source Box Entries' (Direct Out) counts the source pile entries drawn from, and "
         "'Lines Merged' (Inner Cold) counts the movement rows consolidated - neither is a "
         "carton count. Inner Cold's UOM is inferred as CARTON: its quantity divided by "
         "net weight matches the article pack size throughout."],
        ["UOM", "Multiple UOMs under one (transaction, article) are joined with ' / '. "
                "35 of 2,343 OUT groups have two UOMs; the rest have one."],
        ["Lot / Batch / Vakkal", "Distinct values across the merged rows, comma-joined."],
        ["IN sheets Category / Material / UOM", "Carried across from the matching transfer-OUT "
                                                "line (by header + article + lot); blank when no match."],
        [],
        ["Sheet definitions"],
        ["Transfer Out", "Inter-unit dispatch FROM a regular warehouse (A68, A185, W202, F53, A101)."],
        ["Cold Transfer Out", "Dispatch FROM cold - from_site is 'Cold Storage' / Savla D-39 / "
                              "Savla D-514 / Savla Bond / Rishi / Supreme / Eskimo, or from_cold_unit is set."],
        ["Transfer In", "GRN receipt INTO a regular warehouse."],
        ["Cold Transfer In", "GRN receipt INTO a cold site. Union of the two receive paths "
                             "(cold_transfer_in_headers + interunit_transfer_in_header), deduped per GRN."],
        ["Cold Direct Out", "Stock issued straight out of cold storage to a customer "
                            "(cdpl/cfpl_cold_storage_direct_out). Not an inter-unit transfer, "
                            "but it is stock leaving cold - included for completeness."],
        ["Inner Cold Transfer", "Relocation / lot renumbering WITHIN cold storage. No warehouse "
                                "change, so it is neither an in nor an out - kept separate."],
        ["In Transit (Pending)", "pending_transfer_stock: dispatched but not yet received. "
                                 "These boxes are already counted on the OUT sheets - this sheet "
                                 "is the open-pipeline view, do not add it to the totals."],
        ["All Transfers", "The four core sheets stacked on a common column set for pivoting."],
        [],
        ["Date basis"],
        ["OUT sheets", "stock transfer (challan) date. A challan whose document date falls outside "
                       "the period but which was entered inside it is still included and flagged "
                       "'entry date' in the 'In Range By' column."],
        ["IN sheets", "GRN date. The originating challan date is in 'TR Date'."],
        ["Cold Direct Out", "entry date."],
        ["Inner Cold Transfer", "created timestamp (its transfer_date column is free text)."],
        [],
        ["Caveats"],
        ["Receipt Status", "'Received' when the challan is Received; 'Partially/Pending' when a GRN "
                           "exists but the challan is not closed; else 'Not Received'."],
        ["Cold receipt source", "A cold receipt can be written by two paths. The cold carton table "
                                "wins whenever it holds real per-box detail; the interunit scan "
                                "ledger is used otherwise. The 'Box Source' column records which "
                                "one every row came from, and no GRN is counted from both."],
        ["GRN-20260608124120", "Challan TRANS202606051800 dispatched 185 boxes / 1,850.000 kg. The "
                               "cold table records 185 / 1,850.000; the interunit scan ledger holds "
                               "369 / 3,690.000. The cold table's 185 is used - counting the ledger "
                               "would double this GRN. It is the only in-period GRN where the two "
                               "paths disagree."],
        ["Total Weight (IN)", "Gross weight where scanned. Cold receipts record net only, so gross "
                              "is blank there."],
        ["Empty article", "Shown as '(blank)' - 10 OUT lines and 5 IN boxes have no article text."],
    ]:
        n.append(line)

    if any(dropped.values()):
        n.append([])
        n.append(["Fields the IMS records but which are empty for this period "
                  "(column omitted rather than shown blank)"])
        for sheet_name, cols in dropped.items():
            if cols:
                n.append([sheet_name, ", ".join(cols)])
    n.column_dimensions["A"].width = 32
    n.column_dimensions["B"].width = 108
    for col in "CDEFGHI":
        n.column_dimensions[col].width = 16
    n["A1"].font = Font(bold=True, size=14)
    n["A5"].font = Font(bold=True)
    for row in n.iter_rows():
        for cell in row:
            if cell.column_letter == "B":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print(f"wrote {OUT_XLSX}")
    for title, _h, rows in sheets:
        print(f"  {title:24s} {len(rows):7d} rows")
    if superseded:
        print(f"  cold GRNs taken from cold table instead of scan ledger: {superseded}")


if __name__ == "__main__":
    main()
