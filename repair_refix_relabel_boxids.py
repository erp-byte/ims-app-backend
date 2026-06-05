"""Re-fix the relabeled boxes from the '<orig>-R<tid>' format to the inward-style
'{last 8 digits of epoch ms}-{n}' format. Updates BOTH interunit_transfer_boxes
and the tracking pending_transfer_stock rows, and refreshes the Excel sheet.

Identifies the relabel rows = tracking pending rows (source_table='' AND
destination_table='') with a real (non-LINE) box_id.

DRY-RUN by default; --apply commits.
"""
import os, io, sys, time
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
from sqlalchemy import create_engine, text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

raw = os.environ["DATABASE_URL"]
DB = raw.replace("postgresql://", "postgresql+psycopg://", 1) if raw.startswith("postgresql://") else raw
APPLY = "--apply" in sys.argv
eng = create_engine(DB)
XLSX = "Interunit_Transfer_Reconciliation.xlsx"
SHEET = "Relabeled Active Double-Use"


def _exists(c, box_id):
    for tbl in ("interunit_transfer_boxes", "pending_transfer_stock"):
        if c.execute(text(f"SELECT 1 FROM {tbl} WHERE box_id=:b LIMIT 1"), {"b": box_id}).fetchone():
            return True
    return False


FIND = text("""
    SELECT p.id AS pending_id, p.box_id, p.transaction_no, p.lot_no, p.item_description,
           p.transfer_out_id, p.transfer_out_challan_no, p.net_weight
    FROM pending_transfer_stock p
    WHERE p.status='In Transit'
      AND COALESCE(p.source_table,'')='' AND COALESCE(p.destination_table,'')=''
      AND p.box_id NOT LIKE 'LINE-%'
    ORDER BY p.transfer_out_id, p.box_id
""")

out_rows = []
with eng.begin() as c:
    rows = c.execute(FIND).fetchall()
    print(f"Relabel rows to re-fix: {len(rows)}  APPLY={APPLY}\n")
    base = str(int(time.time() * 1000))[-8:]
    seq = 0
    for r in rows:
        m = r._mapping
        old_id = m["box_id"]
        seq += 1
        new_id = f"{base}-{seq}"
        while _exists(c, new_id):
            seq += 1
            new_id = f"{base}-{seq}"
        out_rows.append([
            m["transfer_out_id"], m["transfer_out_challan_no"], m["item_description"],
            m["lot_no"] or "", m["transaction_no"], old_id, new_id,
            float(m["net_weight"] or 0), "DONE" if APPLY else "PLANNED",
        ])
        print(f"  transfer {m['transfer_out_id']}: {old_id}  ->  {new_id}")
        if APPLY:
            c.execute(text("UPDATE pending_transfer_stock SET box_id=:n WHERE id=:pid"),
                      {"n": new_id, "pid": m["pending_id"]})
            c.execute(text("""UPDATE interunit_transfer_boxes SET box_id=:n
                              WHERE box_id=:o AND transaction_no=:t"""),
                      {"n": new_id, "o": old_id, "t": m["transaction_no"]})
    print(f"\nTOTAL re-fixed: {len(out_rows)}")
    if not APPLY:
        print("DRY-RUN only — no DB writes.")

# refresh Excel sheet
wb = openpyxl.load_workbook(XLSX)
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET)
headers = ["Transfer ID", "Challan No", "Article", "Lot No", "Transaction No",
           "Old Box ID", "New Box ID (epoch-format)", "Net Wt (kg)", "Status"]
ws.append(headers)
for col in range(1, len(headers)+1):
    cc = ws.cell(row=1, column=col); cc.fill = PatternFill("solid", fgColor="1F4E78")
    cc.font = Font(bold=True, color="FFFFFF"); cc.alignment = Alignment(horizontal="center", wrap_text=True)
for row in out_rows:
    ws.append(row)
ws.freeze_panes = "A2"
for i, w in enumerate([12, 22, 30, 10, 20, 20, 24, 11, 10], start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
try:
    wb.save(XLSX); saved = XLSX
except PermissionError:
    saved = "Interunit_Transfer_Reconciliation_RELABEL.xlsx"
    wb.save(saved); saved += "  (main file OPEN — close & re-run to merge)"
print(f"Excel sheet '{SHEET}' -> {saved} ({len(out_rows)} rows).")
print("Done.")
